"""
Create Locations Input - Creates Locations_input.xlsx file from Input.xlsx using location mappings

This script reads Mappings.xlsx to find location-related mappings, then reads Input.xlsx
and creates a new file Locations_input.xlsx with only the location-related columns.
"""

import os
import sys
import re
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from difflib import SequenceMatcher
import requests

# Get the project root directory (parent of Logics folder)
LOGICS_DIR = Path(__file__).parent
PROJECT_ROOT = LOGICS_DIR.parent
BACKEND_DIR = PROJECT_ROOT / 'backend'
EXCEL_FILES_DIR = BACKEND_DIR / 'Excel Files'

# Location-related column types
LOCATION_TYPES = [
    'addressLine1',
    'addressLine2',
    'city',
    'state',
    'zip',
    'practiceId',
    'practiceCloudId',
    'practiceName',
    'locationId',
    'locationName',
    'locationTypeRaw'
]

# Map detected_as to readable names for column headers
readable_names = {
    'addressLine1': 'Address Line 1',
    'addressLine2': 'Address Line 2',
    'city': 'City',
    'state': 'State',
    'zip': 'ZIP',
    'practiceId': 'Practice ID',
    'practiceCloudId': 'Practice Cloud ID',
    'practiceName': 'Practice Name',
    'locationId': 'Location ID',
    'locationName': 'Location Name',
    'locationTypeRaw': 'Location Type_Raw'
}

def normalize_string_for_matching(text):
    """Normalize string for fuzzy matching - lowercase, remove extra spaces, remove special chars"""
    if pd.isna(text) or text == '':
        return ''
    text_str = str(text).lower().strip()
    # Remove special characters except spaces
    text_str = re.sub(r'[^a-z0-9\s]', '', text_str)
    # Remove extra spaces
    text_str = re.sub(r'\s+', ' ', text_str)
    return text_str.strip()

def fuzzy_match_score(str1, str2, threshold=0.7):
    """Calculate fuzzy match score between two strings"""
    if not str1 or not str2:
        return 0.0
    
    str1_norm = normalize_string_for_matching(str1)
    str2_norm = normalize_string_for_matching(str2)
    
    if not str1_norm or not str2_norm:
        return 0.0
    
    # Exact match
    if str1_norm == str2_norm:
        return 1.0
    
    # Check if one is a substring of the other
    if str1_norm in str2_norm or str2_norm in str1_norm:
        return 0.9
    
    # Use SequenceMatcher for similarity
    similarity = SequenceMatcher(None, str1_norm, str2_norm).ratio()
    return similarity

def determine_location_type(location_type_raw_value):
    """
    Determine Location Type from Location Type_Raw value using fuzzy matching.
    
    Returns: 'Both', 'Virtual', or 'In Person'
    """
    if pd.isna(location_type_raw_value) or location_type_raw_value == '':
        return 'In Person'  # Default
    
    location_type_raw_str = str(location_type_raw_value).strip()
    
    # Normalize the input
    normalized_input = normalize_string_for_matching(location_type_raw_str)
    
    # Debug logging for Virtual Visit and similar values
    if 'virtual' in normalized_input or 'visit' in normalized_input:
        print(f"DEBUG: Processing Location Type_Raw value: '{location_type_raw_str}' -> normalized: '{normalized_input}'")
    
    # First, check for explicit "In Person" patterns (check this FIRST to avoid false matches)
    # This must be checked before "Both" to prevent "In Person" from being incorrectly matched to "Both"
    # Check for exact or near-exact matches for "In Person" first
    if normalized_input in ['in person', 'in office', 'inperson', 'inoffice']:
        return 'In Person'
    
    # Check if input contains "in person" or "in office" as a phrase (but not combined with virtual keywords)
    if 'in person' in normalized_input or 'in office' in normalized_input:
        # Make sure it doesn't also contain virtual keywords (which would make it "Both")
        virtual_keywords_check = ['vv', 'virtual', 'telehealth', 'video', 'online']
        has_virtual_in_input = any(keyword in normalized_input for keyword in virtual_keywords_check)
        if not has_virtual_in_input:
            return 'In Person'
    
    # Virtual patterns - CHECK THIS BEFORE "Both" to ensure pure virtual values are correctly identified
    virtual_patterns = ['vv', 'virtual visit', 'virtual', 'telehealth', 'video visit', 'video', 'online']
    
    # Check for "Virtual" patterns FIRST (before checking for "Both")
    # This ensures that values like "Virtual", "Virtual Visit", etc. are correctly identified as Virtual
    for pattern in virtual_patterns:
        pattern_norm = normalize_string_for_matching(pattern)
        # Check if pattern is in the input or vice versa (exact or substring match)
        if pattern_norm == normalized_input:
            if 'virtual' in normalized_input or 'visit' in normalized_input:
                print(f"DEBUG: Matched Virtual pattern (exact): '{pattern}' -> '{pattern_norm}'")
            return 'Virtual'  # Exact match
        elif pattern_norm in normalized_input or normalized_input in pattern_norm:
            if 'virtual' in normalized_input or 'visit' in normalized_input:
                print(f"DEBUG: Matched Virtual pattern (substring): '{pattern}' -> '{pattern_norm}' in '{normalized_input}'")
            return 'Virtual'  # Substring match
        # Check fuzzy match
        fuzzy_score = fuzzy_match_score(normalized_input, pattern_norm, threshold=0.7)
        if fuzzy_score >= 0.7:
            if 'virtual' in normalized_input or 'visit' in normalized_input:
                print(f"DEBUG: Matched Virtual pattern (fuzzy, score={fuzzy_score:.2f}): '{pattern}' -> '{pattern_norm}' vs '{normalized_input}'")
            return 'Virtual'
    
    # Both patterns - check for combinations that indicate both in-person and virtual
    # IMPORTANT: Only match "Both" if BOTH in-person AND virtual keywords are present
    # This must come AFTER Virtual check to avoid false positives
    in_person_keywords_for_both = ['in person', 'in office', 'inperson', 'inoffice']
    virtual_keywords = ['vv', 'virtual', 'telehealth', 'video', 'online']
    
    # Check for explicit in-person phrases (not just "person" or "office" which could be in other words)
    has_in_person_phrase = any(phrase in normalized_input for phrase in ['in person', 'in office', 'inperson', 'inoffice'])
    has_virtual_keyword = any(keyword in normalized_input for keyword in virtual_keywords)
    
    # If both in-person phrase AND virtual keyword are present, it's "Both"
    if has_in_person_phrase and has_virtual_keyword:
        print(f"DEBUG: Matched Both pattern: has_in_person_phrase={has_in_person_phrase}, has_virtual_keyword={has_virtual_keyword}")
        return 'Both'
    
    # Also check explicit "both" patterns
    both_patterns = ['in person/vv', 'in office and video', 'both', 'in person and virtual', 'in office and virtual']
    for pattern in both_patterns:
        pattern_norm = normalize_string_for_matching(pattern)
        # Check if pattern is in the input or vice versa
        if pattern_norm in normalized_input or normalized_input in pattern_norm:
            print(f"DEBUG: Matched Both pattern (explicit): '{pattern}' -> '{pattern_norm}'")
            return 'Both'
        # Check fuzzy match (but only if it's not just "in person")
        if pattern_norm != 'in person' and fuzzy_match_score(normalized_input, pattern_norm, threshold=0.7) >= 0.7:
            print(f"DEBUG: Matched Both pattern (fuzzy): '{pattern}' -> '{pattern_norm}'")
            return 'Both'
    
    # Check if it contains "in person" or "in office" keywords (fallback)
    for keyword in in_person_keywords_for_both:
        keyword_norm = normalize_string_for_matching(keyword)
        if keyword_norm in normalized_input:
            return 'In Person'
    
    # Default to 'In Person' if no match found
    return 'In Person'

def split_address_line1(address_value):
    """
    Split Address Line 1 into Address Line 1 and Address Line 2.
    
    Examples:
    - '165 Broadway, 23rd floor, New York, NY, 10006' -> ('165 Broadway', '23rd floor')
    - '165 Broadway 23rd floor' -> ('165 Broadway', '23rd floor')
    - '175 E Hawthorn Pkwy' -> ('175 E Hawthorn Pkwy', '')
    - '4 CENTURY DR STE 100' -> ('4 Century Dr', 'Ste 100')
    
    Args:
        address_value: The address string to split
        
    Returns:
        Tuple of (address_line1, address_line2)
    """
    if pd.isna(address_value) or address_value == '':
        return ('', '')
    
    address_str = str(address_value).strip()
    if not address_str:
        return ('', '')
    
    # Pattern 1: Comma-separated addresses
    # '165 Broadway, 23rd floor, New York, NY, 10006' -> split by comma
    if ',' in address_str:
        parts = [p.strip() for p in address_str.split(',')]
        # Take first part as Address Line 1
        address_line1 = parts[0]
        # Take second part as Address Line 2 if it exists and doesn't look like city/state/zip
        # Check if second part looks like a city/state/zip (usually 2-3 words, might have state abbreviation)
        if len(parts) > 1:
            second_part = parts[1]
            # If second part is short and might be a floor/suite, use it as Address Line 2
            # Otherwise, if there are more parts, the second might be Address Line 2
            if len(parts) >= 3:
                # Multiple commas - second part is likely Address Line 2
                address_line2 = second_part
            elif len(second_part.split()) <= 3 and any(keyword in second_part.upper() for keyword in ['FLOOR', 'FL', 'STE', 'SUITE', 'APT', 'APARTMENT', 'UNIT', 'BLDG', 'BUILDING']):
                # Second part contains address line 2 keywords
                address_line2 = second_part
            elif len(second_part.split()) <= 2:
                # Short second part - likely Address Line 2
                address_line2 = second_part
            else:
                # Long second part - might be city, keep as Address Line 1 only
                address_line2 = ''
        else:
            address_line2 = ''
        
        # Capitalize properly
        address_line1 = ' '.join(word.capitalize() for word in address_line1.split())
        address_line2 = ' '.join(word.capitalize() for word in address_line2.split()) if address_line2 else ''
        
        return (address_line1, address_line2)
    
    # Pattern 2: Look for common address line 2 keywords (STE, SUITE, FLOOR, FL, APT, etc.)
    # Case-insensitive search for keywords
    address_upper = address_str.upper()
    
    # Keywords that typically indicate Address Line 2 - ordered by specificity
    # More specific patterns first (like "23rd floor") then general ones
    address_line2_keywords = [
        r'\d+\w*\s*(ST|ND|RD|TH)\s+FLOOR',  # e.g., "23rd floor", "2nd floor", "1st floor"
        r'\d+\w*\s*(ST|ND|RD|TH)\s+FL\.?',  # e.g., "23rd fl", "2nd fl."
        r'\bSTE\s+\d+',                      # e.g., "STE 100", "STE 200"
        r'\bSUITE\s+\d+',                    # e.g., "SUITE 100"
        r'\bAPT\s+\d+',                      # e.g., "APT 5"
        r'\bAPARTMENT\s+\d+',                 # e.g., "APARTMENT 3"
        r'\bUNIT\s+\d+',                     # e.g., "UNIT 10"
        r'\bSTE\b',                          # e.g., "STE" (standalone)
        r'\bSUITE\b',                        # e.g., "SUITE" (standalone)
        r'\bFLOOR\b',                        # e.g., "FLOOR" (standalone)
        r'\bFL\b',                           # e.g., "FL" (standalone)
        r'\bFL\.\b',                         # e.g., "FL." (standalone)
        r'\bAPT\b',                          # e.g., "APT" (standalone)
        r'\bAPARTMENT\b',                    # e.g., "APARTMENT" (standalone)
        r'\bAPT\.\b',                        # e.g., "APT." (standalone)
        r'\bUNIT\b',                         # e.g., "UNIT" (standalone)
        r'\bBLDG\b',                         # e.g., "BLDG"
        r'\bBUILDING\b',                     # e.g., "BUILDING"
    ]
    
    split_position = None
    split_keyword = None
    
    for keyword_pattern in address_line2_keywords:
        match = re.search(keyword_pattern, address_upper, re.IGNORECASE)
        if match:
            # Found a keyword - split at this position
            split_position = match.start()
            split_keyword = match.group()
            break
    
    if split_position is not None and split_position > 0:
        # Split at the keyword (but make sure we're not splitting at the very beginning)
        address_line1 = address_str[:split_position].strip()
        address_line2 = address_str[split_position:].strip()
        
        # Capitalize properly
        address_line1 = ' '.join(word.capitalize() for word in address_line1.split())
        address_line2 = ' '.join(word.capitalize() for word in address_line2.split())
        
        return (address_line1, address_line2)
    
    # Pattern 3: Look for patterns like "number word word" followed by more words
    # This is a fallback - if we can't find keywords, try to split intelligently
    words = address_str.split()
    if len(words) >= 4:
        # Try to find a natural split point (after street number + street name)
        # Usually: "number direction street type" then rest
        # For example: "4 CENTURY DR STE 100" -> "4 Century Dr" and "Ste 100"
        # Or: "165 Broadway 23rd floor" -> "165 Broadway" and "23rd floor"
        
        # Common street types
        street_types = ['ST', 'STREET', 'AVE', 'AVENUE', 'RD', 'ROAD', 'DR', 'DRIVE', 'BLVD', 'BOULEVARD', 
                       'LN', 'LANE', 'CT', 'COURT', 'CIR', 'CIRCLE', 'PKWY', 'PARKWAY', 'PL', 'PLACE']
        
        # Look for street type in first few words
        for i in range(min(4, len(words))):
            if words[i].upper().rstrip('.,') in street_types:
                # Found street type - split after this word
                address_line1 = ' '.join(words[:i+1])
                address_line2 = ' '.join(words[i+1:])
                
                # Capitalize properly
                address_line1 = ' '.join(word.capitalize() for word in address_line1.split())
                address_line2 = ' '.join(word.capitalize() for word in address_line2.split())
                
                return (address_line1, address_line2)
    
    # Pattern 4: No split found - keep entire address as Address Line 1
    address_line1 = ' '.join(word.capitalize() for word in address_str.split())
    return (address_line1, '')

def create_locations_input_excel(manual_practice_ids=''):
    """
    Create Locations_input.xlsx file from Input.xlsx using location mappings from Mappings.xlsx
    
    Args:
        manual_practice_ids: Optional string of manually entered Practice IDs (comma, semicolon, or space separated)
    
    Returns:
        Path to the created Locations_input.xlsx file
    """
    # Ensure Excel Files directory exists
    EXCEL_FILES_DIR.mkdir(parents=True, exist_ok=True)
    
    # File paths
    mappings_file = EXCEL_FILES_DIR / 'Mappings.xlsx'
    input_file = EXCEL_FILES_DIR / 'Input.xlsx'
    output_file = EXCEL_FILES_DIR / 'Locations_input.xlsx'
    
    # Check if required files exist
    if not mappings_file.exists():
        raise FileNotFoundError(f"Mappings.xlsx not found at {mappings_file}")
    
    if not input_file.exists():
        raise FileNotFoundError(f"Input.xlsx not found at {input_file}")
    
    # Read Mappings.xlsx
    mappings_df = pd.read_excel(mappings_file, sheet_name='Mappings')
    
    # Check required columns
    if 'Detected As' not in mappings_df.columns or 'Column Name' not in mappings_df.columns:
        raise ValueError("Mappings.xlsx must contain 'Detected As' and 'Column Name' columns")
    
    # Find location-related mappings
    location_mappings = {}
    location_id_counter = 0  # Counter for multiple Location ID columns
    
    # Create reverse mapping: readable_name -> internal_name
    readable_to_internal = {v: k for k, v in readable_names.items()}
    
    for _, row in mappings_df.iterrows():
        detected_as = str(row['Detected As']).strip()
        column_name = str(row['Column Name']).strip()
        
        # Handle multiple mappings (comma-separated in Detected As)
        detected_as_list = [d.strip() for d in detected_as.split(',')]
        
        for detected in detected_as_list:
            # Normalize detected_as to match our location types
            detected_normalized = detected.strip()
            
            # Normalize for comparison: replace underscores/spaces and lowercase
            detected_normalized_for_match = detected_normalized.replace('_', ' ').replace('-', ' ').lower().strip()
            
            # Check if this is a location-related mapping
            # First check if it matches a readable name directly
            if detected_normalized in readable_names.values():
                # This is already a readable name, use it directly
                # Special handling for Location ID - number them
                if detected_normalized == 'Location ID':
                    location_id_counter += 1
                    location_mappings[f'Location ID {location_id_counter}'] = column_name
                    print(f"Mapped '{column_name}' -> Location ID {location_id_counter}")
                else:
                    location_mappings[detected_normalized] = column_name
                    print(f"Mapped '{column_name}' -> {detected_normalized}")
            elif detected_normalized in readable_to_internal:
                # This is a readable name that maps to a location type
                internal_name = readable_to_internal[detected_normalized]
                if internal_name in LOCATION_TYPES:
                    # Special handling for Location ID - number them
                    if internal_name == 'locationId':
                        location_id_counter += 1
                        location_mappings[f'Location ID {location_id_counter}'] = column_name
                        print(f"Mapped '{column_name}' -> Location ID {location_id_counter}")
                    else:
                        location_mappings[readable_names[internal_name]] = column_name
                        print(f"Mapped '{column_name}' -> {readable_names[internal_name]}")
            else:
                # Check if it matches an internal name (normalized comparison)
                # Also handle variations with spaces/underscores/dashes
                matched = False
                for internal_name, readable_name in readable_names.items():
                    if internal_name in LOCATION_TYPES:
                        # Normalize readable name for comparison
                        readable_normalized = readable_name.replace('_', ' ').replace('-', ' ').lower().strip()
                        internal_normalized = internal_name.replace('_', ' ').replace('-', ' ').lower().strip()
                        
                        # Check exact matches (case-insensitive, space/underscore insensitive)
                        if (detected_normalized_for_match == internal_normalized or 
                            detected_normalized_for_match == readable_normalized):
                            # Special handling for Location ID - number them
                            if internal_name == 'locationId':
                                location_id_counter += 1
                                location_mappings[f'Location ID {location_id_counter}'] = column_name
                                print(f"Mapped '{column_name}' -> Location ID {location_id_counter} (matched '{detected_normalized}')")
                            else:
                                location_mappings[readable_name] = column_name
                                print(f"Mapped '{column_name}' -> {readable_name} (matched '{detected_normalized}')")
                            matched = True
                            break
                
                # If still not matched, try fuzzy matching on the normalized strings
                if not matched:
                    for internal_name, readable_name in readable_names.items():
                        if internal_name in LOCATION_TYPES:
                            readable_normalized = readable_name.replace('_', ' ').replace('-', ' ').lower().strip()
                            internal_normalized = internal_name.replace('_', ' ').replace('-', ' ').lower().strip()
                            
                            # Check if normalized strings are similar (handles minor variations)
                            if (detected_normalized_for_match in readable_normalized or 
                                readable_normalized in detected_normalized_for_match or
                                detected_normalized_for_match in internal_normalized or
                                internal_normalized in detected_normalized_for_match):
                                if internal_name == 'locationId':
                                    location_id_counter += 1
                                    location_mappings[f'Location ID {location_id_counter}'] = column_name
                                    print(f"Mapped '{column_name}' -> Location ID {location_id_counter} (fuzzy matched '{detected_normalized}')")
                                else:
                                    location_mappings[readable_name] = column_name
                                    print(f"Mapped '{column_name}' -> {readable_name} (fuzzy matched '{detected_normalized}')")
                                matched = True
                                break
    
    if not location_mappings:
        print("Warning: No location-related mappings found. Creating empty Locations_input.xlsx")
        # Create empty file with location column headers
        empty_data = {readable_names[lt]: [] for lt in LOCATION_TYPES}
        output_df = pd.DataFrame(empty_data)
        output_df.to_excel(output_file, index=False, sheet_name='Locations')
        print(f"Created empty Locations_input.xlsx at {output_file}")
        return str(output_file)
    
    # Read Input.xlsx
    input_df = pd.read_excel(input_file)
    
    # Read _Mapped.xlsx for First Name and Last Name columns
    mapped_file = EXCEL_FILES_DIR / '_Mapped.xlsx'
    mapped_df = None
    if mapped_file.exists():
        try:
            mapped_df = pd.read_excel(mapped_file, sheet_name='Mapped')
            print(f"Read _Mapped.xlsx for First Name and Last Name columns")
        except Exception as e:
            print(f"Warning: Could not read _Mapped.xlsx: {str(e)}. First Name and Last Name will not be copied.")
            mapped_df = None
    else:
        print(f"Warning: _Mapped.xlsx not found at {mapped_file}. First Name and Last Name will not be copied.")
    
    # Create new dataframe with location columns
    locations_data = {}
    
    # Track if Address Line 1 is mapped (we'll process it specially)
    address_line1_source = None
    address_line2_source = None
    
    for location_column_name, source_column in location_mappings.items():
        if location_column_name == 'Address Line 1':
            address_line1_source = source_column
        elif location_column_name == 'Address Line 2':
            address_line2_source = source_column
        else:
            # For other columns, just copy the data
            if source_column in input_df.columns:
                locations_data[location_column_name] = input_df[source_column]
                if location_column_name == 'Location Type_Raw':
                    print(f"Successfully copied Location Type_Raw from column '{source_column}' ({len(input_df)} rows)")
            else:
                # If source column doesn't exist, fill with empty values
                locations_data[location_column_name] = [''] * len(input_df)
                print(f"Warning: Source column '{source_column}' not found in Input.xlsx. Filling with empty values.")
    
    # Debug: Check if Location Type_Raw was mapped but not copied
    if 'Location Type_Raw' not in locations_data:
        if 'Location Type_Raw' in location_mappings:
            print(f"Error: Location Type_Raw mapping exists but was not copied! Mapping: '{location_mappings['Location Type_Raw']}'")
        else:
            # Check if it might have been mapped with a different key
            location_type_raw_variations = [k for k in location_mappings.keys() if 'location' in k.lower() and 'type' in k.lower() and 'raw' in k.lower()]
            if location_type_raw_variations:
                print(f"Warning: Found Location Type_Raw-like mappings but not copied: {location_type_raw_variations}")
    
    # Check if we should skip Address Line 1 splitting (only Address Line 1 mapped, no other address components)
    try:
        import sys
        from pathlib import Path
        sys.path.insert(0, str(Path(__file__).parent))
        from Addressline1 import should_skip_address_splitting, process_address_line1_only
    except ImportError:
        try:
            from .Addressline1 import should_skip_address_splitting, process_address_line1_only
        except Exception:
            # If import fails, default to splitting behavior
            should_skip_address_splitting = lambda x: False
            process_address_line1_only = None
    
    # Determine if we should skip splitting
    skip_splitting = should_skip_address_splitting(location_mappings)
    
    # Process Address Line 1 specially - split into Address Line 1 and Address Line 2
    # UNLESS only Address Line 1 is mapped (no Address Line 2, City, State, ZIP)
    if address_line1_source and address_line1_source in input_df.columns:
        if skip_splitting and process_address_line1_only:
            # Only Address Line 1 is mapped - keep it whole, don't split
            print("Only Address Line 1 is mapped. Keeping Address Line 1 whole without splitting.")
            address_line1_values, address_line2_values = process_address_line1_only(input_df, address_line1_source)
        else:
            # Normal processing - split Address Line 1
            address_line1_values = []
            address_line2_values = []
            
            for idx, address_value in input_df[address_line1_source].items():
                # Split the address
                addr1, addr2 = split_address_line1(address_value)
                address_line1_values.append(addr1)
                
                # If Address Line 2 is also mapped separately, use the mapped value if it exists
                # Otherwise, use the split value from Address Line 1
                if address_line2_source and address_line2_source in input_df.columns:
                    # Address Line 2 is mapped separately - use that value, but if empty, use split value
                    addr2_mapped = str(input_df.at[idx, address_line2_source]).strip() if pd.notna(input_df.at[idx, address_line2_source]) else ''
                    address_line2_values.append(addr2_mapped if addr2_mapped else addr2)
                else:
                    # Address Line 2 not mapped separately - use split value
                    address_line2_values.append(addr2)
        
        locations_data['Address Line 1'] = address_line1_values
        locations_data['Address Line 2'] = address_line2_values
    elif address_line1_source:
        # Address Line 1 is mapped but column doesn't exist
        print(f"Warning: Source column '{address_line1_source}' for Address Line 1 not found in Input.xlsx.")
        locations_data['Address Line 1'] = [''] * len(input_df)
        locations_data['Address Line 2'] = [''] * len(input_df)
    else:
        # Address Line 1 not mapped - initialize empty
        locations_data['Address Line 1'] = [''] * len(input_df)
        locations_data['Address Line 2'] = [''] * len(input_df)
    
    # Handle case where Address Line 2 is mapped separately but Address Line 1 is not
    if address_line2_source and address_line2_source in input_df.columns and not address_line1_source:
        if 'Address Line 2' not in locations_data:
            locations_data['Address Line 2'] = input_df[address_line2_source].astype(str).fillna('').tolist()
    
    # Add NPI Number, First Name, and Last Name columns if they're mapped
    # Note: First Name and Last Name will be read from _Mapped.xlsx, not Input.xlsx
    # PFS is NOT included in Locations_input.xlsx
    npi_source_column = None
    first_name_source_column = None
    last_name_source_column = None
    
    for _, row in mappings_df.iterrows():
        detected_as = str(row['Detected As']).strip()
        column_name = str(row['Column Name']).strip()
        
        # Handle multiple mappings (comma-separated in Detected As)
        detected_as_list = [d.strip() for d in detected_as.split(',')]
        
        for detected in detected_as_list:
            detected_normalized = detected.strip().lower()
            # Check if this is NPI Number mapping
            if detected_normalized in ['npi', 'npi number', 'npinumber']:
                npi_source_column = column_name
            # Check if this is First Name mapping (handle various formats: firstName, firstname, first name, first_name)
            elif detected_normalized in ['firstname', 'first name', 'first_name']:
                first_name_source_column = column_name
            # Check if this is Last Name mapping (handle various formats: lastName, lastname, last name, last_name)
            elif detected_normalized in ['lastname', 'last name', 'last_name']:
                last_name_source_column = column_name
            # PFS is intentionally skipped - not included in Locations_input.xlsx
    
    # Add NPI Number to locations_data if mapped
    if npi_source_column and npi_source_column in input_df.columns:
        locations_data['NPI Number'] = input_df[npi_source_column]
        print(f"Added NPI Number column from mapped column '{npi_source_column}'")
    elif npi_source_column:
        # NPI Number is mapped but source column doesn't exist
        locations_data['NPI Number'] = [''] * len(input_df)
        print(f"Warning: Source column '{npi_source_column}' for NPI Number not found in Input.xlsx. Filling with empty values.")
    
    # Add First Name to locations_data from _Mapped.xlsx if mapped
    if first_name_source_column:
        if mapped_df is not None and 'First Name' in mapped_df.columns:
            # Ensure the length matches input_df
            if len(mapped_df) == len(input_df):
                locations_data['First Name'] = mapped_df['First Name'].tolist()
                print(f"Added First Name column from _Mapped.xlsx")
            else:
                print(f"Warning: _Mapped.xlsx has {len(mapped_df)} rows but Input.xlsx has {len(input_df)} rows. First Name will be filled with empty values.")
                locations_data['First Name'] = [''] * len(input_df)
        else:
            # First Name is mapped but not found in _Mapped.xlsx
            locations_data['First Name'] = [''] * len(input_df)
            print(f"Warning: 'First Name' column not found in _Mapped.xlsx. Filling with empty values.")
    
    # Add Last Name to locations_data from _Mapped.xlsx if mapped
    if last_name_source_column:
        if mapped_df is not None and 'Last Name' in mapped_df.columns:
            # Ensure the length matches input_df
            if len(mapped_df) == len(input_df):
                locations_data['Last Name'] = mapped_df['Last Name'].tolist()
                print(f"Added Last Name column from _Mapped.xlsx")
            else:
                print(f"Warning: _Mapped.xlsx has {len(mapped_df)} rows but Input.xlsx has {len(input_df)} rows. Last Name will be filled with empty values.")
                locations_data['Last Name'] = [''] * len(input_df)
        else:
            # Last Name is mapped but not found in _Mapped.xlsx
            locations_data['Last Name'] = [''] * len(input_df)
            print(f"Warning: 'Last Name' column not found in _Mapped.xlsx. Filling with empty values.")
    
    # PFS is intentionally NOT included in Locations_input.xlsx
    
    # Create DataFrame
    locations_df = pd.DataFrame(locations_data)
    
    # Ensure all location columns are present (even if not mapped)
    # For Location ID, we only ensure the base column if no numbered ones exist
    location_id_columns = [col for col in locations_df.columns if col.startswith('Location ID')]
    for location_type in LOCATION_TYPES:
        column_name = readable_names[location_type]
        if location_type == 'locationId':
            # For Location ID, only add base column if no numbered Location ID columns exist
            if not location_id_columns and column_name not in locations_df.columns:
                locations_df[column_name] = [''] * len(locations_df)
        else:
            if column_name not in locations_df.columns:
                locations_df[column_name] = [''] * len(locations_df)
    
    # Reorder columns to match LOCATION_TYPES order
    # For Location ID, include all numbered Location ID columns after the base Location ID
    column_order = []
    for location_type in LOCATION_TYPES:
        column_name = readable_names[location_type]
        if location_type == 'locationId':
            # Add base Location ID if it exists
            if column_name in locations_df.columns:
                column_order.append(column_name)
            # Add all numbered Location ID columns
            numbered_location_ids = sorted([col for col in locations_df.columns if col.startswith('Location ID ') and col != 'Location ID'], 
                                         key=lambda x: int(x.split()[-1]) if x.split()[-1].isdigit() else 999)
            column_order.extend(numbered_location_ids)
        else:
            if column_name in locations_df.columns:
                column_order.append(column_name)
    
    # Add any remaining columns that weren't in the standard order
    remaining_columns = [col for col in locations_df.columns if col not in column_order]
    column_order.extend(remaining_columns)
    
    locations_df = locations_df[column_order]
    
    # Process Location Type_Raw to determine Location Type
    # If Location Type_Raw column exists, use it to determine Location Type
    # If not, default to 'In Person' (except for Case 1 rows which have Location ID columns already mapped)
    if 'Location Type_Raw' in locations_df.columns:
        print("Processing Location Type_Raw to determine Location Type...")
        location_types = []
        
        # Find all Location ID columns to identify Case 1 rows
        location_id_columns = [col for col in locations_df.columns if col.startswith('Location ID') and not col.startswith('Location Cloud ID') and not col.startswith('Location Monolith ID')]
        
        for idx, row in locations_df.iterrows():
            location_type_raw = row.get('Location Type_Raw', '')
            
            # Debug: Log the raw value for first few rows to verify it's being read correctly
            if idx < 3 and pd.notna(location_type_raw) and str(location_type_raw).strip() != '':
                print(f"DEBUG: Row {idx}: Location Type_Raw value = '{location_type_raw}' (type: {type(location_type_raw)})")
            
            # Check if this is a Case 1 row (has Location ID columns with values)
            is_case_1 = False
            if location_id_columns:
                for loc_id_col in location_id_columns:
                    loc_id_value = row.get(loc_id_col, '')
                    if pd.notna(loc_id_value) and str(loc_id_value).strip() != '':
                        is_case_1 = True
                        break
            
            # For Case 1, if Location Type_Raw is empty, don't apply default
            if is_case_1 and (pd.isna(location_type_raw) or str(location_type_raw).strip() == ''):
                location_type = ''  # Leave empty for Case 1
            else:
                location_type = determine_location_type(location_type_raw)
            
            location_types.append(location_type)
        
        # Add Location Type column if it doesn't exist
        if 'Location Type' not in locations_df.columns:
            # Insert Location Type column after Location Type_Raw
            location_type_raw_idx = locations_df.columns.get_loc('Location Type_Raw')
            locations_df.insert(location_type_raw_idx + 1, 'Location Type', location_types)
        else:
            locations_df['Location Type'] = location_types
        
        print(f"Determined Location Types: {location_types.count('In Person')} In Person, {location_types.count('Virtual')} Virtual, {location_types.count('Both')} Both, {location_types.count('')} Empty (Case 1)")
    else:
        # No Location Type_Raw column - default all to 'In Person' (except Case 1 rows)
        print("No Location Type_Raw column found. Defaulting Location Types to 'In Person' (except Case 1 rows)...")
        
        # Find all Location ID columns to identify Case 1 rows
        location_id_columns = [col for col in locations_df.columns if col.startswith('Location ID') and not col.startswith('Location Cloud ID') and not col.startswith('Location Monolith ID')]
        
        location_types = []
        for idx, row in locations_df.iterrows():
            # Check if this is a Case 1 row (has Location ID columns with values)
            is_case_1 = False
            if location_id_columns:
                for loc_id_col in location_id_columns:
                    loc_id_value = row.get(loc_id_col, '')
                    if pd.notna(loc_id_value) and str(loc_id_value).strip() != '':
                        is_case_1 = True
                        break
            
            # For Case 1, leave empty; otherwise default to 'In Person'
            if is_case_1:
                location_types.append('')
            else:
                location_types.append('In Person')
        
        if 'Location Type' not in locations_df.columns:
            locations_df['Location Type'] = location_types
        else:
            locations_df['Location Type'] = location_types
    
    # Save to Excel
    locations_df.to_excel(output_file, index=False, sheet_name='Locations')
    
    # Apply styling to the Excel file
    wb = load_workbook(output_file)
    ws = wb['Locations']
    
    # Style header row
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Highlight Location Type n columns based on Location Type comparison
    if 'Location Type' in locations_df.columns:
        green_fill = PatternFill(start_color='ABFFAB', end_color='ABFFAB', fill_type='solid')  # Light green
        red_fill = PatternFill(start_color='FFA3A3', end_color='FFA3A3', fill_type='solid')  # Light red
        
        # Find Location Type column index
        location_type_col_idx = None
        for col_idx, col_name in enumerate(locations_df.columns, start=1):
            if col_name == 'Location Type':
                location_type_col_idx = col_idx
                break
        
        if location_type_col_idx:
            # Find all Location Type n columns
            location_type_n_columns = []
            for col_idx, col_name in enumerate(locations_df.columns, start=1):
                if col_name.startswith('Location Type ') and col_name != 'Location Type':
                    # Extract the number
                    number = col_name.replace('Location Type ', '').strip()
                    if number.isdigit():
                        location_type_n_columns.append((col_idx, col_name, int(number)))
            
            # Sort by number
            location_type_n_columns.sort(key=lambda x: x[2])
            
            if location_type_n_columns:
                print(f"Highlighting {len(location_type_n_columns)} Location Type n column(s) based on Location Type comparison...")
                
                # Process each row
                for row_idx, (idx, row) in enumerate(locations_df.iterrows(), start=2):  # Start at 2 (row 1 is header)
                    location_type = row.get('Location Type', '')
                    location_type_str = str(location_type).strip() if pd.notna(location_type) else ''
                    
                    # Highlight each Location Type n column based on comparison
                    for col_idx, col_name, number in location_type_n_columns:
                        location_type_n_val = row.get(col_name, '')
                        location_type_n_str = str(location_type_n_val).strip() if pd.notna(location_type_n_val) else ''
                        
                        # Determine highlight color based on rules
                        highlight_color = None
                        
                        if location_type_str == 'Both':
                            # Both → all Location Type n should be green
                            highlight_color = green_fill
                        elif location_type_str == 'In Person':
                            # In Person → green if Location Type n is 'In Person', red otherwise
                            if location_type_n_str == 'In Person':
                                highlight_color = green_fill
                            else:
                                highlight_color = red_fill
                        elif location_type_str == 'Virtual':
                            # Virtual → green if Location Type n is 'Virtual', red otherwise
                            if location_type_n_str == 'Virtual':
                                highlight_color = green_fill
                            else:
                                highlight_color = red_fill
                        
                        # Apply highlighting if color determined
                        if highlight_color:
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.fill = highlight_color
                
                print(f"Highlighted Location Type n columns based on Location Type comparison")
    
    wb.save(output_file)
    
    print(f"Created Locations_input.xlsx with {len(locations_df)} rows and {len(locations_df.columns)} location columns")
    print(f"Location columns included: {', '.join(locations_df.columns.tolist())}")
    
    # Create Practice_Locations.xlsx with unique Practice ID values
    practice_locations_file = EXCEL_FILES_DIR / 'Practice_Locations.xlsx'
    create_practice_locations_excel(mappings_df, input_df, practice_locations_file, manual_practice_ids=manual_practice_ids)
    
    # Format address lines in Locations_input.xlsx (if needed)
    print("\nFormatting address lines...")
    try:
        import sys
        from pathlib import Path
        sys.path.insert(0, str(Path(__file__).parent))
        from format_address_lines import format_locations_input_addresses
        format_locations_input_addresses(manual_practice_ids=manual_practice_ids)
    except ImportError:
        try:
            from .format_address_lines import format_locations_input_addresses
            format_locations_input_addresses(manual_practice_ids=manual_practice_ids)
        except Exception as e:
            print(f"Warning: Could not format address lines: {str(e)}")
    except Exception as e:
        print(f"Warning: Could not format address lines: {str(e)}")
    
    # Fetch location details from API and add to Practice_Locations.xlsx
    # This is called after Locations_input.xlsx is created
    print("\nFetching location details from API...")
    try:
        import sys
        from pathlib import Path
        sys.path.insert(0, str(Path(__file__).parent))
        from fetch_practice_locations import fetch_and_write_locations
        fetch_and_write_locations(practice_locations_file)
        
        # After fetching locations, add Location Cloud ID to Locations_input.xlsx
        add_location_cloud_ids_to_locations_input(output_file, practice_locations_file)
    except ImportError:
        # If import fails, try relative import
        try:
            from .fetch_practice_locations import fetch_and_write_locations
            fetch_and_write_locations(practice_locations_file)
            # After fetching locations, add Location Cloud ID to Locations_input.xlsx
            add_location_cloud_ids_to_locations_input(output_file, practice_locations_file)
        except Exception as e:
            print(f"Warning: Could not fetch location details from API: {str(e)}")
            print("Practice_Locations.xlsx created with Practice IDs only.")
    except Exception as e:
        print(f"Warning: Could not fetch location details from API: {str(e)}")
        print("Practice_Locations.xlsx created with Practice IDs only.")
    
    return str(output_file)


def convert_to_cloud_ids(practice_ids):
    """
    Convert monolith practice IDs to cloud practice IDs
    
    Args:
        practice_ids: List of monolith practice IDs (as strings or integers)
        
    Returns:
        Dictionary mapping monolith_id -> cloud_id
    """
    # Convert all IDs to strings
    practice_ids = [str(pid) for pid in practice_ids]
    
    PRACTICE_IDS_API = 'https://provider-reference-v1.east.zocdoccloud.com/provider-reference/v1/practice/ids-by-monolith-ids~batchGet'
    url = PRACTICE_IDS_API
    headers = {
        'accept': 'application/json',
        'Content-Type': 'application/json'
    }
    data = {
        "monolith_practice_ids": practice_ids
    }
    
    print("Fetching Practice Cloud IDs...")
    try:
        response = requests.post(url, headers=headers, json=data, timeout=30)
        
        cloud_id_map = {}
        reverse_map = {}  # cloud_id -> monolith_id
        if response.status_code == 200:
            result = response.json()
            for item in result.get('practice_ids', []):
                monolith_id = str(item.get('monolith_practice_id'))
                cloud_id = item.get('practice_id')
                if cloud_id:
                    cloud_id_map[monolith_id] = cloud_id
                    reverse_map[cloud_id] = monolith_id
            print(f"Successfully converted {len(cloud_id_map)} Practice IDs to Cloud IDs")
            return cloud_id_map, reverse_map
        else:
            print(f"Failed to fetch cloud IDs. Status code: {response.status_code}")
            print(response.text)
            return {}, {}
    except Exception as e:
        print(f"Error converting Practice IDs to Cloud IDs: {str(e)}")
        return {}, {}


def get_monolith_ids_from_cloud_ids(cloud_ids, existing_reverse_map=None):
    """
    Try to get monolith IDs from Cloud IDs by fetching location details.
    This is a fallback method when we only have Cloud IDs.
    
    Args:
        cloud_ids: List of Cloud IDs
        existing_reverse_map: Optional existing reverse mapping (cloud_id -> monolith_id)
        
    Returns:
        Dictionary mapping cloud_id -> monolith_id (may contain None values if not found)
    """
    if existing_reverse_map is None:
        existing_reverse_map = {}
    
    LOCATION_API = 'https://provider-reference-v1.east.zocdoccloud.com/provider-reference/v1/practice/location~batchGet'
    reverse_map = existing_reverse_map.copy()
    
    # Try to fetch location details to see if we can extract monolith practice ID
    # Note: This may not always work as location API might not return monolith practice ID
    # But we'll try for any Cloud IDs not already in the reverse map
    cloud_ids_to_fetch = [cid for cid in cloud_ids if cid not in reverse_map]
    
    if not cloud_ids_to_fetch:
        return reverse_map
    
    # Fetch locations for Cloud IDs (batch)
    url = LOCATION_API
    headers = {
        'accept': 'application/json',
        'Content-Type': 'application/json'
    }
    data = {"practice_ids": cloud_ids_to_fetch}
    
    try:
        response = requests.post(url, headers=headers, json=data, timeout=30)
        if response.status_code == 200:
            result = response.json()
            locations = result.get('practice_locations', [])
            
            # Try to extract monolith practice ID from location data
            # Note: Location API response structure may vary
            for location in locations:
                cloud_id = location.get('practice_id')
                # Check if location has monolith practice ID (may not be available)
                monolith_id = location.get('monolith_practice_id')
                if cloud_id and monolith_id:
                    reverse_map[cloud_id] = str(monolith_id)
    except Exception as e:
        print(f"Warning: Could not fetch location details to get monolith IDs: {str(e)}")
    
    # For any Cloud IDs not found, set to None
    for cloud_id in cloud_ids:
        if cloud_id not in reverse_map:
            reverse_map[cloud_id] = None
    
    return reverse_map


def create_practice_locations_excel(mappings_df, input_df, output_file, manual_practice_ids=''):
    """
    Create Practice_Locations.xlsx file with unique Practice ID values
    
    This function handles both Practice IDs (monolith) and Practice Cloud IDs.
    - If 'Practice ID' is mapped, it converts to 'Practice Cloud ID' and fills both columns
    - If 'Practice Cloud ID' is mapped, it tries to get 'Practice ID' and fills both columns
    - Manual entry can contain either Practice IDs or Practice Cloud IDs
    
    Args:
        mappings_df: DataFrame containing mappings from Mappings.xlsx
        input_df: DataFrame containing input data from Input.xlsx
        output_file: Path to the output Practice_Locations.xlsx file
        manual_practice_ids: Optional string of manually entered Practice IDs or Practice Cloud IDs 
                            (comma, semicolon, or space separated)
    """
    # Find Practice ID and Practice Cloud ID mappings
    practice_id_column = None
    practice_cloud_id_column = None
    
    # Create reverse mapping: readable_name -> internal_name
    readable_to_internal = {v: k for k, v in readable_names.items()}
    
    for _, row in mappings_df.iterrows():
        detected_as = str(row['Detected As']).strip()
        column_name = str(row['Column Name']).strip()
        
        # Handle multiple mappings (comma-separated in Detected As)
        detected_as_list = [d.strip() for d in detected_as.split(',')]
        
        for detected in detected_as_list:
            detected_normalized = detected.strip()
            
            # Check if this is Practice ID mapping
            # First check if it matches readable name directly
            if detected_normalized == 'Practice ID':
                practice_id_column = column_name
            # Check if it matches internal name
            elif detected_normalized.lower() == 'practiceid' or detected_normalized.lower() == 'practice id':
                practice_id_column = column_name
            # Check if it's in readable_to_internal mapping
            elif detected_normalized in readable_to_internal:
                internal_name = readable_to_internal[detected_normalized]
                if internal_name == 'practiceId':
                    practice_id_column = column_name
            # Check if it matches internal name (lowercase comparison)
            else:
                detected_lower = detected_normalized.lower()
                if detected_lower == 'practiceid' or detected_lower == 'practice id':
                    practice_id_column = column_name
            
            # Check if this is Practice Cloud ID mapping
            if detected_normalized == 'Practice Cloud ID':
                practice_cloud_id_column = column_name
            elif detected_normalized.lower() == 'practicecloudid' or detected_normalized.lower() == 'practice cloud id':
                practice_cloud_id_column = column_name
    
    # Collect Practice IDs and Practice Cloud IDs from multiple sources
    practice_ids_from_column = []  # These are definitely monolith IDs
    practice_cloud_ids_from_column = []  # These are definitely Cloud IDs
    manual_ids = []  # These could be Practice IDs or Practice Cloud IDs
    
    # 1. Get Practice IDs from mapped column if exists (these are monolith IDs)
    if practice_id_column and practice_id_column in input_df.columns:
        practice_ids = input_df[practice_id_column].dropna().astype(str).str.strip()
        practice_ids = practice_ids[practice_ids != '']
        practice_ids = practice_ids.str.replace(r'\.0$', '', regex=True)
        practice_ids_from_column = practice_ids.unique().tolist()
        print(f"Found {len(practice_ids_from_column)} unique Practice ID(s) (monolith) from mapped column '{practice_id_column}'")
    
    # 2. Get Practice Cloud IDs from mapped column if exists (these are Cloud IDs)
    if practice_cloud_id_column and practice_cloud_id_column in input_df.columns:
        cloud_ids = input_df[practice_cloud_id_column].dropna().astype(str).str.strip()
        cloud_ids = cloud_ids[cloud_ids != '']
        cloud_ids = cloud_ids.str.replace(r'\.0$', '', regex=True)
        practice_cloud_ids_from_column = cloud_ids.unique().tolist()
        print(f"Found {len(practice_cloud_ids_from_column)} unique Practice Cloud ID(s) from mapped column '{practice_cloud_id_column}'")
    
    # 3. Get IDs from manual entry if provided (could be Practice IDs or Practice Cloud IDs)
    if manual_practice_ids and manual_practice_ids.strip():
        # Parse manual IDs (comma, semicolon, or space separated)
        # Split by comma, semicolon, or whitespace
        parsed_ids = re.split(r'[,;\s]+', manual_practice_ids.strip())
        parsed_ids = [pid.strip() for pid in parsed_ids if pid.strip()]
        # Remove .0 suffix if present
        parsed_ids = [re.sub(r'\.0$', '', pid) for pid in parsed_ids]
        manual_ids = list(set(parsed_ids))  # Get unique values
        print(f"Found {len(manual_ids)} unique ID(s) from manual entry (will be processed as Practice IDs, converted to Cloud IDs if needed)")
    
    # Combine all IDs
    all_practice_ids = practice_ids_from_column + manual_ids
    all_practice_cloud_ids = practice_cloud_ids_from_column
    
    # If no Practice IDs or Practice Cloud IDs found from any source
    if not all_practice_ids and not all_practice_cloud_ids:
        if not practice_id_column and not practice_cloud_id_column:
            print("Warning: No Practice ID or Practice Cloud ID mapping found and no manual IDs provided. Creating empty Practice_Locations.xlsx")
        else:
            print(f"Warning: Practice ID/Practice Cloud ID columns not found in Input.xlsx and no manual IDs provided. Creating empty Practice_Locations.xlsx")
        empty_df = pd.DataFrame({'Practice ID': [], 'Practice Cloud ID': []})
        empty_df.to_excel(output_file, index=False, sheet_name='Practice')
        return
    
    # Get unique values
    unique_practice_ids = list(set(all_practice_ids)) if all_practice_ids else []
    unique_practice_cloud_ids = list(set(all_practice_cloud_ids)) if all_practice_cloud_ids else []
    
    # Step 1: Convert Practice IDs to Cloud IDs (if Practice IDs are provided)
    practice_id_to_cloud_id = {}  # monolith_id -> cloud_id
    cloud_id_to_practice_id = {}  # cloud_id -> monolith_id
    
    if unique_practice_ids:
        print(f"Converting {len(unique_practice_ids)} Practice ID(s) to Cloud ID(s)...")
        cloud_id_map, reverse_map = convert_to_cloud_ids(unique_practice_ids)
        practice_id_to_cloud_id.update(cloud_id_map)
        cloud_id_to_practice_id.update(reverse_map)
    
    # Step 2: Try to get Practice IDs from Cloud IDs (if Cloud IDs are provided)
    if unique_practice_cloud_ids:
        print(f"Attempting to get Practice ID(s) for {len(unique_practice_cloud_ids)} Cloud ID(s)...")
        reverse_map_from_cloud = get_monolith_ids_from_cloud_ids(unique_practice_cloud_ids, cloud_id_to_practice_id)
        cloud_id_to_practice_id.update(reverse_map_from_cloud)
        
        # Also update forward map for any found monolith IDs
        for cloud_id, monolith_id in reverse_map_from_cloud.items():
            if monolith_id:
                practice_id_to_cloud_id[monolith_id] = cloud_id
    
    # Step 3: Build combined list with both Practice IDs and Cloud IDs
    # Create lookup dictionaries for merging
    by_practice_id = {}  # practice_id -> (practice_id, cloud_id)
    by_cloud_id = {}  # cloud_id -> (practice_id, cloud_id)
    
    # Add pairs from Practice IDs
    for practice_id in unique_practice_ids:
        cloud_id = practice_id_to_cloud_id.get(practice_id)
        pair = (practice_id, cloud_id)
        by_practice_id[practice_id] = pair
        if cloud_id:
            by_cloud_id[cloud_id] = pair
    
    # Add pairs from Cloud IDs and merge with existing if needed
    for cloud_id in unique_practice_cloud_ids:
        monolith_id = cloud_id_to_practice_id.get(cloud_id)
        
        if cloud_id in by_cloud_id:
            # Merge: use the practice_id from existing pair if available, otherwise use found monolith_id
            existing_pid, existing_cid = by_cloud_id[cloud_id]
            final_pid = existing_pid or monolith_id
            merged_pair = (final_pid, cloud_id)
            by_cloud_id[cloud_id] = merged_pair
            if final_pid:
                by_practice_id[final_pid] = merged_pair
        else:
            # New pair
            pair = (monolith_id, cloud_id)
            by_cloud_id[cloud_id] = pair
            if monolith_id:
                by_practice_id[monolith_id] = pair
    
    # Collect all unique pairs
    seen = set()
    unique_pairs = []
    for pair in list(by_practice_id.values()) + list(by_cloud_id.values()):
        if pair not in seen:
            seen.add(pair)
            unique_pairs.append(pair)
    
    # Sort pairs: first by Practice ID (if available), then by Cloud ID
    unique_pairs.sort(key=lambda x: (
        x[0] is None,
        int(x[0]) if x[0] and x[0].isdigit() else (x[0] or ''),
        x[1] or ''
    ))
    
    # Create DataFrame with both columns
    practice_ids_list = [pair[0] for pair in unique_pairs]
    cloud_ids_list = [pair[1] for pair in unique_pairs]
    
    practice_data = {
        'Practice ID': practice_ids_list,
        'Practice Cloud ID': cloud_ids_list
    }
    
    practice_df = pd.DataFrame(practice_data)
    
    # Ensure both columns exist (even if empty)
    if 'Practice ID' not in practice_df.columns:
        practice_df['Practice ID'] = [None] * len(practice_df)
    if 'Practice Cloud ID' not in practice_df.columns:
        practice_df['Practice Cloud ID'] = [None] * len(practice_df)
    
    # Fetch Practice Names from API using Practice Cloud IDs
    # This works for both: Cloud IDs converted from Monolith IDs AND directly provided Cloud IDs
    practice_names_list = [None] * len(practice_df)
    valid_cloud_ids = []
    cloud_id_indices = []
    
    # Process all Cloud IDs in the list (regardless of their source)
    for idx, cloud_id in enumerate(cloud_ids_list):
        if cloud_id and pd.notna(cloud_id):
            cloud_id_str = str(cloud_id).strip()
            # Check if it's a valid Cloud ID (starts with 'pt_')
            if cloud_id_str.startswith('pt_'):
                valid_cloud_ids.append(cloud_id_str)
                cloud_id_indices.append(idx)
    
    # Fetch Practice Names from API if we have valid Cloud IDs
    if valid_cloud_ids:
        print(f"Fetching Practice Names for {len(valid_cloud_ids)} Practice Cloud ID(s)...")
        try:
            # Import the function from Practice_name.py
            logics_dir = Path(__file__).parent
            sys.path.insert(0, str(logics_dir))
            from Practice_name import get_practice_details_from_api
            
            # Call API to get practice names
            details_result = get_practice_details_from_api(valid_cloud_ids)
            
            if details_result['success']:
                practice_names = details_result['practice_names']
                # Map practice names back to DataFrame indices
                for i, cloud_id in enumerate(valid_cloud_ids):
                    idx = cloud_id_indices[i]
                    if cloud_id in practice_names:
                        practice_names_list[idx] = practice_names[cloud_id]
                print(f"Successfully fetched {len([n for n in practice_names_list if n])} Practice Name(s)")
            else:
                print(f"Warning: Could not fetch Practice Names from API: {details_result.get('error', 'Unknown error')}")
        except ImportError:
            print("Warning: Could not import Practice_name module. Skipping Practice Name fetch.")
        except Exception as e:
            print(f"Warning: Error fetching Practice Names: {str(e)}")
    else:
        # Debug: Check if we have Cloud IDs but they're not in the expected format
        non_pt_cloud_ids = [str(cid).strip() for cid in cloud_ids_list if cid and pd.notna(cid) and not str(cid).strip().startswith('pt_')]
        if non_pt_cloud_ids:
            print(f"Note: Found {len(non_pt_cloud_ids)} Cloud ID(s) that don't start with 'pt_' (skipping Practice Name fetch for these)")
    
    # Add Practice NAME column to DataFrame
    practice_df['Practice NAME'] = practice_names_list
    
    # Reorder columns: Practice ID, Practice Cloud ID, Practice NAME
    practice_df = practice_df[['Practice ID', 'Practice Cloud ID', 'Practice NAME']]
    
    # Save to Excel
    practice_df.to_excel(output_file, index=False, sheet_name='Practice')
    
    # Apply styling to the Excel file
    wb = load_workbook(output_file)
    ws = wb['Practice']
    
    # Style header row
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Auto-adjust column widths (now includes Practice NAME column)
    for col_idx, col_letter in enumerate(['A', 'B', 'C'], 1):
        if col_idx <= len(practice_df.columns):
            max_length = 0
            for cell in ws[col_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50) if max_length > 0 else 15
            ws.column_dimensions[col_letter].width = adjusted_width
    
    wb.save(output_file)
    
    # Count non-null values
    total_practice_ids = practice_df['Practice ID'].notna().sum()
    total_cloud_ids = practice_df['Practice Cloud ID'].notna().sum()
    total_practice_names = practice_df['Practice NAME'].notna().sum()
    total_rows = len(practice_df)
    
    print(f"Created Practice_Locations.xlsx with {total_rows} row(s)")
    print(f"  - {total_practice_ids} Practice ID(s) (monolith)")
    print(f"  - {total_cloud_ids} Practice Cloud ID(s)")
    print(f"  - {total_practice_names} Practice NAME(s)")
    if practice_id_column:
        print(f"Practice IDs extracted from column: '{practice_id_column}'")
    if practice_cloud_id_column:
        print(f"Practice Cloud IDs extracted from column: '{practice_cloud_id_column}'")
    if manual_ids:
        print(f"Manual IDs added: {len(manual_ids)} unique value(s)")
    
    # Show mapping summary
    mapped_count = sum(1 for pair in unique_pairs if pair[0] and pair[1])
    if mapped_count > 0:
        print(f"Successfully mapped {mapped_count} Practice ID(s) to Cloud ID(s) (both columns filled)")
    if total_rows > mapped_count:
        print(f"Note: {total_rows - mapped_count} row(s) have only one ID filled (could not find corresponding ID)")


def normalize_string(s):
    """
    Normalize a string for fuzzy matching: lowercase, strip, remove extra spaces
    """
    if pd.isna(s) or s == '':
        return ''
    s = str(s).strip().lower()
    # Remove extra spaces
    s = ' '.join(s.split())
    return s


def extract_numbers(s):
    """
    Extract all numbers from a string and return them as a list of strings
    """
    if pd.isna(s) or s == '':
        return []
    s = str(s)
    # Find all sequences of digits
    numbers = re.findall(r'\d+', s)
    return numbers


def fuzzy_match_score(str1, str2, threshold=0.8):
    """
    Calculate fuzzy match score between two strings using SequenceMatcher.
    Numbers are not considered in the fuzzy matching - if addresses have different numbers,
    they will not match.
    
    Returns True if similarity >= threshold, False otherwise
    """
    str1_norm = normalize_string(str1)
    str2_norm = normalize_string(str2)
    
    if not str1_norm or not str2_norm:
        return False
    
    # Extract numbers from both strings
    numbers1 = extract_numbers(str1)
    numbers2 = extract_numbers(str2)
    
    # If both strings have numbers, they must match exactly
    if numbers1 and numbers2:
        if numbers1 != numbers2:
            return False
    
    # Remove numbers from both strings for fuzzy matching
    str1_no_numbers = re.sub(r'\d+', '', str1_norm).strip()
    str2_no_numbers = re.sub(r'\d+', '', str2_norm).strip()
    
    # Remove extra spaces after removing numbers
    str1_no_numbers = ' '.join(str1_no_numbers.split())
    str2_no_numbers = ' '.join(str2_no_numbers.split())
    
    if not str1_no_numbers or not str2_no_numbers:
        # If after removing numbers one string is empty, check exact match of original
        return str1_norm == str2_norm
    
    # Exact match after normalization and number removal
    if str1_no_numbers == str2_no_numbers:
        return True
    
    # Calculate similarity ratio on strings without numbers
    similarity = SequenceMatcher(None, str1_no_numbers, str2_no_numbers).ratio()
    return similarity >= threshold


def replace_address_values_from_matched_location(locations_df, api_locations_df, row_idx, location_cloud_id):
    """
    Replace address values in Locations_input.xlsx with values from matched location in Practice_Locations.xlsx.
    
    Args:
        locations_df: DataFrame from Locations_input.xlsx
        api_locations_df: DataFrame from Practice_Locations.xlsx Locations sheet
        row_idx: Index of the row in locations_df to update
        location_cloud_id: Location Cloud ID to look up in api_locations_df
    
    Returns:
        True if replacement was made, False otherwise
    """
    if not location_cloud_id or pd.isna(location_cloud_id) or str(location_cloud_id).strip() == '':
        return False
    
    location_cloud_id_str = str(location_cloud_id).strip()
    location_cloud_id_str = re.sub(r'\.0$', '', location_cloud_id_str)
    
    # Find the matched location in api_locations_df
    matched_location = api_locations_df[
        api_locations_df['location_id'].astype(str).str.strip().apply(lambda x: re.sub(r'\.0$', '', x) if pd.notna(x) else '') == location_cloud_id_str
    ]
    
    if matched_location.empty:
        return False
    
    # Get the first match
    location_row = matched_location.iloc[0]
    
    # Extract address values
    address_1 = location_row.get('address_1', '')
    address_2 = location_row.get('address_2', '')
    city = location_row.get('city', '')
    state = location_row.get('state', '')
    zip_code = location_row.get('zip', '')
    
    # Replace values in locations_df (always replace, not just if empty)
    replaced = False
    
    if 'Address Line 1' in locations_df.columns:
        address_1_value = str(address_1).strip() if pd.notna(address_1) and address_1 != '' else ''
        locations_df.at[row_idx, 'Address Line 1'] = address_1_value
        replaced = True
    
    if 'Address Line 2' in locations_df.columns:
        address_2_value = str(address_2).strip() if pd.notna(address_2) and address_2 != '' else ''
        locations_df.at[row_idx, 'Address Line 2'] = address_2_value
        replaced = True
    
    if 'City' in locations_df.columns:
        city_value = str(city).strip() if pd.notna(city) and city != '' else ''
        locations_df.at[row_idx, 'City'] = city_value
        replaced = True
    
    if 'State' in locations_df.columns:
        state_value = str(state).strip() if pd.notna(state) and state != '' else ''
        locations_df.at[row_idx, 'State'] = state_value
        replaced = True
    
    if 'ZIP' in locations_df.columns:
        if pd.notna(zip_code) and zip_code != '':
            zip_value = zip_code
            # Special handling for ZIP to preserve leading zeros
            if isinstance(zip_value, (int, float)):
                zip_value = str(int(zip_value)).zfill(5)
            else:
                zip_str = str(zip_value).strip()
                if zip_str.isdigit() and len(zip_str) < 5:
                    zip_value = zip_str.zfill(5)
                else:
                    zip_value = zip_str
            locations_df.at[row_idx, 'ZIP'] = zip_value
            replaced = True
    
    return replaced


def find_virtual_location_by_state(state_value, api_locations_df):
    """
    Find virtual location by state where address_1 = '--' and state matches.
    
    Args:
        state_value: State value to match
        api_locations_df: DataFrame from Practice_Locations.xlsx Locations sheet
    
    Returns:
        Dictionary with location info if found, None otherwise
    """
    if pd.isna(state_value) or str(state_value).strip() == '':
        return None
    
    state_str = str(state_value).strip()
    state_norm = normalize_string(state_str)
    
    if not state_norm:
        return None
    
    # Check if required columns exist
    if 'address_1' not in api_locations_df.columns or 'state' not in api_locations_df.columns:
        return None
    
    # Find locations where address_1 = '--' and state matches
    virtual_locations = api_locations_df[
        (api_locations_df['address_1'].astype(str).str.strip() == '--') &
        (api_locations_df['state'].astype(str).str.strip().apply(lambda x: normalize_string(x) if pd.notna(x) else '') == state_norm)
    ]
    
    if virtual_locations.empty:
        return None
    
    # Take the first match
    virtual_location = virtual_locations.iloc[0]
    
    # Extract location details
    location_id = virtual_location.get('location_id')
    monolith_location_id = virtual_location.get('monolith_location_id')
    location_type = virtual_location.get('Location Type', '')
    practice_monolith_id = virtual_location.get('Practice Monolith ID', '')
    practice_cloud_id = virtual_location.get('Practice Cloud ID', '')
    
    if pd.notna(location_id):
        # Convert to string and remove .0 suffix if present
        location_id_str = str(location_id).strip()
        location_id_str = re.sub(r'\.0$', '', location_id_str)
        
        # Process monolith_location_id
        monolith_id_str = ''
        if pd.notna(monolith_location_id):
            monolith_id_str = str(monolith_location_id).strip()
            monolith_id_str = re.sub(r'\.0$', '', monolith_id_str)
        
        # Process location_type
        location_type_str = ''
        if pd.notna(location_type):
            location_type_str = str(location_type).strip()
        
        # Process practice_monolith_id
        practice_monolith_id_str = ''
        if pd.notna(practice_monolith_id) and practice_monolith_id != '':
            practice_monolith_id_str = str(practice_monolith_id).strip()
            practice_monolith_id_str = re.sub(r'\.0$', '', practice_monolith_id_str)
        
        # Process practice_cloud_id
        practice_cloud_id_str = ''
        if pd.notna(practice_cloud_id) and practice_cloud_id != '':
            practice_cloud_id_str = str(practice_cloud_id).strip()
        
        return {
            'location_id': location_id_str,
            'monolith_location_id': monolith_id_str,
            'location_type': location_type_str,
            'practice_monolith_id': practice_monolith_id_str,
            'practice_cloud_id': practice_cloud_id_str
        }
    
    return None


def match_location_by_address(locations_row, api_locations_df, practice_id=None, practice_cloud_id=None):
    """
    Match a location row from Locations_input.xlsx with locations from Practice_Locations.xlsx
    using fuzzy matching on address components.
    
    Args:
        locations_row: Series representing a row from Locations_input.xlsx
        api_locations_df: DataFrame from Practice_Locations.xlsx Locations sheet
        practice_id: Practice ID (monolith) value to filter by (optional)
        practice_cloud_id: Practice Cloud ID value to filter by (optional)
        
    Returns:
        List of tuples (monolith_location_id, location_id, location_type) for matched locations
    """
    # Normalize Practice ID if provided
    practice_id_str = None
    if practice_id is not None and pd.notna(practice_id):
        practice_id_str = str(practice_id).strip()
        practice_id_str = re.sub(r'\.0$', '', practice_id_str)
        if not practice_id_str:
            practice_id_str = None
    
    # Normalize Practice Cloud ID if provided
    practice_cloud_id_str = None
    if practice_cloud_id is not None and pd.notna(practice_cloud_id):
        practice_cloud_id_str = str(practice_cloud_id).strip()
        practice_cloud_id_str = re.sub(r'\.0$', '', practice_cloud_id_str)
        if not practice_cloud_id_str:
            practice_cloud_id_str = None
    
    # Filter API locations by Practice Monolith ID and/or Practice Cloud ID (if provided)
    # If no IDs provided, use all locations for matching
    if practice_id_str and practice_cloud_id_str:
        # Both provided - match either
        filtered_df = api_locations_df[
            (api_locations_df['Practice Monolith ID'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True) == practice_id_str) |
            (api_locations_df['Practice Cloud ID'].astype(str).str.strip() == practice_cloud_id_str)
        ].copy()
    elif practice_id_str:
        # Only Practice ID provided
        filtered_df = api_locations_df[
            api_locations_df['Practice Monolith ID'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True) == practice_id_str
        ].copy()
    elif practice_cloud_id_str:
        # Only Practice Cloud ID provided
        filtered_df = api_locations_df[
            api_locations_df['Practice Cloud ID'].astype(str).str.strip() == practice_cloud_id_str
        ].copy()
    else:
        # No Practice ID provided - match against all locations
        filtered_df = api_locations_df.copy()
    
    if filtered_df.empty:
        return []
    
    # Get address components from locations_row
    addr_line1 = locations_row.get('Address Line 1', '') if 'Address Line 1' in locations_row else ''
    addr_line2 = locations_row.get('Address Line 2', '') if 'Address Line 2' in locations_row else ''
    city = locations_row.get('City', '') if 'City' in locations_row else ''
    state = locations_row.get('State', '') if 'State' in locations_row else ''
    zip_code = locations_row.get('ZIP', '') if 'ZIP' in locations_row else ''
    
    # Normalize address components
    addr_line1_norm = normalize_string(addr_line1)
    addr_line2_norm = normalize_string(addr_line2)
    city_norm = normalize_string(city)
    state_norm = normalize_string(state)
    zip_norm = normalize_string(zip_code)
    
    if not addr_line1_norm:
        return []
    
    # Check if only Address Line 1 is mapped (Address Line 2, City, State, ZIP are all empty)
    # In this case, we concatenate address components from Practice_Locations.xlsx and match against full Address Line 1
    only_address_line1_mapped = not addr_line2_norm and not city_norm and not state_norm and not zip_norm
    
    if only_address_line1_mapped:
        # Special case: Only Address Line 1 is mapped
        # Concatenate address_1, address_2, city, state, zip from Practice_Locations.xlsx
        # and match against the full Address Line 1 from Locations_input.xlsx
        matches = []
        for idx, api_row in filtered_df.iterrows():
            # Get address components from API location
            api_addr1 = api_row.get('address_1', '')
            api_addr2 = api_row.get('address_2', '')
            api_city = api_row.get('city', '')
            api_state = api_row.get('state', '')
            api_zip = api_row.get('zip', '')
            
            # Concatenate all address components from Practice_Locations.xlsx
            api_address_parts = []
            if api_addr1 and str(api_addr1).strip():
                api_address_parts.append(str(api_addr1).strip())
            if api_addr2 and str(api_addr2).strip():
                api_address_parts.append(str(api_addr2).strip())
            if api_city and str(api_city).strip():
                api_address_parts.append(str(api_city).strip())
            if api_state and str(api_state).strip():
                api_address_parts.append(str(api_state).strip())
            if api_zip and str(api_zip).strip():
                api_address_parts.append(str(api_zip).strip())
            
            # Join with space to create full concatenated address
            api_address_combined = ' '.join(api_address_parts)
            
            # Normalize the concatenated address
            api_address_combined_norm = normalize_string(api_address_combined)
            
            # Fuzzy match the full Address Line 1 against the concatenated address
            if api_address_combined_norm and fuzzy_match_score(addr_line1_norm, api_address_combined_norm, threshold=0.7):
                matches.append((idx, api_row))
        
        # No further filtering needed - we've already matched against the full concatenated address
    else:
        # Normal case: Multiple address components are mapped
        # Step 1: Fuzzy match Address Line 1
        matches = []
        for idx, api_row in filtered_df.iterrows():
            api_addr1 = api_row.get('address_1', '')
            api_addr1_norm = normalize_string(api_addr1)
            
            if fuzzy_match_score(addr_line1_norm, api_addr1_norm, threshold=0.7):
                matches.append((idx, api_row))
        
        if not matches:
            return []
        
        # Step 2: Filter matches by Address Line 2
        # If Address Line 2 is empty in input, only match with empty address_2
        # If Address Line 2 has a value in input, only match with non-empty address_2 (strict matching)
        filtered_matches = []
        for idx, api_row in matches:
            api_addr2 = api_row.get('address_2', '')
            api_addr2_norm = normalize_string(api_addr2)
            
            # If Address Line 2 is empty in input, only match with empty address_2
            if not addr_line2_norm:
                if not api_addr2_norm:
                    filtered_matches.append((idx, api_row))
            else:
                # If Address Line 2 has a value in input, only match with non-empty address_2
                # This ensures we don't match locations without Address Line 2 when input has one
                if api_addr2_norm:
                    if fuzzy_match_score(addr_line2_norm, api_addr2_norm, threshold=0.7):
                        filtered_matches.append((idx, api_row))
                # If api_addr2_norm is empty but input has Address Line 2, don't add to filtered_matches
                # (this location will be excluded)
        
        # Always use filtered_matches (even if empty) to ensure strict Address Line 2 matching
            matches = filtered_matches
        
        # Step 3: Filter by City if available
        if city_norm:
            filtered_matches = []
            for idx, api_row in matches:
                api_city = api_row.get('city', '')
                api_city_norm = normalize_string(api_city)
                
                if fuzzy_match_score(city_norm, api_city_norm, threshold=0.8):
                    filtered_matches.append((idx, api_row))
            
            if filtered_matches:
                matches = filtered_matches
        
        # Step 4: Filter by State if available
        if state_norm:
            filtered_matches = []
            for idx, api_row in matches:
                api_state = api_row.get('state', '')
                api_state_norm = normalize_string(api_state)
                
                if fuzzy_match_score(state_norm, api_state_norm, threshold=0.8):
                    filtered_matches.append((idx, api_row))
            
            if filtered_matches:
                matches = filtered_matches
        
        # Step 5: Filter by ZIP if available
        if zip_norm:
            filtered_matches = []
            for idx, api_row in matches:
                api_zip = api_row.get('zip', '')
                api_zip_norm = normalize_string(api_zip)
                
                if fuzzy_match_score(zip_norm, api_zip_norm, threshold=0.8):
                    filtered_matches.append((idx, api_row))
            
            if filtered_matches:
                matches = filtered_matches
    
    # Extract monolith_location_id, location_id, location_type, practice_monolith_id, and practice_cloud_id from matches
    result = []
    for idx, api_row in matches:
        monolith_id = api_row.get('monolith_location_id')
        cloud_id = api_row.get('location_id')
        location_type = api_row.get('Location Type', '')
        practice_monolith_id = api_row.get('Practice Monolith ID', '')
        practice_cloud_id = api_row.get('Practice Cloud ID', '')
        
        if pd.notna(monolith_id) and pd.notna(cloud_id):
            monolith_id_str = str(monolith_id).strip()
            monolith_id_str = re.sub(r'\.0$', '', monolith_id_str)
            location_type_str = str(location_type).strip() if pd.notna(location_type) else ''
            practice_monolith_id_str = str(practice_monolith_id).strip() if pd.notna(practice_monolith_id) and practice_monolith_id != '' else ''
            practice_cloud_id_str = str(practice_cloud_id).strip() if pd.notna(practice_cloud_id) and practice_cloud_id != '' else ''
            result.append((monolith_id_str, cloud_id, location_type_str, practice_monolith_id_str, practice_cloud_id_str))
    
    return result


def add_virtual_locations_after_matching(locations_df, api_locations_df):
    """
    Add virtual locations (address_1 = '--') to rows with Location Type 'Both' or 'Virtual'.
    This function runs after all matching is complete.
    
    Args:
        locations_df: DataFrame from Locations_input.xlsx
        api_locations_df: DataFrame from Practice_Locations.xlsx Locations sheet
    """
    # Check required columns exist
    if 'Location Type' not in locations_df.columns:
        print("Warning: 'Location Type' column not found in Locations_input.xlsx. Skipping virtual location addition.")
        return
    
    if 'State' not in locations_df.columns:
        print("Warning: 'State' column not found in Locations_input.xlsx. Skipping virtual location addition.")
        return
    
    # Check if required columns exist in api_locations_df
    if 'state' not in api_locations_df.columns:
        print("Warning: 'state' column not found in Practice_Locations.xlsx. Skipping virtual location addition.")
        return
    
    if 'address_1' not in api_locations_df.columns:
        print("Warning: 'address_1' column not found in Practice_Locations.xlsx. Skipping virtual location addition.")
        return
    
    if 'location_id' not in api_locations_df.columns:
        print("Warning: 'location_id' column not found in Practice_Locations.xlsx. Skipping virtual location addition.")
        return
    
    # Get all Location Cloud ID columns
    all_location_cloud_id_columns = [
        col for col in locations_df.columns 
        if col.startswith('Location Cloud ID')
    ]
    
    # Check if Location Cloud ID 1 exists (or any Location Cloud ID column)
    has_location_cloud_id_1 = 'Location Cloud ID 1' in locations_df.columns or 'Location Cloud ID' in locations_df.columns
    
    if not (has_location_cloud_id_1 or all_location_cloud_id_columns):
        print("Warning: No Location Cloud ID columns found. Skipping virtual location addition.")
        return
    
    # Process each row
    rows_processed = 0
    for idx in locations_df.index:
        # Check Location Type - must be 'Both' or 'Virtual', NOT 'In Person'
        location_type_value = locations_df.at[idx, 'Location Type']
        if pd.isna(location_type_value):
            continue
        
        location_type_str = str(location_type_value).strip()
        
        # Only process if Location Type is 'Both' or 'Virtual' (explicitly NOT 'In Person')
        if location_type_str not in ['Both', 'Virtual']:
            continue
        
        # Get State value from Locations_input.xlsx
        state_value = locations_df.at[idx, 'State']
        if pd.isna(state_value) or str(state_value).strip() == '':
            continue
        
        state_str = str(state_value).strip()
        state_norm = normalize_string(state_str)
        
        if not state_norm:
            continue
        
        # Find all locations in Practice_Locations.xlsx where:
        # 1. state matches (normalized)
        # 2. address_1 = '--'
        matching_states = api_locations_df[
            api_locations_df['state'].astype(str).str.strip().apply(
                lambda x: normalize_string(x) if pd.notna(x) else ''
            ) == state_norm
        ]
        
        # Among those, find where address_1 = '--'
        virtual_locations = matching_states[
            matching_states['address_1'].astype(str).str.strip() == '--'
        ]
        
        if virtual_locations.empty:
            print(f"Row {idx}: No virtual location found for state '{state_str}' (Location Type: {location_type_str})")
            continue
        
        # Get the first virtual location's details
        virtual_location = virtual_locations.iloc[0]
        virtual_location_id = virtual_location.get('location_id')
        virtual_monolith_location_id = virtual_location.get('monolith_location_id')
        virtual_location_type = virtual_location.get('Location Type', '')
        virtual_practice_cloud_id = virtual_location.get('Practice Cloud ID', '')
        
        if pd.isna(virtual_location_id):
            continue
        
        virtual_location_id_str = str(virtual_location_id).strip()
        virtual_location_id_str = re.sub(r'\.0$', '', virtual_location_id_str)
        
        # Process monolith_location_id
        virtual_monolith_id_str = ''
        if pd.notna(virtual_monolith_location_id):
            virtual_monolith_id_str = str(virtual_monolith_location_id).strip()
            virtual_monolith_id_str = re.sub(r'\.0$', '', virtual_monolith_id_str)
        
        # Process location_type
        virtual_location_type_str = ''
        if pd.notna(virtual_location_type):
            virtual_location_type_str = str(virtual_location_type).strip()
        
        # Process practice_cloud_id
        virtual_practice_cloud_id_str = ''
        if pd.notna(virtual_practice_cloud_id) and virtual_practice_cloud_id != '':
            virtual_practice_cloud_id_str = str(virtual_practice_cloud_id).strip()
        
        # Check if this virtual location_id is already in any Location Cloud ID column for this row
        virtual_already_exists = False
        for col in locations_df.columns:
            if col.startswith('Location Cloud ID'):
                existing_value = locations_df.at[idx, col]
                if pd.notna(existing_value):
                    existing_value_str = str(existing_value).strip()
                    existing_value_str = re.sub(r'\.0$', '', existing_value_str)
                    if existing_value_str == virtual_location_id_str:
                        virtual_already_exists = True
                        break
        
        if virtual_already_exists:
            print(f"Row {idx}: Virtual location {virtual_location_id_str} already exists in this row")
            continue
        
        # Find the highest numbered Location Cloud ID column that has a value in THIS ROW
        max_number_for_row = 0
        for col in all_location_cloud_id_columns:
            cell_value = locations_df.at[idx, col]
            if pd.notna(cell_value) and str(cell_value).strip() != '':
                # This column has a value in this row
                if col == 'Location Cloud ID':
                    max_number_for_row = max(max_number_for_row, 1)
                else:
                    # Extract number from "Location Cloud ID 1", "Location Cloud ID 2", etc.
                    number_str = col.replace('Location Cloud ID', '').strip()
                    if number_str.isdigit():
                        max_number_for_row = max(max_number_for_row, int(number_str))
        
        # Create new column number
        new_column_number = max_number_for_row + 1
        
        # Column names to create
        location_id_col_name = f'Location ID {new_column_number}'
        location_cloud_id_col_name = f'Location Cloud ID {new_column_number}'
        location_type_col_name = f'Location Type {new_column_number}'
        practice_cloud_id_col_name = f'Practice Cloud ID {new_column_number}'
        
        # Find where to insert (after the last Location Cloud ID column)
        insert_idx = len(locations_df.columns)
        if all_location_cloud_id_columns:
            # Sort columns by number
            sorted_cols = sorted(
                all_location_cloud_id_columns,
                key=lambda x: (
                    int(x.replace('Location Cloud ID', '').strip()) 
                    if x.replace('Location Cloud ID', '').strip().isdigit() 
                    else 0
                )
            )
            last_col = sorted_cols[-1]
            insert_idx = locations_df.columns.get_loc(last_col) + 1
        
        # Create columns in the correct order: Location ID, Location Cloud ID, Practice Cloud ID, Location Type
        # Create Location ID column if it doesn't exist
        if location_id_col_name not in locations_df.columns:
            locations_df.insert(insert_idx, location_id_col_name, [''] * len(locations_df))
            insert_idx += 1
            print(f"Created new column: {location_id_col_name}")
        
        # Create Location Cloud ID column if it doesn't exist
        if location_cloud_id_col_name not in locations_df.columns:
            locations_df.insert(insert_idx, location_cloud_id_col_name, [''] * len(locations_df))
            insert_idx += 1
            print(f"Created new column: {location_cloud_id_col_name}")
        
        # Create Practice Cloud ID column if it doesn't exist
        if practice_cloud_id_col_name not in locations_df.columns:
            locations_df.insert(insert_idx, practice_cloud_id_col_name, [''] * len(locations_df))
            insert_idx += 1
            print(f"Created new column: {practice_cloud_id_col_name}")
        
        # Create Location Type column if it doesn't exist
        if location_type_col_name not in locations_df.columns:
            locations_df.insert(insert_idx, location_type_col_name, [''] * len(locations_df))
            print(f"Created new column: {location_type_col_name}")
        
        # Populate all columns for this row
        locations_df.at[idx, location_id_col_name] = virtual_monolith_id_str
        locations_df.at[idx, location_cloud_id_col_name] = virtual_location_id_str
        locations_df.at[idx, location_type_col_name] = virtual_location_type_str
        locations_df.at[idx, practice_cloud_id_col_name] = virtual_practice_cloud_id_str
        
        rows_processed += 1
        print(f"Row {idx}: Added virtual location {virtual_location_id_str} (state: {state_str}, Location Type: {location_type_str}) to columns {location_id_col_name}, {location_cloud_id_col_name}, {location_type_col_name}, {practice_cloud_id_col_name}")
    
    print(f"Post-processing complete: Added virtual locations to {rows_processed} row(s)")


def add_location_cloud_ids_to_locations_input(locations_input_file, practice_locations_file):
    """
    Add Location Cloud ID column to Locations_input.xlsx by matching Location IDs
    with the location data from Practice_Locations.xlsx.
    
    Cases:
    - Case 1: Location Monolith ID (Location ID n) is already mapped
      - Takes values from Location ID 1, Location ID 2, etc.
      - Searches in Practice_Locations.xlsx by monolith_location_id
      - Extracts location_id, Practice Cloud ID, and Location Type
      - Creates Location Cloud ID n, Practice Cloud ID n, Location Type n columns
      - Renames Location ID n to Location Monolith ID n after mapping
    
    - Case 2: Only State is mapped (State-only matching)
      - Takes State values (can be comma-separated)
      - Searches in Practice_Locations.xlsx by state
      - Creates Location Monolith ID n, Location Cloud ID n, Location Type n columns
    
    - Case 3: Location ID is empty, all address components mapped separately
      - Address Line 1, Address Line 2, City, State, ZIP are all mapped separately
      - Uses fuzzy matching on individual address components
      - Creates Location ID n, Location Cloud ID n, Location Type n columns
    
    - Case 4: Location ID is empty, only Address Line 1 is mapped (contains all address info)
      - Only Address Line 1 is mapped (Address Line 2, City, State, ZIP are empty)
      - Concatenates address_1, address_2, city, state, zip from Practice_Locations.xlsx
      - Matches full concatenated address against Address Line 1 using fuzzy matching
      - Creates Location ID n, Location Cloud ID n, Location Type n columns
    
    Args:
        locations_input_file: Path to Locations_input.xlsx
        practice_locations_file: Path to Practice_Locations.xlsx (should have Locations sheet)
    """
    locations_input_path = Path(locations_input_file)
    practice_locations_path = Path(practice_locations_file)
    
    if not locations_input_path.exists():
        print(f"Warning: Locations_input.xlsx not found at {locations_input_path}")
        return
    
    if not practice_locations_path.exists():
        print(f"Warning: Practice_Locations.xlsx not found at {practice_locations_path}")
        return
    
    try:
        # Read Locations_input.xlsx
        locations_df = pd.read_excel(locations_input_path, sheet_name='Locations')
        
        # Check if there are any Location ID columns
        location_id_columns = [col for col in locations_df.columns if col.startswith('Location ID')]
        if not location_id_columns:
            print("No Location ID columns found in Locations_input.xlsx. Skipping Location Cloud ID mapping.")
            return
        
        # Try to read Locations sheet from Practice_Locations.xlsx
        try:
            api_locations_df = pd.read_excel(practice_locations_path, sheet_name='Locations')
        except Exception as e:
            print(f"Warning: Could not read Locations sheet from Practice_Locations.xlsx: {str(e)}")
            print("Location Cloud ID will not be added.")
            return
        
        # Check if required columns exist in API locations
        if 'monolith_location_id' not in api_locations_df.columns or 'location_id' not in api_locations_df.columns:
            print("Warning: Required columns (monolith_location_id, location_id) not found in Locations sheet.")
            print("Location Cloud ID will not be added.")
            return
        
        # Check if Location ID columns are empty (need fuzzy matching)
        # We need fuzzy matching only if ALL Location ID columns are empty or mostly empty
        # If at least one Location ID column has values, we use direct mapping
        needs_fuzzy_matching = True  # Start with True, will be set to False if any column has values
        
        for location_id_col in location_id_columns:
            if location_id_col in locations_df.columns:
                # Check if column has any non-empty values
                non_empty_mask = locations_df[location_id_col].notna() & (locations_df[location_id_col].astype(str).str.strip() != '')
                non_empty_count = non_empty_mask.sum()
                total_count = len(locations_df)
                
                # If more than 20% of values are non-empty, this column has values
                if total_count > 0 and (non_empty_count / total_count) > 0.2:
                    needs_fuzzy_matching = False  # At least one column has values, so we'll use direct mapping
                    break  # Found at least one column with values, no need to check others
        
        # Case 2: Only State is mapped (State-only matching)
        # Check if State column exists and has values, and other address components are empty/not mapped
        # IMPORTANT: Case 2 should ONLY be used if State is the ONLY address element mapped
        # If ANY other address component (Address Line 1, Address Line 2, City, ZIP) has values, use Case 3 instead
        has_state = 'State' in locations_df.columns
        has_address_line1 = 'Address Line 1' in locations_df.columns
        has_address_line2 = 'Address Line 2' in locations_df.columns
        has_city = 'City' in locations_df.columns
        has_zip = 'ZIP' in locations_df.columns
        
        # Check if State has values and other address components are empty
        only_state_mapped = False
        if has_state and needs_fuzzy_matching:
            # Check if State column has values
            state_has_values = not locations_df['State'].isna().all() and not (locations_df['State'].astype(str).str.strip() == '').all()
            
            if not state_has_values:
                only_state_mapped = False
            else:
                # Check if ANY other address component has values (not just if column exists)
                # If ANY address component has values, it's NOT Case 2
                addr1_has_values = False
                addr2_has_values = False
                city_has_values = False
                zip_has_values = False
                
                if has_address_line1:
                    addr1_has_values = not locations_df['Address Line 1'].isna().all() and not (locations_df['Address Line 1'].astype(str).str.strip() == '').all()
                if has_address_line2:
                    addr2_has_values = not locations_df['Address Line 2'].isna().all() and not (locations_df['Address Line 2'].astype(str).str.strip() == '').all()
                if has_city:
                    city_has_values = not locations_df['City'].isna().all() and not (locations_df['City'].astype(str).str.strip() == '').all()
                if has_zip:
                    zip_has_values = not locations_df['ZIP'].isna().all() and not (locations_df['ZIP'].astype(str).str.strip() == '').all()
                
                # Only State is mapped if State has values AND NO other address components have values
                # If ANY other address component has values, it's Case 3, not Case 2
                only_state_mapped = state_has_values and not addr1_has_values and not addr2_has_values and not city_has_values and not zip_has_values
        
        if only_state_mapped:
            # Case 2: Only State is mapped (State-only matching)
            print("Case 2: Only State is mapped. Matching by State and populating Location ID columns...")
            
            # Check if 'state' column exists in API locations
            if 'state' not in api_locations_df.columns:
                print("Warning: 'state' column not found in Practice_Locations.xlsx Locations sheet.")
                print("Cannot match by State.")
                return
            
            # Check if 'location_id' column exists in API locations
            if 'location_id' not in api_locations_df.columns:
                print("Warning: 'location_id' column not found in Practice_Locations.xlsx Locations sheet.")
                print("Cannot populate Location ID.")
                return
            
            # Process each row in Locations_input.xlsx
            matched_count = 0
            max_location_id_columns = 1  # Track maximum number of Location ID columns needed
            
            for idx, row in locations_df.iterrows():
                state_value = row.get('State', '')
                
                # Skip if State is empty
                if pd.isna(state_value) or str(state_value).strip() == '':
                    continue
                
                state_value_str = str(state_value).strip()
                
                # Split by comma to handle multiple states
                state_list = [s.strip() for s in state_value_str.split(',') if s.strip()]
                
                if not state_list:
                    continue
                
                # Track matches for this row (store full match info)
                matches_for_row = []
                
                # Process each state in the comma-separated list
                for state_item in state_list:
                    state_item_norm = normalize_string(state_item)
                    
                    if not state_item_norm:
                        continue
                    
                    # Find matching rows in Practice_Locations.xlsx by state
                    matching_rows = api_locations_df[
                        api_locations_df['state'].astype(str).str.strip().apply(lambda x: normalize_string(x) if pd.notna(x) else '') == state_item_norm
                    ]
                    
                    # If multiple matches, take the first one for each state
                    if not matching_rows.empty:
                        first_match = matching_rows.iloc[0]
                        location_id = first_match.get('location_id')
                        monolith_location_id = first_match.get('monolith_location_id')
                        location_type = first_match.get('Location Type', '')
                        practice_monolith_id = first_match.get('Practice Monolith ID', '')
                        practice_cloud_id = first_match.get('Practice Cloud ID', '')
                        
                        if pd.notna(location_id):
                            # Convert to string and remove .0 suffix if present
                            location_id_str = str(location_id).strip()
                            location_id_str = re.sub(r'\.0$', '', location_id_str)
                            
                            # Process monolith_location_id
                            monolith_id_str = ''
                            if pd.notna(monolith_location_id):
                                monolith_id_str = str(monolith_location_id).strip()
                                monolith_id_str = re.sub(r'\.0$', '', monolith_id_str)
                            
                            # Process location_type
                            location_type_str = ''
                            if pd.notna(location_type):
                                location_type_str = str(location_type).strip()
                            
                            # Process practice_monolith_id
                            practice_monolith_id_str = ''
                            if pd.notna(practice_monolith_id) and practice_monolith_id != '':
                                practice_monolith_id_str = str(practice_monolith_id).strip()
                                practice_monolith_id_str = re.sub(r'\.0$', '', practice_monolith_id_str)
                            
                            # Process practice_cloud_id
                            practice_cloud_id_str = ''
                            if pd.notna(practice_cloud_id) and practice_cloud_id != '':
                                practice_cloud_id_str = str(practice_cloud_id).strip()
                            
                            matches_for_row.append({
                                'location_id': location_id_str,
                                'monolith_location_id': monolith_id_str,
                                'location_type': location_type_str,
                                'practice_monolith_id': practice_monolith_id_str,
                                'practice_cloud_id': practice_cloud_id_str
                            })
                
                # Populate Location Monolith ID, Location Cloud ID, Location Type, Practice Monolith ID n, and Practice Cloud ID n columns
                # Order: Location Monolith ID {n}, Location Cloud ID {n}, Location Type {n}, Practice Monolith ID {n}, Practice Cloud ID {n} for each n
                if matches_for_row:
                    # First, create all columns in the correct order if they don't exist
                    # We need to create columns for all match numbers up to max_location_id_columns
                    # But we'll create them as we process matches
                    for i, match_info in enumerate(matches_for_row, start=1):
                        location_monolith_id_col = f'Location Monolith ID {i}'
                        location_cloud_id_col = f'Location Cloud ID {i}'
                        location_type_col = f'Location Type {i}'
                        practice_monolith_id_col = f'Practice Monolith ID {i}'
                        practice_cloud_id_col = f'Practice Cloud ID {i}'
                        
                        # Determine insertion position
                        insert_idx = len(locations_df.columns)
                        
                        # Find where to insert: after the last Location Type column of the previous match, or after Location ID columns
                        if i == 1:
                            # For first match, find position after Location ID columns (if any)
                            existing_location_id_cols = [col for col in locations_df.columns if col.startswith('Location ID') and not col.startswith('Location Cloud ID') and not col.startswith('Location Monolith ID')]
                            if existing_location_id_cols:
                                # Sort to get the last one
                                sorted_cols = sorted(existing_location_id_cols, 
                                                   key=lambda x: int(x.split()[-1]) if x.split()[-1].isdigit() else 0)
                                last_col = sorted_cols[-1]
                                insert_idx = locations_df.columns.get_loc(last_col) + 1
                        else:
                            # For subsequent matches, insert after Location Type of previous match
                            prev_location_type_col = f'Location Type {i-1}'
                            if prev_location_type_col in locations_df.columns:
                                insert_idx = locations_df.columns.get_loc(prev_location_type_col) + 1
                        
                        # Create Location Monolith ID column if it doesn't exist
                        if location_monolith_id_col not in locations_df.columns:
                            locations_df.insert(insert_idx, location_monolith_id_col, [''] * len(locations_df))
                            insert_idx += 1
                        else:
                            insert_idx = locations_df.columns.get_loc(location_monolith_id_col) + 1
                        
                        # Create Location Cloud ID column if it doesn't exist (insert after Location Monolith ID)
                        if location_cloud_id_col not in locations_df.columns:
                            locations_df.insert(insert_idx, location_cloud_id_col, [''] * len(locations_df))
                            insert_idx += 1
                        else:
                            insert_idx = locations_df.columns.get_loc(location_cloud_id_col) + 1
                        
                        # Create Location Type column if it doesn't exist (insert after Location Cloud ID)
                        if location_type_col not in locations_df.columns:
                            locations_df.insert(insert_idx, location_type_col, [''] * len(locations_df))
                            insert_idx += 1
                        else:
                            insert_idx = locations_df.columns.get_loc(location_type_col) + 1
                        
                        # Create Practice Monolith ID column if it doesn't exist (insert after Location Type)
                        if practice_monolith_id_col not in locations_df.columns:
                            locations_df.insert(insert_idx, practice_monolith_id_col, [''] * len(locations_df))
                            insert_idx += 1
                        else:
                            insert_idx = locations_df.columns.get_loc(practice_monolith_id_col) + 1
                        
                        # Create Practice Cloud ID column if it doesn't exist (insert after Practice Monolith ID)
                        if practice_cloud_id_col not in locations_df.columns:
                            locations_df.insert(insert_idx, practice_cloud_id_col, [''] * len(locations_df))
                        
                        # Populate the values
                        locations_df.at[idx, location_monolith_id_col] = match_info['monolith_location_id']
                        locations_df.at[idx, location_cloud_id_col] = match_info['location_id']
                        # Case 2 exception: Location Type column should be empty for State-only matching
                        locations_df.at[idx, location_type_col] = ''  # Empty for Case 2
                        locations_df.at[idx, practice_monolith_id_col] = match_info.get('practice_monolith_id', '')
                        locations_df.at[idx, practice_cloud_id_col] = match_info.get('practice_cloud_id', '')
                        matched_count += 1
                    
                    # Case 2 exception: Also clear the main 'Location Type' column (not numbered) for this row
                    if 'Location Type' in locations_df.columns:
                        locations_df.at[idx, 'Location Type'] = ''
                    
                    # Replace address values with values from the first matched location
                    if matches_for_row and len(matches_for_row) > 0:
                        first_match_cloud_id = matches_for_row[0].get('location_id', '')
                        if first_match_cloud_id:
                            replace_address_values_from_matched_location(locations_df, api_locations_df, idx, first_match_cloud_id)
                    
                    # Update max_location_id_columns if needed
                    if len(matches_for_row) > max_location_id_columns:
                        max_location_id_columns = len(matches_for_row)
            
            print(f"State-only matching completed. Matched {matched_count} location(s) across up to {max_location_id_columns} Location Cloud ID column(s).")
            
            # Save updated Locations_input.xlsx
            with pd.ExcelWriter(locations_input_path, engine='openpyxl', mode='w') as writer:
                locations_df.to_excel(writer, index=False, sheet_name='Locations')
            
            # Reapply styling
            wb = load_workbook(locations_input_path)
            ws = wb['Locations']
            
            # Style header row
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(locations_input_path)
            
            print(f"Added Location Cloud ID column(s) (up to Location Cloud ID {max_location_id_columns}) to Locations_input.xlsx based on State matching")
            
            # Case 2 is done - mark that we don't need fuzzy matching (to prevent Cases 3/4 from running)
            needs_fuzzy_matching = False
        
        if needs_fuzzy_matching:
            # Check which fuzzy matching case we're in
            # Case 3: All address components mapped separately
            # Case 4: Only Address Line 1 is mapped (contains all address info)
            
            has_address_line1 = 'Address Line 1' in locations_df.columns
            has_address_line2 = 'Address Line 2' in locations_df.columns
            has_city = 'City' in locations_df.columns
            has_state = 'State' in locations_df.columns
            has_zip = 'ZIP' in locations_df.columns
            
            # Check if only Address Line 1 is mapped (Case 4)
            addr1_has_values = False
            addr2_empty = True
            city_empty = True
            state_empty = True
            zip_empty = True
            
            if has_address_line1:
                addr1_has_values = not locations_df['Address Line 1'].isna().all() and not (locations_df['Address Line 1'].astype(str).str.strip() == '').all()
            if has_address_line2:
                addr2_empty = locations_df['Address Line 2'].isna().all() or (locations_df['Address Line 2'].astype(str).str.strip() == '').all()
            if has_city:
                city_empty = locations_df['City'].isna().all() or (locations_df['City'].astype(str).str.strip() == '').all()
            if has_state:
                state_empty = locations_df['State'].isna().all() or (locations_df['State'].astype(str).str.strip() == '').all()
            if has_zip:
                zip_empty = locations_df['ZIP'].isna().all() or (locations_df['ZIP'].astype(str).str.strip() == '').all()
            
            only_address_line1_mapped = addr1_has_values and addr2_empty and city_empty and state_empty and zip_empty
            
            if only_address_line1_mapped:
                # Case 4: Only Address Line 1 is mapped (contains all address info)
                print("Case 4: Only Address Line 1 is mapped. Using concatenated address matching...")
            else:
                # Case 3: All address components mapped separately
                print("Case 3: All address components mapped separately. Using fuzzy matching on individual components...")
            
            # Remove existing Location ID and Location Cloud ID columns if they exist (we'll rename them)
            if 'Location ID' in locations_df.columns:
                locations_df = locations_df.drop(columns=['Location ID'])
            if 'Location Cloud ID' in locations_df.columns:
                locations_df = locations_df.drop(columns=['Location Cloud ID'])
            
            # Check required columns for fuzzy matching
            # Need Address Line 1 (Practice ID is optional - if not provided, will match against all locations)
            has_practice_id = 'Practice ID' in locations_df.columns
            has_practice_monolith_id = 'Practice Monolith ID' in locations_df.columns
            has_practice_cloud_id = 'Practice Cloud ID' in locations_df.columns
            
            if not has_address_line1:
                print("Warning: Required column 'Address Line 1' for fuzzy matching not found.")
                print("Location Cloud ID will not be added via fuzzy matching.")
                return
            
            # Process each row
            max_matches = 1  # Track maximum number of matches found for any row
            all_location_ids = [[] for _ in range(len(locations_df))]
            all_cloud_ids = [[] for _ in range(len(locations_df))]
            all_location_types = [[] for _ in range(len(locations_df))]
            all_practice_monolith_ids = [[] for _ in range(len(locations_df))]
            all_practice_cloud_ids = [[] for _ in range(len(locations_df))]
            
            for idx, row in locations_df.iterrows():
                # Check for Practice ID, Practice Monolith ID, or Practice Cloud ID (optional)
                practice_id = row.get('Practice ID') or row.get('Practice Monolith ID')
                practice_cloud_id = row.get('Practice Cloud ID')
                
                # Check if IDs are available (optional - if not available, will match against all locations)
                practice_id_valid = practice_id is not None and pd.notna(practice_id) and str(practice_id).strip() != ''
                practice_cloud_id_valid = practice_cloud_id is not None and pd.notna(practice_cloud_id) and str(practice_cloud_id).strip() != ''
                
                # Match location by address (pass IDs if available, otherwise None to match against all locations)
                matches = match_location_by_address(
                    row, 
                    api_locations_df, 
                    practice_id=practice_id if practice_id_valid else None,
                    practice_cloud_id=practice_cloud_id if practice_cloud_id_valid else None
                )
                
                if matches:
                    for monolith_id, cloud_id, location_type, practice_monolith_id, practice_cloud_id in matches:
                        all_location_ids[idx].append(monolith_id)
                        all_cloud_ids[idx].append(cloud_id)
                        all_location_types[idx].append(location_type)
                        all_practice_monolith_ids[idx].append(practice_monolith_id)
                        all_practice_cloud_ids[idx].append(practice_cloud_id)
                    
                    if len(matches) > max_matches:
                        max_matches = len(matches)
            
            # Create columns for all possible matches (up to max_matches)
            for match_num in range(1, max_matches + 1):
                location_id_col = f'Location ID {match_num}'
                cloud_id_col = f'Location Cloud ID {match_num}'
                location_type_col = f'Location Type {match_num}'
                practice_monolith_id_col = f'Practice Monolith ID {match_num}'
                practice_cloud_id_col = f'Practice Cloud ID {match_num}'
                
                if location_id_col not in locations_df.columns:
                    locations_df[location_id_col] = [''] * len(locations_df)
                if cloud_id_col not in locations_df.columns:
                    locations_df[cloud_id_col] = [''] * len(locations_df)
                if location_type_col not in locations_df.columns:
                    locations_df[location_type_col] = [''] * len(locations_df)
                if practice_monolith_id_col not in locations_df.columns:
                    locations_df[practice_monolith_id_col] = [''] * len(locations_df)
                if practice_cloud_id_col not in locations_df.columns:
                    locations_df[practice_cloud_id_col] = [''] * len(locations_df)
            
            # Fill Location ID, Location Cloud ID, Location Type, Practice Monolith ID n, and Practice Cloud ID n columns
            for idx in range(len(locations_df)):
                if all_location_ids[idx]:
                    for match_idx in range(len(all_location_ids[idx])):
                        col_num = match_idx + 1
                        location_id_col = f'Location ID {col_num}'
                        cloud_id_col = f'Location Cloud ID {col_num}'
                        location_type_col = f'Location Type {col_num}'
                        practice_monolith_id_col = f'Practice Monolith ID {col_num}'
                        practice_cloud_id_col = f'Practice Cloud ID {col_num}'
                        
                        locations_df.at[idx, location_id_col] = all_location_ids[idx][match_idx]
                        locations_df.at[idx, cloud_id_col] = all_cloud_ids[idx][match_idx]
                        locations_df.at[idx, location_type_col] = all_location_types[idx][match_idx]
                        if all_practice_monolith_ids[idx] and match_idx < len(all_practice_monolith_ids[idx]):
                            locations_df.at[idx, practice_monolith_id_col] = all_practice_monolith_ids[idx][match_idx]
                        if all_practice_cloud_ids[idx] and match_idx < len(all_practice_cloud_ids[idx]):
                            locations_df.at[idx, practice_cloud_id_col] = all_practice_cloud_ids[idx][match_idx]
                    
                    # Replace address values with values from the first matched location
                    if all_cloud_ids[idx] and len(all_cloud_ids[idx]) > 0:
                        first_match_cloud_id = all_cloud_ids[idx][0]
                        if first_match_cloud_id:
                            replace_address_values_from_matched_location(locations_df, api_locations_df, idx, first_match_cloud_id)
            
            print(f"Fuzzy matching completed. Found matches for locations.")
            
        else:
            # Case 1: Direct mapping - Location Monolith ID is already mapped
            print("Case 1: Location Monolith ID columns are mapped. Using direct mapping...")
            
            # Create mapping from monolith_location_id to location_id (cloud ID), practice_monolith_id, practice_cloud_id, location_type, and address fields
            location_id_map = {}
            practice_monolith_id_map = {}
            practice_cloud_id_map = {}
            location_type_map = {}
            address_line1_map = {}
            address_line2_map = {}
            city_map = {}
            state_map = {}
            zip_map = {}
            
            for _, row in api_locations_df.iterrows():
                monolith_id = str(row['monolith_location_id']).strip() if pd.notna(row['monolith_location_id']) else None
                cloud_id = row['location_id'] if pd.notna(row['location_id']) else None
                practice_monolith_id = row.get('Practice Monolith ID', '') if 'Practice Monolith ID' in row else ''
                practice_cloud_id = row.get('Practice Cloud ID', '') if 'Practice Cloud ID' in row else ''
                location_type = row.get('Location Type', '') if 'Location Type' in row else ''
                
                # Extract address fields
                address_1 = row.get('address_1', '') if 'address_1' in row else ''
                address_2 = row.get('address_2', '') if 'address_2' in row else ''
                city = row.get('city', '') if 'city' in row else ''
                state = row.get('state', '') if 'state' in row else ''
                zip_code = row.get('zip', '') if 'zip' in row else ''
                
                if monolith_id and cloud_id:
                    monolith_id = re.sub(r'\.0$', '', monolith_id)
                    if monolith_id not in location_id_map:
                        location_id_map[monolith_id] = cloud_id
                        practice_monolith_id_map[monolith_id] = str(practice_monolith_id).strip() if pd.notna(practice_monolith_id) and practice_monolith_id != '' else ''
                        practice_cloud_id_map[monolith_id] = str(practice_cloud_id).strip() if pd.notna(practice_cloud_id) and practice_cloud_id != '' else ''
                        location_type_map[monolith_id] = str(location_type).strip() if pd.notna(location_type) else ''
            
                        # Store address fields (convert to string and strip, handle NaN)
                        address_line1_map[monolith_id] = str(address_1).strip() if pd.notna(address_1) and address_1 != '' else ''
                        address_line2_map[monolith_id] = str(address_2).strip() if pd.notna(address_2) and address_2 != '' else ''
                        city_map[monolith_id] = str(city).strip() if pd.notna(city) and city != '' else ''
                        state_map[monolith_id] = str(state).strip() if pd.notna(state) and state != '' else ''
                        # Special handling for ZIP to preserve leading zeros
                        if pd.notna(zip_code) and zip_code != '':
                            zip_value = zip_code
                            if isinstance(zip_value, (int, float)):
                                zip_map[monolith_id] = str(int(zip_value)).zfill(5)
                            else:
                                zip_str = str(zip_value).strip()
                                if zip_str.isdigit() and len(zip_str) < 5:
                                    zip_map[monolith_id] = zip_str.zfill(5)
                                else:
                                    zip_map[monolith_id] = zip_str
                        else:
                            zip_map[monolith_id] = ''
            
            print(f"Created mapping for {len(location_id_map)} Location Monolith IDs to Cloud IDs, Practice Cloud IDs, Location Types, and Address fields")
            
            # Add Location Cloud ID, Practice Cloud ID, Practice Monolith ID, and Location Type columns for each Location ID column
            for location_id_col in location_id_columns:
                # Create corresponding column names
                if location_id_col == 'Location ID':
                    cloud_id_col_name = 'Location Cloud ID'
                    practice_cloud_id_col_name = 'Practice Cloud ID'
                    practice_monolith_id_col_name = 'Practice Monolith ID'
                    location_type_col_name = 'Location Type'
                else:
                    # Extract number from "Location ID 1" -> "Location Cloud ID 1", "Practice Cloud ID 1", "Practice Monolith ID 1", "Location Type 1"
                    # Handle both "Location ID 1" and "Location ID1" formats
                    number = location_id_col.replace('Location ID', '').strip()
                    # Ensure proper spacing: "Location Cloud ID 1" not "Location Cloud ID1"
                    if number:
                        cloud_id_col_name = f'Location Cloud ID {number}'
                        practice_cloud_id_col_name = f'Practice Cloud ID {number}'
                        practice_monolith_id_col_name = f'Practice Monolith ID {number}'
                        location_type_col_name = f'Location Type {number}'
                    else:
                        cloud_id_col_name = 'Location Cloud ID'
                        practice_cloud_id_col_name = 'Practice Cloud ID'
                        practice_monolith_id_col_name = 'Practice Monolith ID'
                        location_type_col_name = 'Location Type'
                
                # If Location Cloud ID column doesn't exist, create it (insert after Location ID column)
                if cloud_id_col_name not in locations_df.columns:
                    location_id_col_idx = locations_df.columns.get_loc(location_id_col)
                    locations_df.insert(location_id_col_idx + 1, cloud_id_col_name, [''] * len(locations_df))
                
                # If Location Type column doesn't exist, create it (insert after Location Cloud ID)
                if location_type_col_name not in locations_df.columns:
                    cloud_id_col_idx = locations_df.columns.get_loc(cloud_id_col_name)
                    locations_df.insert(cloud_id_col_idx + 1, location_type_col_name, [''] * len(locations_df))
                
                # If Practice Monolith ID column doesn't exist, create it (insert after Location Type)
                if practice_monolith_id_col_name not in locations_df.columns:
                    location_type_col_idx = locations_df.columns.get_loc(location_type_col_name)
                    locations_df.insert(location_type_col_idx + 1, practice_monolith_id_col_name, [''] * len(locations_df))
                
                # If Practice Cloud ID column doesn't exist, create it (insert after Practice Monolith ID)
                if practice_cloud_id_col_name not in locations_df.columns:
                    practice_monolith_id_col_idx = locations_df.columns.get_loc(practice_monolith_id_col_name)
                    locations_df.insert(practice_monolith_id_col_idx + 1, practice_cloud_id_col_name, [''] * len(locations_df))
                
                # Map Location Monolith IDs to Location Cloud IDs, Practice Cloud IDs, Practice Monolith IDs, and Location Types
                for idx in locations_df.index:
                    loc_id = locations_df.at[idx, location_id_col]
                    if pd.notna(loc_id) and str(loc_id).strip() != '':
                        loc_id_str = str(loc_id).strip()
                        loc_id_str = re.sub(r'\.0$', '', loc_id_str)
                        cloud_id = location_id_map.get(loc_id_str, '')
                        practice_monolith_id = practice_monolith_id_map.get(loc_id_str, '')
                        practice_cloud_id = practice_cloud_id_map.get(loc_id_str, '')
                        location_type = location_type_map.get(loc_id_str, '')
                        locations_df.at[idx, cloud_id_col_name] = cloud_id
                        locations_df.at[idx, practice_cloud_id_col_name] = practice_cloud_id
                        locations_df.at[idx, practice_monolith_id_col_name] = practice_monolith_id
                        locations_df.at[idx, location_type_col_name] = location_type
            
            # Replace address fields (Address Line 1, Address Line 2, City, State, ZIP) from matched locations
            # Replace values in Locations_input.xlsx with values from Practice_Locations.xlsx
            for idx in locations_df.index:
                # Find the first Location ID column that has a value and a match
                matched_cloud_id = None
                for location_id_col in location_id_columns:
                    loc_id = locations_df.at[idx, location_id_col]
                    if pd.notna(loc_id) and str(loc_id).strip() != '':
                        loc_id_str = str(loc_id).strip()
                        loc_id_str = re.sub(r'\.0$', '', loc_id_str)
                        
                        # Check if this Location ID has a match in our maps
                        if loc_id_str in location_id_map:
                            matched_cloud_id = location_id_map.get(loc_id_str, '')
                            break  # Found a match, use this Location ID
                
                # If we found a match, replace all address fields from that match
                if matched_cloud_id:
                    replace_address_values_from_matched_location(locations_df, api_locations_df, idx, matched_cloud_id)
            
            # Rename Location ID columns to Location Monolith ID after mapping is complete
            print("Renaming Location ID columns to Location Monolith ID...")
            rename_mapping = {}
            for location_id_col in location_id_columns:
                if location_id_col in locations_df.columns:
                    if location_id_col == 'Location ID':
                        rename_mapping[location_id_col] = 'Location Monolith ID'
                    else:
                        # Extract number from "Location ID 1" -> "Location Monolith ID 1"
                        number = location_id_col.replace('Location ID', '').strip()
                        if number:
                            rename_mapping[location_id_col] = f'Location Monolith ID {number}'
                        else:
                            rename_mapping[location_id_col] = 'Location Monolith ID'
            
            # Perform the renaming
            locations_df = locations_df.rename(columns=rename_mapping)
            print(f"Renamed {len(rename_mapping)} Location ID column(s) to Location Monolith ID")
        
        # Reorder columns to match the specified sequence
        # Order: NPI Number, First Name, Last Name, Address Line 1, Address Line 2, City, State, ZIP, Practice ID,
        #        Location Monolith ID 1, Location Cloud ID 1, Practice Cloud ID 1, Location Type 1,
        #        Location Monolith ID 2, Location Cloud ID 2, Practice Cloud ID 2, Location Type 2, etc.
        # Note: PFS is NOT included in Locations_input.xlsx
        
        # Define the base column order
        base_columns = [
            'NPI Number',
            'First Name',
            'Last Name',
            'Address Line 1',
            'Address Line 2',
            'City',
            'State',
            'ZIP',
            'Practice ID'
        ]
        
        # Find all Location Monolith ID columns and sort them by number
        # (They were renamed from Location ID columns in Case 1)
        location_id_cols = [col for col in locations_df.columns if col.startswith('Location Monolith ID') or col.startswith('Location ID')]
        location_id_cols_sorted = sorted(
            location_id_cols,
            key=lambda x: int(x.split()[-1]) if x.split()[-1].isdigit() else 999
        )
        
        # Build the final column order
        column_order = []
        
        # Add base columns first (only if they exist in the dataframe)
        for base_col in base_columns:
            if base_col in locations_df.columns:
                column_order.append(base_col)
        
        # Add Location Monolith ID (or Location ID), Location Cloud ID, Practice Cloud ID, and Location Type columns in sequence
        for location_id_col in location_id_cols_sorted:
            # Extract number from Location Monolith ID or Location ID column name
            if location_id_col == 'Location ID' or location_id_col == 'Location Monolith ID':
                number = ''
            elif location_id_col.startswith('Location Monolith ID'):
                number = location_id_col.replace('Location Monolith ID', '').strip()
            else:
                number = location_id_col.replace('Location ID', '').strip()
            
            # Add Location ID column
            if location_id_col not in column_order:
                column_order.append(location_id_col)
            
            # Add corresponding Location Cloud ID column
            if number:
                cloud_col = f'Location Cloud ID {number}'
            else:
                cloud_col = 'Location Cloud ID'
            
            if cloud_col in locations_df.columns and cloud_col not in column_order:
                column_order.append(cloud_col)
            
            # Add corresponding Practice Cloud ID column
            if number:
                practice_cloud_col = f'Practice Cloud ID {number}'
            else:
                practice_cloud_col = 'Practice Cloud ID'
            
            if practice_cloud_col in locations_df.columns and practice_cloud_col not in column_order:
                column_order.append(practice_cloud_col)
            
            # Add corresponding Location Type column
            if number:
                location_type_col = f'Location Type {number}'
            else:
                location_type_col = 'Location Type'
            
            if location_type_col in locations_df.columns and location_type_col not in column_order:
                column_order.append(location_type_col)
        
        # Add any remaining columns that weren't in the specified order
        for col in locations_df.columns:
            if col not in column_order:
                column_order.append(col)
        
        # Reorder the dataframe
        locations_df = locations_df[column_order]
        
        # Final reordering: Ensure columns are in the correct order (before virtual locations are added)
        print("\nFinal column reordering...")
        # Define the base column order
        base_columns = [
            'NPI Number',
            'First Name',
            'Last Name',
            'Address Line 1',
            'Address Line 2',
            'City',
            'State',
            'ZIP',
            'Practice ID'
        ]
        
        # Find all Location ID columns (both Location ID and Location Monolith ID)
        location_id_cols = [col for col in locations_df.columns if col.startswith('Location Monolith ID') or col.startswith('Location ID')]
        location_id_cols_sorted = sorted(
            location_id_cols,
            key=lambda x: int(x.split()[-1]) if x.split()[-1].isdigit() else 999
        )
        
        # Build the final column order
        final_column_order = []
        
        # Add base columns first (only if they exist in the dataframe)
        for base_col in base_columns:
            if base_col in locations_df.columns:
                final_column_order.append(base_col)
        
        # Add Location ID (or Location Monolith ID), Location Cloud ID, Practice Cloud ID, and Location Type columns in sequence
        for location_id_col in location_id_cols_sorted:
            # Extract number from Location Monolith ID or Location ID column name
            if location_id_col == 'Location ID' or location_id_col == 'Location Monolith ID':
                number = ''
            elif location_id_col.startswith('Location Monolith ID'):
                number = location_id_col.replace('Location Monolith ID', '').strip()
            else:
                number = location_id_col.replace('Location ID', '').strip()
            
            # Add Location ID column
            if location_id_col not in final_column_order:
                final_column_order.append(location_id_col)
            
            # Add corresponding Location Cloud ID column
            if number:
                cloud_col = f'Location Cloud ID {number}'
            else:
                cloud_col = 'Location Cloud ID'
            
            if cloud_col in locations_df.columns and cloud_col not in final_column_order:
                final_column_order.append(cloud_col)
            
            # Add corresponding Practice Cloud ID column
            if number:
                practice_cloud_col = f'Practice Cloud ID {number}'
            else:
                practice_cloud_col = 'Practice Cloud ID'
            
            if practice_cloud_col in locations_df.columns and practice_cloud_col not in final_column_order:
                final_column_order.append(practice_cloud_col)
            
            # Add corresponding Location Type column
            if number:
                location_type_col = f'Location Type {number}'
            else:
                location_type_col = 'Location Type'
            
            if location_type_col in locations_df.columns and location_type_col not in final_column_order:
                final_column_order.append(location_type_col)
        
        # Add any remaining columns that weren't in the specified order
        for col in locations_df.columns:
            if col not in final_column_order:
                final_column_order.append(col)
        
        # Reorder the dataframe
        locations_df = locations_df[final_column_order]
        print("Column reordering complete.")
        
        # Save updated Locations_input.xlsx
        with pd.ExcelWriter(locations_input_path, engine='openpyxl', mode='w') as writer:
            locations_df.to_excel(writer, index=False, sheet_name='Locations')
        
        # Reapply styling
        wb = load_workbook(locations_input_path)
        ws = wb['Locations']
        
        # Style header row
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Highlight Location Type n columns based on Location Type comparison
        if 'Location Type' in locations_df.columns:
            green_fill = PatternFill(start_color='ABFFAB', end_color='ABFFAB', fill_type='solid')  # Light green
            red_fill = PatternFill(start_color='FFA3A3', end_color='FFA3A3', fill_type='solid')  # Light red
            
            # Find Location Type column index
            location_type_col_idx = None
            for col_idx, col_name in enumerate(locations_df.columns, start=1):
                if col_name == 'Location Type':
                    location_type_col_idx = col_idx
                    break
            
            if location_type_col_idx:
                # Find all Location Type n columns
                location_type_n_columns = []
                for col_idx, col_name in enumerate(locations_df.columns, start=1):
                    if col_name.startswith('Location Type ') and col_name != 'Location Type':
                        # Extract the number
                        number = col_name.replace('Location Type ', '').strip()
                        if number.isdigit():
                            location_type_n_columns.append((col_idx, col_name, int(number)))
                
                # Sort by number
                location_type_n_columns.sort(key=lambda x: x[2])
                
                if location_type_n_columns:
                    print(f"Highlighting {len(location_type_n_columns)} Location Type n column(s) based on Location Type comparison...")
                    
                    # Process each row
                    for row_idx, (idx, row) in enumerate(locations_df.iterrows(), start=2):  # Start at 2 (row 1 is header)
                        location_type = row.get('Location Type', '')
                        location_type_str = str(location_type).strip() if pd.notna(location_type) else ''
                        
                        # Highlight each Location Type n column based on comparison
                        for col_idx, col_name, number in location_type_n_columns:
                            location_type_n_val = row.get(col_name, '')
                            location_type_n_str = str(location_type_n_val).strip() if pd.notna(location_type_n_val) else ''
                            
                            # Determine highlight color based on rules
                            highlight_color = None
                            
                            if location_type_str == 'Both':
                                # Both → all Location Type n should be green
                                highlight_color = green_fill
                            elif location_type_str == 'In Person':
                                # In Person → green if Location Type n is 'In Person', red otherwise
                                if location_type_n_str == 'In Person':
                                    highlight_color = green_fill
                                else:
                                    highlight_color = red_fill
                            elif location_type_str == 'Virtual':
                                # Virtual → green if Location Type n is 'Virtual', red otherwise
                                if location_type_n_str == 'Virtual':
                                    highlight_color = green_fill
                                else:
                                    highlight_color = red_fill
                            
                            # Apply highlighting if color determined
                            if highlight_color:
                                cell = ws.cell(row=row_idx, column=col_idx)
                                cell.fill = highlight_color
                    
                    print(f"Highlighted Location Type n columns based on Location Type comparison")
        
        # Add vertical borders to group Location ID n, Location Cloud ID n, Practice Cloud ID n, and Location Type n columns
        print("Adding vertical borders to group location columns...")
        # Create a right border style
        right_border = Border(
            right=Side(style='thin', color='000000')
        )
        
        # Find all Location Type n columns to add borders after them
        location_type_n_columns = []
        for col_idx, col_name in enumerate(locations_df.columns, start=1):
            if col_name.startswith('Location Type ') and col_name != 'Location Type':
                # Extract the number
                number = col_name.replace('Location Type ', '').strip()
                if number.isdigit():
                    location_type_n_columns.append((col_idx, col_name, int(number)))
        
        # Also check for base Location Type column (if it exists and is not numbered)
        base_location_type_idx = None
        for col_idx, col_name in enumerate(locations_df.columns, start=1):
            if col_name == 'Location Type':
                base_location_type_idx = col_idx
                break
        
        # Sort by number
        location_type_n_columns.sort(key=lambda x: x[2])
        
        # Add borders to all cells in Location Type n columns (right border to separate groups)
        all_location_type_cols = []
        if base_location_type_idx:
            all_location_type_cols.append((base_location_type_idx, 'Location Type'))
        all_location_type_cols.extend([(idx, name) for idx, name, _ in location_type_n_columns])
        
        for col_idx, col_name in all_location_type_cols:
            column_letter = get_column_letter(col_idx)
            # Apply right border to all cells in this column (including header)
            max_row = ws.max_row
            for row_idx in range(1, max_row + 1):
                cell = ws[f"{column_letter}{row_idx}"]
                # Get existing border and add right border
                existing_border = cell.border if cell.border else Border()
                cell.border = Border(
                    left=existing_border.left if existing_border.left else Side(),
                    right=Side(style='thin', color='000000'),
                    top=existing_border.top if existing_border.top else Side(),
                    bottom=existing_border.bottom if existing_border.bottom else Side()
                )
        
        print("Added vertical borders to group location columns")
        
        wb.save(locations_input_path)
        
        # Post-processing: Add virtual locations for rows with Location Type 'Both' or 'Virtual'
        # This runs after ALL cases (1, 2, 3, 4) are complete
        print("\nPost-processing: Adding virtual locations for 'Both' and 'Virtual' Location Types...")
        # Reload the dataframe to ensure we have the latest data (important for Case 2 which saved earlier)
        locations_df = pd.read_excel(locations_input_path, sheet_name='Locations')
        add_virtual_locations_after_matching(locations_df, api_locations_df)
        
        # Final column reordering after virtual locations are added
        print("\nFinal column reordering after virtual locations...")
        # Define the base column order
        base_columns = [
            'NPI Number',
            'First Name',
            'Last Name',
            'Address Line 1',
            'Address Line 2',
            'City',
            'State',
            'ZIP',
            'Practice ID'
        ]
        
        # Find all Location ID columns (both Location ID and Location Monolith ID)
        location_id_cols = [col for col in locations_df.columns if col.startswith('Location Monolith ID') or col.startswith('Location ID')]
        location_id_cols_sorted = sorted(
            location_id_cols,
            key=lambda x: int(x.split()[-1]) if x.split()[-1].isdigit() else 999
        )
        
        # Build the final column order
        final_column_order = []
        
        # Add base columns first (only if they exist in the dataframe)
        for base_col in base_columns:
            if base_col in locations_df.columns:
                final_column_order.append(base_col)
        
        # Add Location ID (or Location Monolith ID), Location Cloud ID, Practice Cloud ID, and Location Type columns in sequence
        for location_id_col in location_id_cols_sorted:
            # Extract number from Location Monolith ID or Location ID column name
            if location_id_col == 'Location ID' or location_id_col == 'Location Monolith ID':
                number = ''
            elif location_id_col.startswith('Location Monolith ID'):
                number = location_id_col.replace('Location Monolith ID', '').strip()
            else:
                number = location_id_col.replace('Location ID', '').strip()
            
            # Add Location ID column
            if location_id_col not in final_column_order:
                final_column_order.append(location_id_col)
            
            # Add corresponding Location Cloud ID column
            if number:
                cloud_col = f'Location Cloud ID {number}'
            else:
                cloud_col = 'Location Cloud ID'
            
            if cloud_col in locations_df.columns and cloud_col not in final_column_order:
                final_column_order.append(cloud_col)
            
            # Add corresponding Practice Cloud ID column
            if number:
                practice_cloud_col = f'Practice Cloud ID {number}'
            else:
                practice_cloud_col = 'Practice Cloud ID'
            
            if practice_cloud_col in locations_df.columns and practice_cloud_col not in final_column_order:
                final_column_order.append(practice_cloud_col)
            
            # Add corresponding Location Type column
            if number:
                location_type_col = f'Location Type {number}'
            else:
                location_type_col = 'Location Type'
            
            if location_type_col in locations_df.columns and location_type_col not in final_column_order:
                final_column_order.append(location_type_col)
        
        # Add any remaining columns that weren't in the specified order
        for col in locations_df.columns:
            if col not in final_column_order:
                final_column_order.append(col)
        
        # Reorder the dataframe
        locations_df = locations_df[final_column_order]
        print("Column reordering complete after virtual locations.")
        
        # Save again after virtual locations and final reordering
        with pd.ExcelWriter(locations_input_path, engine='openpyxl', mode='w') as writer:
            locations_df.to_excel(writer, index=False, sheet_name='Locations')
        
        # Reapply styling after virtual locations
        wb = load_workbook(locations_input_path)
        ws = wb['Locations']
        
        # Style header row
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Highlight Location Type n columns based on Location Type comparison
        if 'Location Type' in locations_df.columns:
            green_fill = PatternFill(start_color='ABFFAB', end_color='ABFFAB', fill_type='solid')  # Light green
            red_fill = PatternFill(start_color='FFA3A3', end_color='FFA3A3', fill_type='solid')  # Light red
            
            # Find Location Type column index
            location_type_col_idx = None
            for col_idx, col_name in enumerate(locations_df.columns, start=1):
                if col_name == 'Location Type':
                    location_type_col_idx = col_idx
                    break
            
            if location_type_col_idx:
                # Find all Location Type n columns
                location_type_n_columns = []
                for col_idx, col_name in enumerate(locations_df.columns, start=1):
                    if col_name.startswith('Location Type ') and col_name != 'Location Type':
                        # Extract the number
                        number = col_name.replace('Location Type ', '').strip()
                        if number.isdigit():
                            location_type_n_columns.append((col_idx, col_name, int(number)))
                
                # Sort by number
                location_type_n_columns.sort(key=lambda x: x[2])
                
                if location_type_n_columns:
                    print(f"Highlighting {len(location_type_n_columns)} Location Type n column(s) based on Location Type comparison...")
                    
                    # Process each row
                    for row_idx, (idx, row) in enumerate(locations_df.iterrows(), start=2):  # Start at 2 (row 1 is header)
                        location_type = row.get('Location Type', '')
                        location_type_str = str(location_type).strip() if pd.notna(location_type) else ''
                        
                        # Highlight each Location Type n column based on comparison
                        for col_idx, col_name, number in location_type_n_columns:
                            location_type_n_val = row.get(col_name, '')
                            location_type_n_str = str(location_type_n_val).strip() if pd.notna(location_type_n_val) else ''
                            
                            # Determine highlight color based on rules
                            highlight_color = None
                            
                            if location_type_str == 'Both':
                                # Both → all Location Type n should be green
                                highlight_color = green_fill
                            elif location_type_str == 'In Person':
                                # In Person → green if Location Type n is 'In Person', red otherwise
                                if location_type_n_str == 'In Person':
                                    highlight_color = green_fill
                                else:
                                    highlight_color = red_fill
                            elif location_type_str == 'Virtual':
                                # Virtual → green if Location Type n is 'Virtual', red otherwise
                                if location_type_n_str == 'Virtual':
                                    highlight_color = green_fill
                                else:
                                    highlight_color = red_fill
                            
                            # Apply highlighting if color determined
                            if highlight_color:
                                cell = ws.cell(row=row_idx, column=col_idx)
                                cell.fill = highlight_color
                    
                    print(f"Highlighted Location Type n columns based on Location Type comparison")
        
        # Add vertical borders to group location columns
        print("Adding vertical borders to group location columns...")
        right_border = Border(
            right=Side(style='thin', color='000000')
        )
        
        # Find all Location Type n columns to add borders after them
        location_type_n_columns = []
        for col_idx, col_name in enumerate(locations_df.columns, start=1):
            if col_name.startswith('Location Type ') and col_name != 'Location Type':
                number = col_name.replace('Location Type ', '').strip()
                if number.isdigit():
                    location_type_n_columns.append((col_idx, col_name, int(number)))
        
        # Also check for base Location Type column
        base_location_type_idx = None
        for col_idx, col_name in enumerate(locations_df.columns, start=1):
            if col_name == 'Location Type':
                base_location_type_idx = col_idx
                break
        
        location_type_n_columns.sort(key=lambda x: x[2])
        
        all_location_type_cols = []
        if base_location_type_idx:
            all_location_type_cols.append((base_location_type_idx, 'Location Type'))
        all_location_type_cols.extend([(idx, name) for idx, name, _ in location_type_n_columns])
        
        for col_idx, col_name in all_location_type_cols:
            column_letter = get_column_letter(col_idx)
            max_row = ws.max_row
            for row_idx in range(1, max_row + 1):
                cell = ws[f"{column_letter}{row_idx}"]
                existing_border = cell.border if cell.border else Border()
                cell.border = Border(
                    left=existing_border.left if existing_border.left else Side(),
                    right=Side(style='thin', color='000000'),
                    top=existing_border.top if existing_border.top else Side(),
                    bottom=existing_border.bottom if existing_border.bottom else Side()
                )
        
        print("Added vertical borders to group location columns")
        
        wb.save(locations_input_path)
        
        print(f"Added Location Cloud ID columns to Locations_input.xlsx")
        
    except Exception as e:
        print(f"Warning: Error adding Location Cloud ID to Locations_input.xlsx: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == '__main__':
    # For testing purposes
    try:
        result = create_locations_input_excel()
        print(f"Successfully created Locations_input.xlsx at {result}")
    except Exception as e:
        print(f"Error creating Locations_input.xlsx: {str(e)}")
        sys.exit(1)

