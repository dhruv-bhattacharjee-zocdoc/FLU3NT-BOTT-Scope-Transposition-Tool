"""
Create Mapped - Creates _Mapped.xlsx file from Input.xlsx using Mappings.xlsx

This script reads Mappings.xlsx to understand the column mappings, then reads Input.xlsx
and creates a new file _Mapped.xlsx with the mapped columns.
"""

import os
import sys
import re
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Get the project root directory (parent of Logics folder)
LOGICS_DIR = Path(__file__).parent
PROJECT_ROOT = LOGICS_DIR.parent
BACKEND_DIR = PROJECT_ROOT / 'backend'
EXCEL_FILES_DIR = BACKEND_DIR / 'Excel Files'

def create_mapped_excel():
    """
    Create _Mapped.xlsx file from Input.xlsx using column mappings from Mappings.xlsx
    
    Returns:
        Path to the created _Mapped.xlsx file
    """
    # Ensure Excel Files directory exists
    EXCEL_FILES_DIR.mkdir(parents=True, exist_ok=True)
    
    # File paths
    mappings_file = EXCEL_FILES_DIR / 'Mappings.xlsx'
    input_file = EXCEL_FILES_DIR / 'Input.xlsx'
    output_file = EXCEL_FILES_DIR / '_Mapped.xlsx'
    
    # Check if required files exist
    if not mappings_file.exists():
        raise FileNotFoundError(f"Mappings.xlsx not found at {mappings_file}")
    
    if not input_file.exists():
        raise FileNotFoundError(f"Input.xlsx not found at {input_file}")
    
    # Read Mappings.xlsx
    mappings_df = pd.read_excel(mappings_file, sheet_name='Mappings')
    
    # Check required columns
    if 'Detected As' not in mappings_df.columns or 'Column Name' not in mappings_df.columns:
        raise ValueError("Mappings.xlsx must contain 'Detected As' and 'Column Name' columns")
    
    # Create mapping dictionary: Detected As -> Column Name
    column_mapping = {}
    # Track if both First Name and Last Name map to the same column
    first_name_column = None
    last_name_column = None
    same_column_for_names = None
    
    for _, row in mappings_df.iterrows():
        detected_as = str(row['Detected As']).strip()
        column_name = str(row['Column Name']).strip()
        
        # Handle multiple mappings (comma-separated in Detected As)
        detected_as_list = [d.strip() for d in detected_as.split(',')]
        
        for detected in detected_as_list:
            if detected and column_name:
                # Check if this is First Name or Last Name
                if detected in ['First Name', 'firstName']:
                    first_name_column = column_name
                elif detected in ['Last Name', 'lastName']:
                    last_name_column = column_name
                
                # If multiple columns map to same detected_as, use the first one
                if detected not in column_mapping:
                    column_mapping[detected] = column_name
    
    # Check if both First Name and Last Name map to the same column
    if first_name_column and last_name_column and first_name_column == last_name_column:
        same_column_for_names = first_name_column
        print(f"Both First Name and Last Name map to the same column: {same_column_for_names}")
        print("Will split concatenated names into separate First Name and Last Name columns")
    
    # Read Input.xlsx
    input_df = pd.read_excel(input_file)
    
    # Create new dataframe with mapped columns
    mapped_data = {}
    
    for detected_as, source_column in column_mapping.items():
        if source_column in input_df.columns:
            mapped_data[detected_as] = input_df[source_column]
        else:
            # If source column doesn't exist, fill with empty values
            mapped_data[detected_as] = [''] * len(input_df)
    
    # Handle case where both First Name and Last Name map to the same column
    if same_column_for_names and same_column_for_names in input_df.columns:
        # Split concatenated names into First Name and Last Name
        full_names = input_df[same_column_for_names].astype(str)
        first_names = []
        last_names = []
        
        for full_name in full_names:
            if pd.isna(full_name) or full_name == 'nan' or full_name.strip() == '':
                first_names.append('')
                last_names.append('')
            else:
                # Split by space - first word is first name, rest is last name
                name_parts = full_name.strip().split()
                if len(name_parts) >= 2:
                    first_names.append(name_parts[0])
                    last_names.append(' '.join(name_parts[1:]))
                elif len(name_parts) == 1:
                    # Only one word - assume it's first name
                    first_names.append(name_parts[0])
                    last_names.append('')
                else:
                    first_names.append('')
                    last_names.append('')
        
        # Update mapped data with split names
        if 'First Name' in mapped_data or 'firstName' in mapped_data:
            key = 'First Name' if 'First Name' in mapped_data else 'firstName'
            mapped_data[key] = first_names
        else:
            mapped_data['First Name'] = first_names
        
        if 'Last Name' in mapped_data or 'lastName' in mapped_data:
            key = 'Last Name' if 'Last Name' in mapped_data else 'lastName'
            mapped_data[key] = last_names
        else:
            mapped_data['Last Name'] = last_names
        
        print(f"Split {len([n for n in first_names if n])} concatenated names into First Name and Last Name")
    
    # Remove symbols (except '.') from First Name and Last Name columns
    def remove_symbols_except_period(text):
        """Remove all symbols except period (.) from text"""
        if pd.isna(text) or text == '':
            return text
        text_str = str(text)
        # Keep only alphanumeric characters, spaces, and periods
        cleaned = re.sub(r'[^a-zA-Z0-9\s.]', '', text_str)
        return cleaned.strip()
    
    # Clean First Name and Last Name columns
    if 'First Name' in mapped_data:
        mapped_data['First Name'] = [remove_symbols_except_period(name) for name in mapped_data['First Name']]
    elif 'firstName' in mapped_data:
        mapped_data['firstName'] = [remove_symbols_except_period(name) for name in mapped_data['firstName']]
    
    if 'Last Name' in mapped_data:
        mapped_data['Last Name'] = [remove_symbols_except_period(name) for name in mapped_data['Last Name']]
    elif 'lastName' in mapped_data:
        mapped_data['lastName'] = [remove_symbols_except_period(name) for name in mapped_data['lastName']]
    
    # Create the mapped dataframe
    mapped_df = pd.DataFrame(mapped_data)
    
    # Read NPI-Extracts.xlsx for name comparison
    npi_extracts_file = EXCEL_FILES_DIR / 'NPI-Extracts.xlsx'
    npi_extracts_df = None
    npi_name_lookup = {}  # NPI Number -> {FIRST_NAME, LAST_NAME}
    
    if npi_extracts_file.exists():
        try:
            npi_extracts_df = pd.read_excel(npi_extracts_file, sheet_name='NPI Extracts')
            print(f"Reading NPI-Extracts.xlsx for name comparison...")
            
            # Create lookup dictionary: NPI Number -> {FIRST_NAME, LAST_NAME}
            if 'NPI Number' in npi_extracts_df.columns:
                for idx, row in npi_extracts_df.iterrows():
                    npi_val = row['NPI Number']
                    # Convert to string and remove .0 suffix if present
                    if pd.notna(npi_val):
                        npi = str(npi_val).replace('.0', '').strip() if isinstance(npi_val, float) else str(npi_val).strip()
                    else:
                        npi = None
                    first_name = str(row['FIRST_NAME']).strip() if pd.notna(row.get('FIRST_NAME')) else ''
                    last_name = str(row['LAST_NAME']).strip() if pd.notna(row.get('LAST_NAME')) else ''
                    
                    if npi and npi != 'nan':
                        npi_name_lookup[npi] = {
                            'FIRST_NAME': first_name.upper() if first_name else '',
                            'LAST_NAME': last_name.upper() if last_name else ''
                        }
                
                print(f"Created name lookup dictionary with {len(npi_name_lookup)} entries")
        except Exception as e:
            print(f"Warning: Could not read NPI-Extracts.xlsx for name comparison: {str(e)}")
    else:
        print(f"Warning: NPI-Extracts.xlsx not found at {npi_extracts_file}. Skipping name comparison.")
    
    # Write to _Mapped.xlsx with formatting
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        mapped_df.to_excel(writer, sheet_name='Mapped', index=False)
        
        # Get the workbook and worksheet for formatting
        workbook = writer.book
        worksheet = writer.sheets['Mapped']
        
        # Format header row
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Find First Name and Last Name columns for highlighting
        first_name_col_idx = None
        last_name_col_idx = None
        npi_col_idx = None
        
        for col_idx, col_name in enumerate(mapped_df.columns, start=1):
            col_name_str = str(col_name).strip()
            if col_name_str in ['First Name', 'firstName']:
                first_name_col_idx = col_idx
            elif col_name_str in ['Last Name', 'lastName']:
                last_name_col_idx = col_idx
            elif col_name_str in ['NPI Number', 'NPI', 'npi']:
                npi_col_idx = col_idx
        
        # Highlight mismatches if we have NPI lookup and name columns
        if npi_name_lookup and first_name_col_idx and last_name_col_idx and npi_col_idx:
            highlight_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')  # Light red
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
            mismatch_count = 0
            empty_name_count = 0
            checked_count = 0
            not_found_count = 0
            
            # Get column names for easier access
            first_name_col_name = mapped_df.columns[first_name_col_idx - 1]
            last_name_col_name = mapped_df.columns[last_name_col_idx - 1]
            npi_col_name = mapped_df.columns[npi_col_idx - 1]
            
            print(f"Checking name matches: First Name column='{first_name_col_name}', Last Name column='{last_name_col_name}', NPI column='{npi_col_name}'")
            print(f"NPI lookup dictionary has {len(npi_name_lookup)} entries")
            
            # Compare names row by row
            for row_idx, (idx, row) in enumerate(mapped_df.iterrows(), start=2):  # Start at 2 (row 1 is header)
                # Get NPI value from the row
                npi_val = row.get(npi_col_name) if npi_col_name in row else None
                
                # Convert to string and remove .0 suffix if present
                if pd.notna(npi_val):
                    npi = str(npi_val).replace('.0', '').strip() if isinstance(npi_val, float) else str(npi_val).strip()
                else:
                    npi = None
                
                if npi and npi != 'nan':
                    if npi in npi_name_lookup:
                        checked_count += 1
                    # Get names from NPI-Extracts
                    npi_first_name = npi_name_lookup[npi]['FIRST_NAME']
                    npi_last_name = npi_name_lookup[npi]['LAST_NAME']
                    
                    # Check if both FIRST_NAME and LAST_NAME are empty in NPI-Extracts
                    if (not npi_first_name or npi_first_name.strip() == '') and (not npi_last_name or npi_last_name.strip() == ''):
                        # Highlight NPI Number cell in yellow
                        npi_cell = worksheet.cell(row=row_idx, column=npi_col_idx)
                        npi_cell.fill = yellow_fill
                        empty_name_count += 1
                    
                    # Get names from _Mapped.xlsx
                    mapped_first_name_val = row.get(first_name_col_name) if first_name_col_name in row else None
                    mapped_last_name_val = row.get(last_name_col_name) if last_name_col_name in row else None
                    
                    mapped_first_name = ''
                    mapped_last_name = ''
                    
                    if pd.notna(mapped_first_name_val) and mapped_first_name_val != '':
                        mapped_first_name = str(mapped_first_name_val).strip().upper()
                    if pd.notna(mapped_last_name_val) and mapped_last_name_val != '':
                        mapped_last_name = str(mapped_last_name_val).strip().upper()
                    
                    # Compare and highlight if different (case-insensitive comparison)
                    # Highlight if both are non-empty and different, OR if one is empty and the other is not
                    first_name_mismatch = False
                    last_name_mismatch = False
                    
                    if mapped_first_name or npi_first_name:
                        if mapped_first_name != npi_first_name:
                            first_name_mismatch = True
                    
                    if mapped_last_name or npi_last_name:
                        if mapped_last_name != npi_last_name:
                            last_name_mismatch = True
                    
                    if first_name_mismatch:
                        # Highlight First Name cell
                        first_name_cell = worksheet.cell(row=row_idx, column=first_name_col_idx)
                        first_name_cell.fill = highlight_fill
                        mismatch_count += 1
                    
                    if last_name_mismatch:
                        # Highlight Last Name cell
                        last_name_cell = worksheet.cell(row=row_idx, column=last_name_col_idx)
                        last_name_cell.fill = highlight_fill
                        mismatch_count += 1
                    else:
                        not_found_count += 1
            
            print(f"Checked {checked_count} rows with matching NPIs, {not_found_count} rows with NPIs not found in lookup")
            if mismatch_count > 0:
                print(f"Highlighted {mismatch_count} name mismatches in _Mapped.xlsx")
            else:
                print("No name mismatches found - all names match NPI-Extracts.xlsx")
            
            if empty_name_count > 0:
                print(f"Highlighted {empty_name_count} NPI Number(s) in yellow (FIRST_NAME and LAST_NAME are empty in NPI-Extracts.xlsx)")
        else:
            if not npi_name_lookup:
                print("Warning: NPI name lookup is empty - cannot highlight mismatches")
            if not first_name_col_idx:
                print("Warning: First Name column not found - cannot highlight mismatches")
            if not last_name_col_idx:
                print("Warning: Last Name column not found - cannot highlight mismatches")
            if not npi_col_idx:
                print("Warning: NPI Number column not found - cannot highlight mismatches")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook to ensure highlights are persisted
        workbook.save(output_file)
    
    return str(output_file)

if __name__ == '__main__':
    # Test the function
    try:
        result_path = create_mapped_excel()
        print(f"Excel file created at: {result_path}")
    except Exception as e:
        print(f"Error: {str(e)}")

