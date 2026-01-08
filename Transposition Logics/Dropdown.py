"""
Dropdown - Applies data validation dropdowns to specific columns in Template copy.xlsx

This script applies data validation (dropdown lists) to columns in the Provider and Location sheets
of Template copy.xlsx. Dropdowns are organized by section (Provider and Locations).
"""

import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Get the project root directory (parent of Transposition Logics folder)
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
BACKEND_DIR = PROJECT_ROOT / 'backend'
EXCEL_FILES_DIR = BACKEND_DIR / 'Excel Files'

def apply_dropdowns_to_template():
    """
    Apply data validation dropdowns to columns in Template copy.xlsx
    
    Returns:
        bool: True if successful, False otherwise
    """
    template_file = EXCEL_FILES_DIR / 'Template copy.xlsx'
    
    # Check if template file exists
    if not template_file.exists():
        return False
    
    try:
        # Load the template workbook
        wb = load_workbook(template_file)
        
        # ========== PROVIDER SECTION ==========
        if 'Provider' in wb.sheetnames:
            provider_sheet = wb['Provider']
            header_row = 1
            max_row = provider_sheet.max_row
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            
            # Gender: Male,Female,NonBinary,Not Applicable
            gender_column_letter = None
            for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
                if cell.value and str(cell.value).strip().lower() == 'gender':
                    gender_column_letter = get_column_letter(col_idx)
                    break
            
            if gender_column_letter:
                gender_validation = DataValidation(type="list", formula1='"Male,Female,NonBinary,Not Applicable"')
                gender_validation.add(f"{gender_column_letter}2:{gender_column_letter}{max_row}")
                provider_sheet.add_data_validation(gender_validation)
            
            # Professional Suffix 1-3: ValidationAndReference!$G$2:$G$511
            professional_suffix_columns = {}
            for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip()
                    if cell_value == 'Professional Suffix 1':
                        professional_suffix_columns[1] = get_column_letter(col_idx)
                    elif cell_value == 'Professional Suffix 2':
                        professional_suffix_columns[2] = get_column_letter(col_idx)
                    elif cell_value == 'Professional Suffix 3':
                        professional_suffix_columns[3] = get_column_letter(col_idx)
            
            for suffix_num in [1, 2, 3]:
                if suffix_num in professional_suffix_columns:
                    col_letter = professional_suffix_columns[suffix_num]
                    suffix_validation = DataValidation(type="list", formula1="ValidationAndReference!$G$2:$G$511")
                    suffix_validation.add(f"{col_letter}2:{col_letter}{max_row}")
                    provider_sheet.add_data_validation(suffix_validation)
            
            # Specialty 1-5: ValidationAndReference!$K:$K
            specialty_columns = {}
            for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip()
                    if cell_value == 'Specialty 1':
                        specialty_columns[1] = get_column_letter(col_idx)
                    elif cell_value == 'Specialty 2':
                        specialty_columns[2] = get_column_letter(col_idx)
                    elif cell_value == 'Specialty 3':
                        specialty_columns[3] = get_column_letter(col_idx)
                    elif cell_value == 'Specialty 4':
                        specialty_columns[4] = get_column_letter(col_idx)
                    elif cell_value == 'Specialty 5':
                        specialty_columns[5] = get_column_letter(col_idx)
            
            for specialty_num in [1, 2, 3, 4, 5]:
                if specialty_num in specialty_columns:
                    col_letter = specialty_columns[specialty_num]
                    specialty_validation = DataValidation(type="list", formula1="ValidationAndReference!$K:$K")
                    specialty_validation.add(f"{col_letter}2:{col_letter}{max_row}")
                    provider_sheet.add_data_validation(specialty_validation)
            
            # Board Certification 1-5: ValidationAndReference!$N$2:$N$299
            board_cert_columns = {}
            for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip()
                    if cell_value == 'Board Certification 1':
                        board_cert_columns[1] = get_column_letter(col_idx)
                    elif cell_value == 'Board Certification 2':
                        board_cert_columns[2] = get_column_letter(col_idx)
                    elif cell_value == 'Board Certification 3':
                        board_cert_columns[3] = get_column_letter(col_idx)
                    elif cell_value == 'Board Certification 4':
                        board_cert_columns[4] = get_column_letter(col_idx)
                    elif cell_value == 'Board Certification 5':
                        board_cert_columns[5] = get_column_letter(col_idx)
            
            for cert_num in [1, 2, 3, 4, 5]:
                if cert_num in board_cert_columns:
                    col_letter = board_cert_columns[cert_num]
                    cert_validation = DataValidation(type="list", formula1="ValidationAndReference!$N$2:$N$299")
                    cert_validation.add(f"{col_letter}2:{col_letter}{max_row}")
                    provider_sheet.add_data_validation(cert_validation)
            
            # Sub Board Certification 1-5: ValidationAndReference!$N$2:$N$294
            sub_board_cert_columns = {}
            for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip()
                    if cell_value == 'Sub Board Certification 1':
                        sub_board_cert_columns[1] = get_column_letter(col_idx)
                    elif cell_value == 'Sub Board Certification 2':
                        sub_board_cert_columns[2] = get_column_letter(col_idx)
                    elif cell_value == 'Sub Board Certification 3':
                        sub_board_cert_columns[3] = get_column_letter(col_idx)
                    elif cell_value == 'Sub Board Certification 4':
                        sub_board_cert_columns[4] = get_column_letter(col_idx)
                    elif cell_value == 'Sub Board Certification 5':
                        sub_board_cert_columns[5] = get_column_letter(col_idx)
            
            for sub_cert_num in [1, 2, 3, 4, 5]:
                if sub_cert_num in sub_board_cert_columns:
                    col_letter = sub_board_cert_columns[sub_cert_num]
                    sub_cert_validation = DataValidation(type="list", formula1="ValidationAndReference!$N$2:$N$294")
                    sub_cert_validation.add(f"{col_letter}2:{col_letter}{max_row}")
                    provider_sheet.add_data_validation(sub_cert_validation)
            
            # Hospital Affiliation 1-5: ValidationAndReference!$T$2:$T$7258
            hospital_affiliation_columns = {}
            for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip()
                    if cell_value == 'Hospital Affiliation 1':
                        hospital_affiliation_columns[1] = get_column_letter(col_idx)
                    elif cell_value == 'Hospital Affiliation 2':
                        hospital_affiliation_columns[2] = get_column_letter(col_idx)
                    elif cell_value == 'Hospital Affiliation 3':
                        hospital_affiliation_columns[3] = get_column_letter(col_idx)
                    elif cell_value == 'Hospital Affiliation 4':
                        hospital_affiliation_columns[4] = get_column_letter(col_idx)
                    elif cell_value == 'Hospital Affiliation 5':
                        hospital_affiliation_columns[5] = get_column_letter(col_idx)
            
            for hosp_num in [1, 2, 3, 4, 5]:
                if hosp_num in hospital_affiliation_columns:
                    col_letter = hospital_affiliation_columns[hosp_num]
                    hosp_validation = DataValidation(type="list", formula1="ValidationAndReference!$T$2:$T$7258")
                    hosp_validation.add(f"{col_letter}2:{col_letter}{max_row}")
                    provider_sheet.add_data_validation(hosp_validation)
            
            # Additional Languages Spoken 1-3: ValidationAndReference!$W$2:$W$144
            additional_languages_columns = {}
            for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip().lower()
                    if cell_value == 'additional language spoken 1' or cell_value == 'additional languages spoken 1':
                        additional_languages_columns[1] = get_column_letter(col_idx)
                    elif cell_value == 'additional language spoken 2' or cell_value == 'additional languages spoken 2':
                        additional_languages_columns[2] = get_column_letter(col_idx)
                    elif cell_value == 'additional language spoken 3' or cell_value == 'additional languages spoken 3':
                        additional_languages_columns[3] = get_column_letter(col_idx)
            
            for lang_num in [1, 2, 3]:
                if lang_num in additional_languages_columns:
                    col_letter = additional_languages_columns[lang_num]
                    lang_validation = DataValidation(type="list", formula1="ValidationAndReference!$W$2:$W$144")
                    lang_validation.add(f"{col_letter}2:{col_letter}{max_row}")
                    provider_sheet.add_data_validation(lang_validation)
                    
                    # Color the Additional Languages Spoken column header with green (#00FF00)
                    header_cell = provider_sheet[f"{col_letter}{header_row}"]
                    header_cell.fill = green_fill
            
            # Provider type: ValidationAndReference!$Q$2:$Q$9 (Default value - 'Practitioner - Full Profile')
            provider_type_column_letter = None
            for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip().lower()
                    if cell_value == 'provider type' or cell_value == 'provider type (substatus)':
                        provider_type_column_letter = get_column_letter(col_idx)
                        break
            
            if provider_type_column_letter:
                provider_type_validation = DataValidation(type="list", formula1="ValidationAndReference!$Q$2:$Q$9")
                provider_type_validation.add(f"{provider_type_column_letter}2:{provider_type_column_letter}{max_row}")
                provider_sheet.add_data_validation(provider_type_validation)
                
                # Set default value 'Practitioner - Full Profile' for empty cells
                for row_idx in range(2, max_row + 1):
                    cell = provider_sheet[f"{provider_type_column_letter}{row_idx}"]
                    if cell.value is None or str(cell.value).strip() == '':
                        cell.value = 'Practitioner - Full Profile'
                
                # Color the Provider Type column header with green (#00FF00)
                header_cell = provider_sheet[f"{provider_type_column_letter}{header_row}"]
                header_cell.fill = green_fill
            
            # Enterprise Scheduling Flag: Yes, No (Default Value - 'No')
            enterprise_scheduling_column_letter = None
            for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip().lower()
                    if cell_value == 'enterprise scheduling flag':
                        enterprise_scheduling_column_letter = get_column_letter(col_idx)
                        break
            
            if enterprise_scheduling_column_letter:
                enterprise_validation = DataValidation(type="list", formula1='"Yes,No"')
                enterprise_validation.add(f"{enterprise_scheduling_column_letter}2:{enterprise_scheduling_column_letter}{max_row}")
                provider_sheet.add_data_validation(enterprise_validation)
                
                # Set default value 'No' for empty cells
                for row_idx in range(2, max_row + 1):
                    cell = provider_sheet[f"{enterprise_scheduling_column_letter}{row_idx}"]
                    if cell.value is None or str(cell.value).strip() == '':
                        cell.value = 'No'
                
                # Color the Enterprise Scheduling Flag column header with green (#00FF00)
                header_cell = provider_sheet[f"{enterprise_scheduling_column_letter}{header_row}"]
                header_cell.fill = green_fill
            
            # Location 1-5: Location!$X:$X
            location_columns = {}
            for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip()
                    if cell_value == 'Location 1':
                        location_columns[1] = get_column_letter(col_idx)
                    elif cell_value == 'Location 2':
                        location_columns[2] = get_column_letter(col_idx)
                    elif cell_value == 'Location 3':
                        location_columns[3] = get_column_letter(col_idx)
                    elif cell_value == 'Location 4':
                        location_columns[4] = get_column_letter(col_idx)
                    elif cell_value == 'Location 5':
                        location_columns[5] = get_column_letter(col_idx)
            
            for location_num in [1, 2, 3, 4, 5]:
                if location_num in location_columns:
                    col_letter = location_columns[location_num]
                    location_validation = DataValidation(type="list", formula1="Location!$X:$X")
                    location_validation.add(f"{col_letter}2:{col_letter}{max_row}")
                    provider_sheet.add_data_validation(location_validation)
        
        # ========== LOCATIONS SECTION ==========
        if 'Location' in wb.sheetnames:
            location_sheet = wb['Location']
            header_row = 1
            max_row = location_sheet.max_row
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            
            # Location Type: In Person,Virtual
            location_type_column_letter = None
            for col_idx, cell in enumerate(location_sheet[header_row], start=1):
                if cell.value and str(cell.value).strip().lower() == 'location type':
                    location_type_column_letter = get_column_letter(col_idx)
                    break
            
            if location_type_column_letter:
                location_type_validation = DataValidation(type="list", formula1='"In Person,Virtual"')
                location_type_validation.add(f"{location_type_column_letter}2:{location_type_column_letter}{max_row}")
                location_sheet.add_data_validation(location_type_validation)
            
            # State: ValidationAndReference!$A$2:$A$55
            state_column_letter = None
            for col_idx, cell in enumerate(location_sheet[header_row], start=1):
                if cell.value and str(cell.value).strip().lower() == 'state':
                    state_column_letter = get_column_letter(col_idx)
                    break
            
            if state_column_letter:
                state_validation = DataValidation(type="list", formula1="ValidationAndReference!$A$2:$A$55")
                state_validation.add(f"{state_column_letter}2:{state_column_letter}{max_row}")
                location_sheet.add_data_validation(state_validation)
            
            # Virtual Visit Type: ValidationAndReference!$Y$2:$Y$3
            virtual_visit_type_column_letter = None
            for col_idx, cell in enumerate(location_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip().lower()
                    if cell_value == 'virtual visit type':
                        virtual_visit_type_column_letter = get_column_letter(col_idx)
                        break
            
            if virtual_visit_type_column_letter:
                virtual_visit_validation = DataValidation(type="list", formula1="ValidationAndReference!$Y$2:$Y$3")
                virtual_visit_validation.add(f"{virtual_visit_type_column_letter}2:{virtual_visit_type_column_letter}{max_row}")
                location_sheet.add_data_validation(virtual_visit_validation)
            
            # Show name in search?: Yes, No (Default Value - 'Yes')
            show_name_column_letter = None
            for col_idx, cell in enumerate(location_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip().lower()
                    if cell_value == 'show name in search?' or cell_value == 'show name in search':
                        show_name_column_letter = get_column_letter(col_idx)
                        break
            
            if show_name_column_letter:
                show_name_validation = DataValidation(type="list", formula1='"Yes,No"')
                show_name_validation.add(f"{show_name_column_letter}2:{show_name_column_letter}{max_row}")
                location_sheet.add_data_validation(show_name_validation)
                
                # Set default value 'Yes' for empty cells
                for row_idx in range(2, max_row + 1):
                    cell = location_sheet[f"{show_name_column_letter}{row_idx}"]
                    if cell.value is None or str(cell.value).strip() == '':
                        cell.value = 'Yes'
                
                # Color the Show name in search? column header with green (#00FF00)
                header_cell = location_sheet[f"{show_name_column_letter}{header_row}"]
                header_cell.fill = green_fill
            
            # Scheduling Software: ValidationAndReference!$D$2:$D$751
            scheduling_software_column_letter = None
            for col_idx, cell in enumerate(location_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip().lower()
                    if cell_value == 'scheduling software':
                        scheduling_software_column_letter = get_column_letter(col_idx)
                        break
            
            if scheduling_software_column_letter:
                scheduling_software_validation = DataValidation(type="list", formula1="ValidationAndReference!$D$2:$D$751")
                scheduling_software_validation.add(f"{scheduling_software_column_letter}2:{scheduling_software_column_letter}{max_row}")
                location_sheet.add_data_validation(scheduling_software_validation)
        
        # Save the workbook
        wb.save(template_file)
        return True
        
    except Exception as e:
        return False

if __name__ == "__main__":
    apply_dropdowns_to_template()

