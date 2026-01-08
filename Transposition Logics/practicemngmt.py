"""
Practice Management - Extracts Practice Cloud ID and Practice NAME values from Practice_Locations.xlsx and writes to Template copy.xlsx

This script reads the 'Practice Cloud ID' and 'Practice NAME' columns from the 'Practice' sheet in Practice_Locations.xlsx,
and writes the values to columns AD and AE respectively of the 'ValidationAndReference' sheet in Template copy.xlsx.
"""

import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

# Get the project root directory (parent of Transposition Logics folder)
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
BACKEND_DIR = PROJECT_ROOT / 'backend'
EXCEL_FILES_DIR = BACKEND_DIR / 'Excel Files'

def extract_practice_cloud_id_to_template():
    """
    Extract Practice Cloud ID and Practice NAME values from Practice_Locations.xlsx and write to Template copy.xlsx
    Writes to columns AD (Practice Cloud ID) and AE (Practice Name) of the ValidationAndReference sheet.
    
    Returns:
        bool: True if successful, False otherwise
    """
    # File paths
    practice_locations_file = EXCEL_FILES_DIR / 'Practice_Locations.xlsx'
    template_file = EXCEL_FILES_DIR / 'Template copy.xlsx'
    
    # Check if Practice_Locations file exists
    if not practice_locations_file.exists():
        return False
    
    # Check if template file exists
    if not template_file.exists():
        return False
    
    try:
        # Read Practice_Locations.xlsx (Practice sheet)
        practice_df = pd.read_excel(practice_locations_file, sheet_name='Practice')
        
        # Check if 'Practice Cloud ID' column exists
        if 'Practice Cloud ID' not in practice_df.columns:
            return False
        
        # Find the Practice Name column (check for various possible column names)
        practice_name_column_name = None
        possible_practice_name_columns = [
            'Practice NAME',
            'Practice Name',
            'Practice name',
            'practice_name',
            'practice name',
            'NAME',
            'Name',
            'name'
        ]
        
        for col_name in possible_practice_name_columns:
            if col_name in practice_df.columns:
                practice_name_column_name = col_name
                break
        
        # Load the template workbook
        wb = load_workbook(template_file)
        
        # Check if 'ValidationAndReference' sheet exists
        if 'ValidationAndReference' not in wb.sheetnames:
            return False
        
        # Get the ValidationAndReference sheet
        validation_sheet = wb['ValidationAndReference']
        
        # Column AD for Practice Cloud ID, Column AE for Practice Name
        practice_cloud_id_column = 'AD'
        practice_name_column = 'AE'
        header_row = 1
        
        # Define styles: bold font, purple fill color, and all borders
        bold_font = Font(bold=True)
        purple_fill = PatternFill(start_color="ECC9FF", end_color="ECC9FF", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set headers if they don't exist
        header_cell = validation_sheet[f"{practice_cloud_id_column}{header_row}"]
        if not header_cell.value or str(header_cell.value).strip() == '':
            header_cell.value = 'Practice Cloud ID'
        # Apply styles to header (always apply, even if header already exists)
        header_cell.font = bold_font
        header_cell.fill = purple_fill
        header_cell.border = thin_border
        
        header_cell = validation_sheet[f"{practice_name_column}{header_row}"]
        if not header_cell.value or str(header_cell.value).strip() == '':
            header_cell.value = 'Practice Name'
        # Apply styles to header (always apply, even if header already exists)
        header_cell.font = bold_font
        header_cell.fill = purple_fill
        header_cell.border = thin_border
        
        # Create a mapping of Practice Cloud ID to Practice Name
        # Use a dictionary to handle duplicates (keep the first non-empty Practice Name for each Cloud ID)
        practice_data = {}  # cloud_id -> practice_name
        for _, row in practice_df.iterrows():
            cloud_id = row.get('Practice Cloud ID', '')
            if pd.notna(cloud_id) and cloud_id != '':
                cloud_id_str = str(cloud_id).strip()
                # Only add if not already in dictionary
                if cloud_id_str not in practice_data:
                    practice_name = ''
                    if practice_name_column_name:
                        practice_name = row.get(practice_name_column_name, '')
                    if pd.notna(practice_name) and practice_name != '':
                        practice_data[cloud_id_str] = str(practice_name).strip()
                    else:
                        practice_data[cloud_id_str] = ''
        
        if not practice_data:
            return False
        
        # Get existing Practice Cloud IDs from ValidationAndReference sheet to avoid duplicates
        existing_cloud_ids = set()
        existing_cloud_id_to_row = {}  # Map cloud_id to row_idx for updating Practice Name
        max_row = validation_sheet.max_row
        if max_row > 1:
            for row_idx in range(2, max_row + 1):
                cell = validation_sheet[f"{practice_cloud_id_column}{row_idx}"]
                if cell.value:
                    cloud_id_str = str(cell.value).strip()
                    existing_cloud_ids.add(cloud_id_str)
                    existing_cloud_id_to_row[cloud_id_str] = row_idx
        
        # Filter out Practice Cloud IDs that already exist
        new_practice_data = {
            cloud_id: practice_name 
            for cloud_id, practice_name in practice_data.items()
            if cloud_id not in existing_cloud_ids
        }
        
        # Update Practice Name for existing Practice Cloud IDs if missing
        for cloud_id, practice_name in practice_data.items():
            if cloud_id in existing_cloud_ids and practice_name:
                existing_row_idx = existing_cloud_id_to_row[cloud_id]
                # Check if Practice Name is missing or empty for this row
                name_cell = validation_sheet[f"{practice_name_column}{existing_row_idx}"]
                if not name_cell.value or str(name_cell.value).strip() == '':
                    name_cell.value = practice_name
        
        if new_practice_data:
            # Find the next empty row
            next_row = max_row + 1
            
            # Check if there are empty rows we can use
            for row_idx in range(2, max_row + 1):
                cell = validation_sheet[f"{practice_cloud_id_column}{row_idx}"]
                if cell.value is None or str(cell.value).strip() == '':
                    next_row = row_idx
                    break
            
            # Write new Practice Cloud IDs and Practice Names to ValidationAndReference sheet
            for i, (cloud_id, practice_name) in enumerate(new_practice_data.items()):
                row_idx = next_row + i
                # Write Practice Cloud ID
                cell = validation_sheet[f"{practice_cloud_id_column}{row_idx}"]
                cell.value = cloud_id
                # Write Practice Name
                if practice_name:
                    cell = validation_sheet[f"{practice_name_column}{row_idx}"]
                    cell.value = practice_name
        
        # Now populate Practice Name in Location sheet based on Practice Cloud ID
        if 'Location' in wb.sheetnames:
            location_sheet = wb['Location']
            
            # Find Practice Cloud ID column in Location sheet
            location_practice_cloud_id_column_letter = None
            location_practice_name_column_letter = None
            
            for col_idx, cell in enumerate(location_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip().lower()
                    if cell_value == 'practice cloud id':
                        location_practice_cloud_id_column_letter = get_column_letter(col_idx)
                    elif cell_value == 'practice name':
                        location_practice_name_column_letter = get_column_letter(col_idx)
            
            # If Practice Name column doesn't exist, find the next available column after Practice Cloud ID
            if location_practice_cloud_id_column_letter and not location_practice_name_column_letter:
                # Find the column after Practice Cloud ID
                practice_cloud_id_col_num = None
                for col_idx, cell in enumerate(location_sheet[header_row], start=1):
                    if cell.value and str(cell.value).strip().lower() == 'practice cloud id':
                        practice_cloud_id_col_num = col_idx
                        break
                
                if practice_cloud_id_col_num:
                    location_practice_name_column_letter = get_column_letter(practice_cloud_id_col_num + 1)
                    # Set header
                    header_cell = location_sheet[f"{location_practice_name_column_letter}{header_row}"]
                    header_cell.value = 'Practice Name'
            
            # Create lookup dictionary from ValidationAndReference sheet (Practice Cloud ID -> Practice Name)
            validation_lookup = {}  # cloud_id -> practice_name
            validation_max_row = validation_sheet.max_row
            if validation_max_row > 1:
                for row_idx in range(2, validation_max_row + 1):
                    cloud_id_cell = validation_sheet[f"{practice_cloud_id_column}{row_idx}"]
                    name_cell = validation_sheet[f"{practice_name_column}{row_idx}"]
                    if cloud_id_cell.value:
                        cloud_id_str = str(cloud_id_cell.value).strip()
                        practice_name_value = ''
                        if name_cell.value:
                            practice_name_value = str(name_cell.value).strip()
                        validation_lookup[cloud_id_str] = practice_name_value
            
            # Populate Practice Name in Location sheet
            if location_practice_cloud_id_column_letter and location_practice_name_column_letter:
                location_max_row = location_sheet.max_row
                has_filled_values = False
                for row_idx in range(2, location_max_row + 1):
                    cloud_id_cell = location_sheet[f"{location_practice_cloud_id_column_letter}{row_idx}"]
                    if cloud_id_cell.value:
                        cloud_id_str = str(cloud_id_cell.value).strip()
                        # Look up Practice Name from ValidationAndReference sheet
                        if cloud_id_str in validation_lookup:
                            practice_name_value = validation_lookup[cloud_id_str]
                            if practice_name_value:
                                name_cell = location_sheet[f"{location_practice_name_column_letter}{row_idx}"]
                                name_cell.value = practice_name_value
                                has_filled_values = True
                
                # Check if Practice Name column has any filled values and color header green if so
                if location_practice_name_column_letter:
                    # Check all rows to see if any have values
                    if not has_filled_values:
                        for row_idx in range(2, location_max_row + 1):
                            name_cell = location_sheet[f"{location_practice_name_column_letter}{row_idx}"]
                            if name_cell.value and str(name_cell.value).strip() != '':
                                has_filled_values = True
                                break
                    
                    if has_filled_values:
                        header_cell = location_sheet[f"{location_practice_name_column_letter}{header_row}"]
                        header_cell.fill = green_fill
        
        # Now populate Practice Name in Provider sheet based on Practice Cloud ID
        if 'Provider' in wb.sheetnames:
            provider_sheet = wb['Provider']
            
            # Find Practice Cloud ID column in Provider sheet
            provider_practice_cloud_id_column_letter = None
            provider_practice_name_column_letter = None
            
            for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip().lower()
                    if cell_value == 'practice cloud id':
                        provider_practice_cloud_id_column_letter = get_column_letter(col_idx)
                    elif cell_value == 'practice name':
                        provider_practice_name_column_letter = get_column_letter(col_idx)
            
            # If Practice Name column doesn't exist, find the next available column after Practice Cloud ID
            if provider_practice_cloud_id_column_letter and not provider_practice_name_column_letter:
                # Find the column after Practice Cloud ID
                practice_cloud_id_col_num = None
                for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
                    if cell.value and str(cell.value).strip().lower() == 'practice cloud id':
                        practice_cloud_id_col_num = col_idx
                        break
                
                if practice_cloud_id_col_num:
                    provider_practice_name_column_letter = get_column_letter(practice_cloud_id_col_num + 1)
                    # Set header
                    header_cell = provider_sheet[f"{provider_practice_name_column_letter}{header_row}"]
                    header_cell.value = 'Practice Name'
            
            # Use the practice_data dictionary already created from Practice_Locations.xlsx
            # Populate Practice Name in Provider sheet
            if provider_practice_cloud_id_column_letter and provider_practice_name_column_letter:
                provider_max_row = provider_sheet.max_row
                has_filled_values = False
                for row_idx in range(2, provider_max_row + 1):
                    cloud_id_cell = provider_sheet[f"{provider_practice_cloud_id_column_letter}{row_idx}"]
                    if cloud_id_cell.value:
                        cloud_id_str = str(cloud_id_cell.value).strip()
                        # Look up Practice Name from practice_data dictionary (from Practice_Locations.xlsx)
                        if cloud_id_str in practice_data:
                            practice_name_value = practice_data[cloud_id_str]
                            if practice_name_value:
                                name_cell = provider_sheet[f"{provider_practice_name_column_letter}{row_idx}"]
                                name_cell.value = practice_name_value
                                has_filled_values = True
                
                # Check if Practice Name column has any filled values and color header green if so
                if provider_practice_name_column_letter:
                    # Check all rows to see if any have values
                    if not has_filled_values:
                        for row_idx in range(2, provider_max_row + 1):
                            name_cell = provider_sheet[f"{provider_practice_name_column_letter}{row_idx}"]
                            if name_cell.value and str(name_cell.value).strip() != '':
                                has_filled_values = True
                                break
                    
                    if has_filled_values:
                        header_cell = provider_sheet[f"{provider_practice_name_column_letter}{header_row}"]
                        header_cell.fill = green_fill
        
        # Save the workbook
        wb.save(template_file)
        return True
        
    except Exception as e:
        return False

if __name__ == "__main__":
    extract_practice_cloud_id_to_template()
