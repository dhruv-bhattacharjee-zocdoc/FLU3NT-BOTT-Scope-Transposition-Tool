"""
SimilarProvider - Merge duplicate providers based on NPI Number

This script identifies duplicate providers in Template copy.xlsx based on the 'NPI Number' column,
merges them into a single row, and consolidates unique values from:
- Professional Suffix n
- Specialty n
- Location ID n

A border is added to the NPI cell to indicate that it was merged from duplicate rows.
"""

import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

# Get the project root directory (parent of Transposition Logics folder)
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
BACKEND_DIR = PROJECT_ROOT / 'backend'
EXCEL_FILES_DIR = BACKEND_DIR / 'Excel Files'


def normalize_npi(value):
    """
    Normalize NPI value for comparison (handle None, NaN, strings, numbers)
    
    Args:
        value: The NPI value to normalize
        
    Returns:
        str: Normalized string value or empty string
    """
    if value is None:
        return ""
    
    # Convert to string and strip whitespace
    normalized = str(value).strip()
    
    # Handle float values (e.g., 1234567890.0 -> 1234567890)
    if isinstance(value, float):
        if value.is_integer():
            normalized = str(int(value))
        else:
            normalized = str(value)
    
    # Remove .0 suffix if present
    if normalized.endswith('.0'):
        normalized = normalized[:-2]
    
    return normalized


def get_unique_values(values_list):
    """
    Get unique non-empty values from a list, preserving order
    
    Args:
        values_list: List of values
        
    Returns:
        list: List of unique non-empty values
    """
    seen = set()
    unique_values = []
    
    for value in values_list:
        normalized = normalize_npi(value)
        if normalized and normalized not in seen:
            seen.add(normalized)
            unique_values.append(value)  # Keep original value, not normalized
    
    return unique_values


def merge_duplicate_providers():
    """
    Merge duplicate providers in Template copy.xlsx based on NPI Number.
    Consolidates unique values from Professional Suffix n, Specialty n, and Location ID n columns.
    Adds a border to the NPI cell to indicate merging.
    
    Returns:
        bool: True if successful, False otherwise
    """
    template_file = EXCEL_FILES_DIR / 'Template copy.xlsx'
    
    # Check if template file exists
    if not template_file.exists():
        print(f"Error: Template copy.xlsx not found at {template_file}")
        return False
    
    try:
        # Load the template workbook
        wb = load_workbook(template_file)
        
        # Check if 'Provider' sheet exists
        if 'Provider' not in wb.sheetnames:
            print("Error: 'Provider' sheet not found in Template copy.xlsx")
            return False
        
        # Get the Provider sheet
        provider_sheet = wb['Provider']
        header_row = 1
        max_row = provider_sheet.max_row
        
        # Find column headers
        npi_column_letter = None
        professional_suffix_columns = {}  # {1: 'A', 2: 'B', ...}
        specialty_columns = {}  # {1: 'A', 2: 'B', ...}
        location_id_columns = {}  # {1: 'A', 2: 'B', ...}
        
        # Search for headers in the first row
        for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
            if cell.value:
                cell_value = str(cell.value).strip().lower()
                
                if cell_value == 'npi number':
                    npi_column_letter = get_column_letter(col_idx)
                
                # Find Professional Suffix columns
                if cell_value.startswith('professional suffix'):
                    # Extract number from "Professional Suffix 1", "Professional Suffix 2", etc.
                    if cell_value == 'professional suffix':
                        professional_suffix_columns[1] = get_column_letter(col_idx)
                    else:
                        try:
                            # Remove "professional suffix" and get the number
                            number_str = cell_value.replace('professional suffix', '').strip()
                            if number_str:
                                number = int(number_str)
                                professional_suffix_columns[number] = get_column_letter(col_idx)
                        except:
                            pass
                
                # Find Specialty columns
                if cell_value.startswith('specialty'):
                    if cell_value == 'specialty':
                        specialty_columns[1] = get_column_letter(col_idx)
                    else:
                        try:
                            number_str = cell_value.replace('specialty', '').strip()
                            if number_str:
                                number = int(number_str)
                                specialty_columns[number] = get_column_letter(col_idx)
                        except:
                            pass
                
                # Find Location ID columns
                if cell_value.startswith('location id'):
                    if cell_value == 'location id':
                        location_id_columns[1] = get_column_letter(col_idx)
                    else:
                        try:
                            number_str = cell_value.replace('location id', '').strip()
                            if number_str:
                                number = int(number_str)
                                location_id_columns[number] = get_column_letter(col_idx)
                        except:
                            pass
        
        # Check if NPI column was found
        if npi_column_letter is None:
            print("Error: 'NPI Number' column not found in Provider sheet")
            return False
        
        # Find maximum column numbers for each type
        max_prof_suffix = max(professional_suffix_columns.keys()) if professional_suffix_columns else 0
        max_specialty = max(specialty_columns.keys()) if specialty_columns else 0
        max_location_id = max(location_id_columns.keys()) if location_id_columns else 0
        
        # Build a dictionary to group rows by NPI
        npi_to_rows = {}
        
        # Collect all rows with their NPI values
        for row_idx in range(2, max_row + 1):  # Start from row 2 (skip header)
            npi_cell = provider_sheet[f"{npi_column_letter}{row_idx}"]
            npi_value = npi_cell.value
            npi_key = normalize_npi(npi_value)
            
            if npi_key:  # Only process rows with NPI
                if npi_key not in npi_to_rows:
                    npi_to_rows[npi_key] = []
                npi_to_rows[npi_key].append(row_idx)
        
        # Find duplicates (NPIs with more than one row)
        duplicates = {npi: rows for npi, rows in npi_to_rows.items() if len(rows) > 1}
        
        if not duplicates:
            print("No duplicate providers found based on NPI Number")
            return True
        
        # Border style for merged NPI cells
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Process each duplicate NPI
        rows_to_delete = []
        
        for npi_key, row_indices in duplicates.items():
            # Sort row indices to keep the first one
            row_indices.sort()
            keep_row = row_indices[0]  # Keep the first row
            merge_rows = row_indices[1:]  # Merge these rows into the first
            
            # Collect all values from merge_rows
            prof_suffix_values = []
            specialty_values = []
            location_id_values = []
            
            # First, get values from the keep_row
            for num in range(1, max_prof_suffix + 1):
                if num in professional_suffix_columns:
                    cell = provider_sheet[f"{professional_suffix_columns[num]}{keep_row}"]
                    if cell.value:
                        prof_suffix_values.append(cell.value)
            
            for num in range(1, max_specialty + 1):
                if num in specialty_columns:
                    cell = provider_sheet[f"{specialty_columns[num]}{keep_row}"]
                    if cell.value:
                        specialty_values.append(cell.value)
            
            for num in range(1, max_location_id + 1):
                if num in location_id_columns:
                    cell = provider_sheet[f"{location_id_columns[num]}{keep_row}"]
                    if cell.value:
                        location_id_values.append(cell.value)
            
            # Then, get values from merge_rows
            for merge_row in merge_rows:
                for num in range(1, max_prof_suffix + 1):
                    if num in professional_suffix_columns:
                        cell = provider_sheet[f"{professional_suffix_columns[num]}{merge_row}"]
                        if cell.value:
                            prof_suffix_values.append(cell.value)
                
                for num in range(1, max_specialty + 1):
                    if num in specialty_columns:
                        cell = provider_sheet[f"{specialty_columns[num]}{merge_row}"]
                        if cell.value:
                            specialty_values.append(cell.value)
                
                for num in range(1, max_location_id + 1):
                    if num in location_id_columns:
                        cell = provider_sheet[f"{location_id_columns[num]}{merge_row}"]
                        if cell.value:
                            location_id_values.append(cell.value)
            
            # Get unique values
            unique_prof_suffix = get_unique_values(prof_suffix_values)
            unique_specialty = get_unique_values(specialty_values)
            unique_location_id = get_unique_values(location_id_values)
            
            # Clear existing values in keep_row
            for num in range(1, max_prof_suffix + 1):
                if num in professional_suffix_columns:
                    provider_sheet[f"{professional_suffix_columns[num]}{keep_row}"].value = None
            
            for num in range(1, max_specialty + 1):
                if num in specialty_columns:
                    provider_sheet[f"{specialty_columns[num]}{keep_row}"].value = None
            
            for num in range(1, max_location_id + 1):
                if num in location_id_columns:
                    provider_sheet[f"{location_id_columns[num]}{keep_row}"].value = None
            
            # Write unique values to keep_row, distributing across columns
            for idx, value in enumerate(unique_prof_suffix, start=1):
                if idx <= max_prof_suffix and idx in professional_suffix_columns:
                    provider_sheet[f"{professional_suffix_columns[idx]}{keep_row}"].value = value
            
            for idx, value in enumerate(unique_specialty, start=1):
                if idx <= max_specialty and idx in specialty_columns:
                    provider_sheet[f"{specialty_columns[idx]}{keep_row}"].value = value
            
            for idx, value in enumerate(unique_location_id, start=1):
                if idx <= max_location_id and idx in location_id_columns:
                    provider_sheet[f"{location_id_columns[idx]}{keep_row}"].value = value
            
            # Add border to NPI cell
            npi_cell = provider_sheet[f"{npi_column_letter}{keep_row}"]
            npi_cell.border = border
            
            # Mark merge_rows for deletion
            rows_to_delete.extend(merge_rows)
        
        # Delete duplicate rows (in reverse order to maintain row indices)
        rows_to_delete.sort(reverse=True)
        for row_idx in rows_to_delete:
            provider_sheet.delete_rows(row_idx)
        
        # Save the workbook
        wb.save(template_file)
        return True
        
    except Exception as e:
        print(f"Error in merge_duplicate_providers: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    merge_duplicate_providers()

