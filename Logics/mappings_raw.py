"""
Mappings Raw - Creates Excel file from mappings data

This script creates an Excel file called 'Mappings.xlsx' in the backend/Excel Files folder
containing all the mappings from the Mappings table.
"""

import os
import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Get the project root directory (parent of Logics folder)
LOGICS_DIR = Path(__file__).parent
PROJECT_ROOT = LOGICS_DIR.parent
BACKEND_DIR = PROJECT_ROOT / 'backend'
EXCEL_FILES_DIR = BACKEND_DIR / 'Excel Files'

def create_mappings_excel(mappings_data, output_filename='Mappings.xlsx'):
    """
    Create Excel file from mappings data
    
    Args:
        mappings_data: List of mapping objects with structure:
            [
                {
                    'columnName': 'Column Name',
                    'detectedAs': 'npi' or ['npi', 'firstName'] (single string or array)
                },
                ...
            ]
        output_filename: Name of the output Excel file (default: 'Mappings.xlsx')
    
    Returns:
        Path to the created Excel file
    """
    # Ensure Excel Files directory exists
    EXCEL_FILES_DIR.mkdir(parents=True, exist_ok=True)
    
    # Prepare data for DataFrame
    rows = []
    for mapping in mappings_data:
        column_name = mapping.get('columnName', '')
        detected_as = mapping.get('detectedAs', '')
        
        # Handle both single string and array formats
        if isinstance(detected_as, list):
            detected_as_str = ', '.join(detected_as)
        else:
            detected_as_str = str(detected_as)
        
        # Map detected_as to readable names
        readable_names = {
            'npi': 'NPI Number',
            'firstName': 'First Name',
            'lastName': 'Last Name',
            'gender': 'Gender',
            'professionalSuffix': 'Professional Suffix 1-3',
            'headshot': 'Headshot Link',
            'additionalLanguages': 'Additional Languages Spoken 1-3',
            'patientsAccepted': 'Patients Accepted',
            'specialty': 'Specialty',
            'locationId': 'Location ID',
            'locationName': 'Location Name',
            'locationTypeRaw': 'Location Type_Raw',
            'practiceId': 'Practice ID',
            'practiceCloudId': 'Practice Cloud ID',
            'practiceName': 'Practice Name',
            'addressLine1': 'Address Line 1',
            'addressLine2': 'Address Line 2',
            'city': 'City',
            'state': 'State',
            'zip': 'ZIP'
        }
        
        # Convert detected_as to readable format
        if isinstance(detected_as, list):
            readable_detected_as = ', '.join([readable_names.get(d, d) for d in detected_as])
        else:
            readable_detected_as = readable_names.get(detected_as, detected_as)
        
        rows.append({
            'Column Name': column_name,
            'Detected As': readable_detected_as,
            'Raw Type': detected_as_str
        })
    
    # Create DataFrame
    df = pd.DataFrame(rows)
    
    # Create Excel file path
    excel_path = EXCEL_FILES_DIR / output_filename
    
    # Write to Excel with formatting
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Mappings', index=False)
        
        # Get the workbook and worksheet for formatting
        workbook = writer.book
        worksheet = writer.sheets['Mappings']
        
        # Format header row
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    return str(excel_path)

if __name__ == '__main__':
    # Test the function
    test_mappings = [
        {'columnName': 'NPI', 'detectedAs': 'npi'},
        {'columnName': 'First Name', 'detectedAs': 'firstName'},
        {'columnName': 'Last Name', 'detectedAs': 'lastName'},
        {'columnName': 'Headshot', 'detectedAs': 'headshot'}
    ]
    
    result_path = create_mappings_excel(test_mappings)
    print(f"Excel file created at: {result_path}")

