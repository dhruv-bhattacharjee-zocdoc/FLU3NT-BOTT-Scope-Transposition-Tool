"""
Additional Languages - Extracts Additional Languages Spoken column from _Mapped.xlsx and writes to Template copy.xlsx

This script reads the 'Additional Languages Spoken 1-3' column from backend\Excel Files\_Mapped.xlsx
and writes it to the 'Additional Language Spoken 1', 'Additional Language Spoken 2', and 'Additional Language Spoken 3' columns
in the 'Provider' sheet in Template copy.xlsx. Values can be separated by comma, semicolon, plus sign, or other separators.
It also changes the header colors to green.
"""

import os
import re
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Get the project root directory (parent of Transposition Logics folder)
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
BACKEND_DIR = PROJECT_ROOT / 'backend'
EXCEL_FILES_DIR = BACKEND_DIR / 'Excel Files'

def normalize_language(lang_str):
    """
    Normalize a language string by removing punctuation, spaces, and converting to uppercase.
    Example: "English" -> "ENGLISH", "Eng lish" -> "ENGLISH"
    """
    if not lang_str:
        return ""
    # Remove all punctuation and spaces, convert to uppercase
    normalized = re.sub(r'[^\w]', '', str(lang_str).upper())
    return normalized

def find_close_match(value, valid_languages):
    """
    Find a close match for a value in the valid languages list.
    Returns the original valid language if a close match is found, None otherwise.
    Checks for:
    1. Exact match (case-insensitive) - no replacement needed
    2. Normalized match (without punctuation/spaces) - e.g., "Eng lish" matches "English"
    3. Prefix match - e.g., "ABC" matches "ABCd"
    """
    if not value or not valid_languages:
        return None
    
    normalized_value = normalize_language(value)
    
    # First check exact match (case-insensitive)
    for valid_language in valid_languages:
        if str(value).strip().lower() == str(valid_language).strip().lower():
            return None  # Exact match, no need to replace
    
    # Check normalized match (without punctuation/spaces)
    for valid_language in valid_languages:
        normalized_valid = normalize_language(valid_language)
        if normalized_value == normalized_valid:
            return valid_language  # Close match found, return the valid version
    
    # Check if normalized input is a prefix of any normalized valid language
    # Only match if the input is at least 2 characters to avoid too broad matches
    if len(normalized_value) >= 2:
        for valid_language in valid_languages:
            normalized_valid = normalize_language(valid_language)
            # Check if input is a prefix of valid language
            if normalized_valid.startswith(normalized_value):
                return valid_language  # Prefix match found, return the valid version
    
    return None

def extract_languages_from_string(lang_str, valid_languages):
    """
    Extract languages from a string by matching against valid languages list.
    Handles various separators: comma, semicolon, plus sign, "and", etc.
    Also handles cases where languages are embedded in longer text like "i speak English, Hindi and Chinese".
    Returns a list of up to 3 matched languages.
    """
    if not lang_str or not valid_languages:
        return [None, None, None]
    
    lang_str = str(lang_str).strip()
    matched_languages = []
    
    # First, try to find valid languages anywhere in the string (as whole words)
    # This handles cases like "i speak English, Hindi and Chinese"
    lang_str_lower = lang_str.lower()
    for valid_lang in valid_languages:
        valid_lang_str = str(valid_lang).strip()
        if not valid_lang_str:
            continue
        
        # Check if the valid language appears as a whole word in the string
        # Use word boundaries to avoid partial matches
        pattern = r'\b' + re.escape(valid_lang_str.lower()) + r'\b'
        if re.search(pattern, lang_str_lower):
            if valid_lang not in matched_languages:
                matched_languages.append(valid_lang)
                if len(matched_languages) >= 3:
                    break
    
    # If we found matches, return them (padded to 3)
    if matched_languages:
        while len(matched_languages) < 3:
            matched_languages.append(None)
        return matched_languages[:3]
    
    # If no matches found with whole word search, try splitting approach
    # Replace "and" with comma for easier parsing
    lang_str_normalized = re.sub(r'\s+and\s+', ', ', lang_str, flags=re.IGNORECASE)
    
    # Split by various separators
    separators = [';', ',', '+', '|', '/']
    parts = [lang_str_normalized]
    
    for sep in separators:
        if sep in lang_str_normalized:
            parts = [part.strip() for part in lang_str_normalized.split(sep)]
            break
    
    # If no separator found, try to match the whole string
    if len(parts) == 1 and parts[0] == lang_str_normalized:
        parts = [lang_str.strip()]
    
    # For each part, try to find a match in valid languages
    for part in parts:
        part = part.strip()
        if not part:
            continue
        
        # Try exact match first
        match_found = None
        for valid_lang in valid_languages:
            if part.lower() == str(valid_lang).lower():
                match_found = valid_lang
                break
        
        # If no exact match, try fuzzy matching
        if not match_found:
            match_found = find_close_match(part, valid_languages)
        
        # If still no match, try to see if the part contains a valid language
        if not match_found:
            # Check if any valid language is contained in the part (or vice versa)
            part_normalized = normalize_language(part)
            for valid_lang in valid_languages:
                valid_lang_normalized = normalize_language(valid_lang)
                # Check if part is contained in valid language or vice versa
                if part_normalized in valid_lang_normalized or valid_lang_normalized in part_normalized:
                    if len(part_normalized) >= 2:  # At least 2 characters to avoid false matches
                        match_found = valid_lang
                        break
        
        if match_found and match_found not in matched_languages:
            matched_languages.append(match_found)
            if len(matched_languages) >= 3:
                break
    
    # Pad to 3 elements
    while len(matched_languages) < 3:
        matched_languages.append(None)
    
    return matched_languages[:3]

def extract_additional_languages_to_template():
    """
    Extract Additional Languages Spoken 1-3 column from _Mapped.xlsx and write to Template copy.xlsx
    Values can be separated by comma, semicolon, plus sign, or other separators.
    Values are split across Additional Language Spoken 1, 2, and 3 columns.
    
    Returns:
        bool: True if successful, False otherwise
    """
    # File paths
    source_file = EXCEL_FILES_DIR / '_Mapped.xlsx'
    template_file = EXCEL_FILES_DIR / 'Template copy.xlsx'
    
    # Check if source file exists
    if not source_file.exists():
        return False
    
    # Check if template file exists
    if not template_file.exists():
        return False
    
    try:
        # Read the Additional Languages Spoken 1-3 column from _Mapped.xlsx
        mapped_df = pd.read_excel(source_file)
        
        # Check if 'Additional Languages Spoken 1-3' column exists
        if 'Additional Languages Spoken 1-3' not in mapped_df.columns:
            return False
        
        # Extract the Additional Languages Spoken 1-3 column
        languages_data = mapped_df['Additional Languages Spoken 1-3'].tolist()
        
        # Load the template workbook
        wb = load_workbook(template_file)
        
        # Check if 'Provider' sheet exists
        if 'Provider' not in wb.sheetnames:
            return False
        
        # Get the Provider sheet
        provider_sheet = wb['Provider']
        
        # Find the Additional Language Spoken column headers
        header_row = 1
        lang_columns = {}
        
        # Search for the headers in the first row
        for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
            if cell.value:
                cell_value = str(cell.value).strip().lower()
                # Check for both singular and plural variations
                if cell_value == 'additional language spoken 1' or cell_value == 'additional languages spoken 1':
                    lang_columns[1] = get_column_letter(col_idx)
                elif cell_value == 'additional language spoken 2' or cell_value == 'additional languages spoken 2':
                    lang_columns[2] = get_column_letter(col_idx)
                elif cell_value == 'additional language spoken 3' or cell_value == 'additional languages spoken 3':
                    lang_columns[3] = get_column_letter(col_idx)
        
        if 1 not in lang_columns:
            return False
        
        # Load valid languages from ValidationAndReference sheet
        valid_languages = set()
        if 'ValidationAndReference' in wb.sheetnames:
            validation_sheet = wb['ValidationAndReference']
            lang_column_letter = None
            
            # Find the 'Lang Description' column in ValidationAndReference sheet
            for col_idx, cell in enumerate(validation_sheet[header_row], start=1):
                if cell.value and str(cell.value).strip().lower() == 'lang descripton':
                    lang_column_letter = get_column_letter(col_idx)
                    break
                # Also try alternative spelling
                if cell.value and str(cell.value).strip().lower() == 'lang description':
                    lang_column_letter = get_column_letter(col_idx)
                    break
            
            # Read all valid language values
            if lang_column_letter:
                max_row = validation_sheet.max_row
                for row_idx in range(2, max_row + 1):  # Start from row 2 (skip header)
                    cell = validation_sheet[f"{lang_column_letter}{row_idx}"]
                    if cell.value:
                        valid_languages.add(str(cell.value).strip())
        
        # Grey fill for invalid values
        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Process and write the Additional Languages data
        for row_idx, languages_value in enumerate(languages_data, start=2):
            # Extract languages by matching against valid languages list
            if pd.isna(languages_value):
                lang_parts = [None, None, None]
            else:
                lang_parts = extract_languages_from_string(languages_value, valid_languages)
            
            # Remove "English" from the list (case-insensitive)
            lang_parts = [lang for lang in lang_parts if lang is None or str(lang).strip().lower() != 'english']
            # Pad back to 3 elements
            while len(lang_parts) < 3:
                lang_parts.append(None)
            lang_parts = lang_parts[:3]
            
            # Check if there are multiple values
            non_empty_count = sum(1 for p in lang_parts if p is not None)
            has_multiple_values = non_empty_count > 1
            
            # Write to Additional Language Spoken 1 through 3
            for i in range(1, 4):
                if i in lang_columns:
                    cell = provider_sheet[f"{lang_columns[i]}{row_idx}"]
                    value = lang_parts[i-1] if lang_parts[i-1] else None
                    
                    # All extracted languages are already matched against valid languages
                    # So we just write them, but highlight if there were multiple values
                    cell.value = value
                    
                    # Highlight if has multiple values (split across columns)
                    if has_multiple_values and value:
                        cell.fill = grey_fill
        
        # Color the header cells green
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        
        for i in range(1, 4):
            if i in lang_columns:
                header = provider_sheet[f"{lang_columns[i]}{header_row}"]
                header.fill = green_fill
        
        # Save the workbook
        wb.save(template_file)
        return True
        
    except Exception as e:
        return False

if __name__ == "__main__":
    extract_additional_languages_to_template()

