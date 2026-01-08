"""
Name - Extracts First Name, Middle Name, and Last Name columns from _Mapped.xlsx and writes to Template copy.xlsx

This script reads the 'First Name', 'Middle Name' (optional), and 'Last Name' columns from backend\Excel Files\_Mapped.xlsx
and writes them to the 'Provider' sheet in Template copy.xlsx. 
- If Last Name contains multiple words (e.g., "Lewis Mayor"), the first word is added to First Name
- Middle Name is combined with First Name
- It also changes the header colors to green.
"""

import os
import re
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Get the project root directory (parent of Transposition Logics folder)
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
BACKEND_DIR = PROJECT_ROOT / 'backend'
EXCEL_FILES_DIR = BACKEND_DIR / 'Excel Files'

def extract_names_to_template():
    """
    Extract First Name, Middle Name (optional), and Last Name columns from _Mapped.xlsx and write to Template copy.xlsx
    - If Last Name has multiple words (e.g., "Lewis Mayor"), the first word is added to First Name (e.g., "John" + "Lewis" = "John Lewis")
    - Middle Name is combined with First Name (e.g., "John F" if First Name is "John" and Middle Name is "F")
    
    Returns:
        bool: True if successful, False otherwise
    """
    # File paths
    source_file = EXCEL_FILES_DIR / '_Mapped.xlsx'
    template_file = EXCEL_FILES_DIR / 'Template copy.xlsx'
    
    # Check if source file exists
    if not source_file.exists():
        return False
    
    # Check if template file exists
    if not template_file.exists():
        return False
    
    try:
        # Read the columns from _Mapped.xlsx
        mapped_df = pd.read_excel(source_file)
        
        # Check if required columns exist
        missing_columns = []
        if 'First Name' not in mapped_df.columns:
            missing_columns.append('First Name')
        if 'Last Name' not in mapped_df.columns:
            missing_columns.append('Last Name')
        
        if missing_columns:
            return False
        
        # Check if Middle Name column exists (optional)
        has_middle_name = 'Middle Name' in mapped_df.columns
        
        # Extract the columns
        first_name_data = mapped_df['First Name'].tolist()
        middle_name_data = mapped_df['Middle Name'].tolist() if has_middle_name else [None] * len(first_name_data)
        last_name_data = mapped_df['Last Name'].tolist()
        
        # Process names: handle cases where Last Name contains multiple words
        processed_first_name_data = []
        processed_last_name_data = []
        
        for i in range(len(first_name_data)):
            first_name = first_name_data[i]
            last_name = last_name_data[i]
            
            # Convert to strings and handle NaN/empty values
            first_name_str = str(first_name).strip() if not pd.isna(first_name) and first_name else ""
            last_name_str = str(last_name).strip() if not pd.isna(last_name) and last_name else ""
            
            # If Last Name has multiple words, take the first word and add it to First Name
            if last_name_str:
                name_parts = last_name_str.split()
                if len(name_parts) >= 2:
                    # First word goes to First Name, rest goes to Last Name
                    first_word_from_last = name_parts[0]
                    remaining_last_name = ' '.join(name_parts[1:])
                    
                    # Add first word from Last Name to First Name
                    if first_name_str:
                        first_name_str = f"{first_name_str} {first_word_from_last}".strip()
                    else:
                        first_name_str = first_word_from_last
                    
                    last_name_str = remaining_last_name
            
            processed_first_name_data.append(first_name_str if first_name_str else None)
            processed_last_name_data.append(last_name_str if last_name_str else None)
        
        # Combine First Name and Middle Name
        combined_first_name_data = []
        for i, first_name_str in enumerate(processed_first_name_data):
            first_name_str = first_name_str if first_name_str else ""
            middle_name_str = str(middle_name_data[i]).strip() if has_middle_name and not pd.isna(middle_name_data[i]) and middle_name_data[i] else ""
            
            # Combine First Name and Middle Name with a space
            if first_name_str and middle_name_str:
                combined_name = f"{first_name_str} {middle_name_str}".strip()
            elif first_name_str:
                combined_name = first_name_str
            else:
                combined_name = ""
            
            combined_first_name_data.append(combined_name if combined_name else None)
        
        # Load the template workbook
        wb = load_workbook(template_file)
        
        # Check if 'Provider' sheet exists
        if 'Provider' not in wb.sheetnames:
            return False
        
        # Get the Provider sheet
        provider_sheet = wb['Provider']
        
        # Find the column headers
        header_row = 1
        first_name_column_letter = None
        last_name_column_letter = None
        
        # Search for the headers in the first row
        for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
            if cell.value:
                cell_value = str(cell.value).strip().lower()
                if cell_value == 'first name':
                    first_name_column_letter = get_column_letter(col_idx)
                elif cell_value == 'last name':
                    last_name_column_letter = get_column_letter(col_idx)
        
        if first_name_column_letter is None:
            return False
        
        if last_name_column_letter is None:
            return False
        
        # Helper function to check if a value contains symbols (except '.')
        def contains_symbols_except_period(value):
            """Check if value contains any symbols except period (.)"""
            if value is None:
                return False
            value_str = str(value)
            # Check if value contains any character that is not alphanumeric, space, or period
            # This will match '-', ',', '/', and any other symbols except '.'
            return bool(re.search(r'[^a-zA-Z0-9\s.]', value_str))
        
        # Light grey fill for cells with symbols
        light_grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Write the combined First Name (with Middle Name) data to the column (starting from row 2, as row 1 is the header)
        for row_idx, combined_name_value in enumerate(combined_first_name_data, start=2):
            cell = provider_sheet[f"{first_name_column_letter}{row_idx}"]
            # The value is already processed, just set it
            cell.value = combined_name_value
            
            # Check if cell contains symbols (except '.') and highlight in light grey
            if combined_name_value and contains_symbols_except_period(combined_name_value):
                cell.fill = light_grey_fill
        
        # Write the processed Last Name data to the column (starting from row 2, as row 1 is the header)
        for row_idx, last_name_value in enumerate(processed_last_name_data, start=2):
            cell = provider_sheet[f"{last_name_column_letter}{row_idx}"]
            # The value is already processed, just set it
            cell.value = last_name_value
            
            # Check if cell contains symbols (except '.') and highlight in light grey
            if last_name_value and contains_symbols_except_period(last_name_value):
                cell.fill = light_grey_fill
        
        # Color the header cells green
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        
        first_name_header = provider_sheet[f"{first_name_column_letter}{header_row}"]
        first_name_header.fill = green_fill
        
        last_name_header = provider_sheet[f"{last_name_column_letter}{header_row}"]
        last_name_header.fill = green_fill
        
        # Save the workbook
        wb.save(template_file)
        return True
        
    except Exception as e:
        return False

if __name__ == "__main__":
    extract_names_to_template()

