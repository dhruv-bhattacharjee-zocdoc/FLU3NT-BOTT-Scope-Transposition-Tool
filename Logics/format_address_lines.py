"""
Format Address Lines - Formats address lines in Locations_input.xlsx and performs fuzzy matching

This script handles cases where the uploaded sheet has the whole address in one column.
It formats the address by converting directions and street suffixes to abbreviations,
then performs fuzzy matching with Practice_Locations.xlsx.
"""

import os
import sys
import re
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from difflib import SequenceMatcher

# Get the project root directory (parent of Logics folder)
LOGICS_DIR = Path(__file__).parent
PROJECT_ROOT = LOGICS_DIR.parent
BACKEND_DIR = PROJECT_ROOT / 'backend'
EXCEL_FILES_DIR = BACKEND_DIR / 'Excel Files'

# Directional abbreviations mapping
DIRECTIONAL_ABBREVIATIONS = {
    'north': 'N',
    'south': 'S',
    'east': 'E',
    'west': 'W',
    'northeast': 'NE',
    'north east': 'NE',
    'north-east': 'NE',
    'northwest': 'NW',
    'north west': 'NW',
    'north-west': 'NW',
    'southeast': 'SE',
    'south east': 'SE',
    'south-east': 'SE',
    'southwest': 'SW',
    'south west': 'SW',
    'south-west': 'SW',
}

# Load street suffix abbreviations from Excel file
def load_street_suffix_abbreviations():
    """
    Load street suffix abbreviations from C1 Street Suffix Abbreviations.xlsx
    
    Returns:
        Dictionary mapping full form to abbreviation (e.g., {'Parkway': 'Pkwy', 'Street': 'St'})
    """
    abbreviations_file = BACKEND_DIR / 'C1 Street Suffix Abbreviations.xlsx'
    
    if not abbreviations_file.exists():
        print(f"Warning: Street suffix abbreviations file not found at {abbreviations_file}")
        return {}
    
    try:
        df = pd.read_excel(abbreviations_file)
        
        # Find the correct column names (case-insensitive)
        full_form_col = None
        abbrev_col = None
        
        for col in df.columns:
            col_lower = col.lower()
            if 'commonly used' in col_lower or 'street suffix' in col_lower or 'abbreviation' in col_lower:
                if full_form_col is None:
                    full_form_col = col
            if 'postal service' in col_lower or 'standard suffix' in col_lower:
                if abbrev_col is None:
                    abbrev_col = col
        
        if full_form_col is None or abbrev_col is None:
            print(f"Warning: Could not find required columns in {abbreviations_file}")
            print(f"Available columns: {', '.join(df.columns)}")
            return {}
        
        # Create mapping dictionary
        suffix_map = {}
        for _, row in df.iterrows():
            full_form = str(row[full_form_col]).strip() if pd.notna(row[full_form_col]) else ''
            abbrev = str(row[abbrev_col]).strip() if pd.notna(row[abbrev_col]) else ''
            
            if full_form and abbrev:
                suffix_map[full_form.lower()] = abbrev
        
        print(f"Loaded {len(suffix_map)} street suffix abbreviations")
        return suffix_map
    
    except Exception as e:
        print(f"Warning: Error loading street suffix abbreviations: {str(e)}")
        return {}


def normalize_string(s):
    """
    Normalize a string: lowercase, strip, remove extra spaces
    """
    if pd.isna(s) or s == '':
        return ''
    s = str(s).strip().lower()
    s = ' '.join(s.split())
    return s


def extract_numbers(s):
    """
    Extract all numbers from a string and return them as a list of strings
    """
    if pd.isna(s) or s == '':
        return []
    s = str(s)
    numbers = re.findall(r'\d+', s)
    return numbers


def convert_directions_to_abbreviations(address):
    """
    Convert directional words to abbreviations in an address string
    
    Args:
        address: Address string
        
    Returns:
        Address string with directions converted to abbreviations
    """
    if pd.isna(address) or address == '':
        return address
    
    address_str = str(address)
    
    # Convert to lowercase for matching, but preserve original case structure
    address_lower = address_str.lower()
    
    # Sort by length (longest first) to match compound directions first
    sorted_directions = sorted(DIRECTIONAL_ABBREVIATIONS.items(), key=lambda x: len(x[0]), reverse=True)
    
    for full_form, abbrev in sorted_directions:
        # Use word boundaries to match whole words only
        pattern = r'\b' + re.escape(full_form) + r'\b'
        if re.search(pattern, address_lower, re.IGNORECASE):
            # Replace with abbreviation, preserving case of first letter
            def replace_func(match):
                matched_text = match.group(0)
                # Preserve capitalization: if first letter is uppercase, make abbreviation uppercase
                if matched_text[0].isupper():
                    return abbrev.upper()
                else:
                    return abbrev.lower()
            
            address_str = re.sub(pattern, replace_func, address_str, flags=re.IGNORECASE)
            address_lower = address_str.lower()
    
    return address_str


def convert_street_suffixes_to_abbreviations(address, suffix_map):
    """
    Convert street suffixes to abbreviations using the suffix map
    
    Args:
        address: Address string
        suffix_map: Dictionary mapping full form to abbreviation
        
    Returns:
        Address string with suffixes converted to abbreviations
    """
    if pd.isna(address) or address == '' or not suffix_map:
        return address
    
    address_str = str(address)
    address_lower = address_str.lower()
    
    # Sort by length (longest first) to match longer suffixes first
    sorted_suffixes = sorted(suffix_map.items(), key=lambda x: len(x[0]), reverse=True)
    
    for full_form, abbrev in sorted_suffixes:
        # Match at the end of a word (likely a street suffix)
        pattern = r'\b' + re.escape(full_form) + r'\b'
        if re.search(pattern, address_lower, re.IGNORECASE):
            # Replace with abbreviation, preserving case
            def replace_func(match):
                matched_text = match.group(0)
                if matched_text[0].isupper():
                    return abbrev.upper()
                else:
                    return abbrev.lower()
            
            address_str = re.sub(pattern, replace_func, address_str, flags=re.IGNORECASE)
            address_lower = address_str.lower()
    
    return address_str


def format_address_line1(address):
    """
    Format Address Line 1 by converting directions and street suffixes to abbreviations
    
    Args:
        address: Address string to format
        
    Returns:
        Formatted address string
    """
    if pd.isna(address) or address == '':
        return address
    
    # Load street suffix abbreviations
    suffix_map = load_street_suffix_abbreviations()
    
    # Step 1: Convert directions to abbreviations
    formatted = convert_directions_to_abbreviations(address)
    
    # Step 2: Convert street suffixes to abbreviations
    formatted = convert_street_suffixes_to_abbreviations(formatted, suffix_map)
    
    return formatted


def fuzzy_match_score(str1, str2, threshold=0.8):
    """
    Calculate fuzzy match score between two strings using SequenceMatcher.
    Numbers are not considered in the fuzzy matching - if addresses have different numbers,
    they will not match.
    
    Returns True if similarity >= threshold, False otherwise
    """
    str1_norm = normalize_string(str1)
    str2_norm = normalize_string(str2)
    
    if not str1_norm or not str2_norm:
        return False
    
    # Extract numbers from both strings
    numbers1 = extract_numbers(str1)
    numbers2 = extract_numbers(str2)
    
    # If both strings have numbers, they must match exactly
    if numbers1 and numbers2:
        if numbers1 != numbers2:
            return False
    
    # Remove numbers from both strings for fuzzy matching
    str1_no_numbers = re.sub(r'\d+', '', str1_norm).strip()
    str2_no_numbers = re.sub(r'\d+', '', str2_norm).strip()
    
    # Remove extra spaces after removing numbers
    str1_no_numbers = ' '.join(str1_no_numbers.split())
    str2_no_numbers = ' '.join(str2_no_numbers.split())
    
    if not str1_no_numbers or not str2_no_numbers:
        # If after removing numbers one string is empty, check exact match of original
        return str1_norm == str2_norm
    
    # Exact match after normalization and number removal
    if str1_no_numbers == str2_no_numbers:
        return True
    
    # Calculate similarity ratio on strings without numbers
    similarity = SequenceMatcher(None, str1_no_numbers, str2_no_numbers).ratio()
    return similarity >= threshold


def combine_api_address(api_row):
    """
    Combine address components from API row into a single string
    
    Args:
        api_row: Row from Practice_Locations.xlsx Locations sheet
        
    Returns:
        Combined address string
    """
    components = []
    
    addr1 = api_row.get('address_1', '')
    if pd.notna(addr1) and str(addr1).strip():
        components.append(str(addr1).strip())
    
    addr2 = api_row.get('address_2', '')
    if pd.notna(addr2) and str(addr2).strip():
        components.append(str(addr2).strip())
    
    city = api_row.get('city', '')
    if pd.notna(city) and str(city).strip():
        components.append(str(city).strip())
    
    state = api_row.get('state', '')
    if pd.notna(state) and str(state).strip():
        components.append(str(state).strip())
    
    zip_code = api_row.get('zip', '')
    if pd.notna(zip_code) and str(zip_code).strip():
        components.append(str(zip_code).strip())
    
    return ' '.join(components)


def format_locations_input_addresses(manual_practice_ids=''):
    """
    Format address lines in Locations_input.xlsx and perform fuzzy matching
    
    Args:
        manual_practice_ids: Optional string of manually entered Practice IDs (comma, semicolon, or space separated)
    """
    locations_input_file = EXCEL_FILES_DIR / 'Locations_input.xlsx'
    practice_locations_file = EXCEL_FILES_DIR / 'Practice_Locations.xlsx'
    
    if not locations_input_file.exists():
        print(f"Warning: Locations_input.xlsx not found at {locations_input_file}")
        return
    
    try:
        # Read Locations_input.xlsx
        locations_df = pd.read_excel(locations_input_file, sheet_name='Locations')
        
        # Check if Address Line 1 has values but other address components are mostly empty
        has_addr1 = 'Address Line 1' in locations_df.columns
        has_addr2 = 'Address Line 2' in locations_df.columns
        has_city = 'City' in locations_df.columns
        has_state = 'State' in locations_df.columns
        has_zip = 'ZIP' in locations_df.columns
        has_practice_id = 'Practice ID' in locations_df.columns
        
        if not has_addr1:
            print("Address Line 1 column not found. Skipping address formatting.")
            return
        
        # Check if Address Line 1 has values but other components are empty
        addr1_has_values = locations_df['Address Line 1'].notna().any() and (locations_df['Address Line 1'].astype(str).str.strip() != '').any()
        
        addr2_empty = True
        city_empty = True
        state_empty = True
        zip_empty = True
        
        if has_addr2:
            addr2_empty = locations_df['Address Line 2'].isna().all() or (locations_df['Address Line 2'].astype(str).str.strip() == '').all()
        if has_city:
            city_empty = locations_df['City'].isna().all() or (locations_df['City'].astype(str).str.strip() == '').all()
        if has_state:
            state_empty = locations_df['State'].isna().all() or (locations_df['State'].astype(str).str.strip() == '').all()
        if has_zip:
            zip_empty = locations_df['ZIP'].isna().all() or (locations_df['ZIP'].astype(str).str.strip() == '').all()
        
        # Check if Practice ID is empty
        practice_id_empty = True
        if has_practice_id:
            practice_id_empty = locations_df['Practice ID'].isna().all() or (locations_df['Practice ID'].astype(str).str.strip() == '').all()
        
        # Only proceed if Address Line 1 has values and other components are empty
        if not addr1_has_values or not (addr2_empty and city_empty and state_empty and zip_empty):
            print("Address components are already populated. Skipping address formatting.")
            return
        
        # Step 1: Fill Practice ID if manual_practice_ids has only one ID
        if practice_id_empty and manual_practice_ids and manual_practice_ids.strip():
            # Parse manual Practice IDs
            manual_ids = re.split(r'[,;\s]+', manual_practice_ids.strip())
            manual_ids = [pid.strip() for pid in manual_ids if pid.strip()]
            manual_ids = [re.sub(r'\.0$', '', pid) for pid in manual_ids]
            
            # If only one Practice ID, fill the column
            if len(manual_ids) == 1:
                if has_practice_id:
                    locations_df['Practice ID'] = manual_ids[0]
                    print(f"Filled Practice ID column with manually entered ID: {manual_ids[0]}")
                else:
                    locations_df['Practice ID'] = manual_ids[0]
                    print(f"Added Practice ID column with manually entered ID: {manual_ids[0]}")
        
        # Step 2: Format Address Line 1 values
        print("Formatting Address Line 1 values...")
        locations_df['Address Line 1'] = locations_df['Address Line 1'].apply(format_address_line1)
        
        # Step 3: Perform fuzzy matching with Practice_Locations.xlsx
        if practice_locations_file.exists():
            try:
                api_locations_df = pd.read_excel(practice_locations_file, sheet_name='Locations')
                
                if 'address_1' in api_locations_df.columns:
                    print("Performing fuzzy matching with Practice_Locations.xlsx...")
                    
                    # Create a mapping of formatted addresses to API locations
                    matched_count = 0
                    
                    for idx, row in locations_df.iterrows():
                        addr1_formatted = str(row['Address Line 1']).strip() if pd.notna(row['Address Line 1']) else ''
                        
                        if not addr1_formatted:
                            continue
                        
                        # Try to find matches in API locations
                        best_match = None
                        best_score = 0
                        
                        for api_idx, api_row in api_locations_df.iterrows():
                            api_address_combined = combine_api_address(api_row)
                            
                            if fuzzy_match_score(addr1_formatted, api_address_combined, threshold=0.7):
                                # Calculate similarity score
                                addr1_norm = normalize_string(addr1_formatted)
                                api_addr_norm = normalize_string(api_address_combined)
                                
                                # Remove numbers for scoring
                                addr1_no_num = re.sub(r'\d+', '', addr1_norm).strip()
                                api_addr_no_num = re.sub(r'\d+', '', api_addr_norm).strip()
                                
                                score = SequenceMatcher(None, addr1_no_num, api_addr_no_num).ratio()
                                
                                if score > best_score:
                                    best_score = score
                                    best_match = api_row
                        
                        # If we found a good match, populate address components
                        if best_match and best_score >= 0.7:
                            # Do NOT populate Address Line 2 - keep it empty
                            
                            # Populate City if empty
                            if city_empty and has_city:
                                api_city = best_match.get('city', '')
                                if pd.notna(api_city) and str(api_city).strip():
                                    locations_df.at[idx, 'City'] = str(api_city).strip()
                            
                            # Populate State if empty
                            if state_empty and has_state:
                                api_state = best_match.get('state', '')
                                if pd.notna(api_state) and str(api_state).strip():
                                    locations_df.at[idx, 'State'] = str(api_state).strip()
                            
                            # Populate ZIP if empty
                            if zip_empty and has_zip:
                                api_zip = best_match.get('zip', '')
                                if pd.notna(api_zip) and str(api_zip).strip():
                                    locations_df.at[idx, 'ZIP'] = str(api_zip).strip()
                            
                            matched_count += 1
                    
                    print(f"Fuzzy matching completed. Matched {matched_count} addresses.")
                
            except Exception as e:
                print(f"Warning: Could not perform fuzzy matching with Practice_Locations.xlsx: {str(e)}")
        
        # Save updated Locations_input.xlsx
        with pd.ExcelWriter(locations_input_file, engine='openpyxl', mode='w') as writer:
            locations_df.to_excel(writer, index=False, sheet_name='Locations')
        
        # Reapply styling
        wb = load_workbook(locations_input_file)
        ws = wb['Locations']
        
        # Style header row
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(locations_input_file)
        
        print(f"Address formatting completed and saved to {locations_input_file}")
        
    except Exception as e:
        print(f"Error formatting address lines: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == '__main__':
    # For testing purposes
    try:
        format_locations_input_addresses()
        print("Successfully formatted address lines")
    except Exception as e:
        print(f"Error: {str(e)}")
        sys.exit(1)

