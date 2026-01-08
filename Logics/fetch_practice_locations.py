"""
Fetch Practice Locations - Fetches location details from Zocdoc API for Practice IDs

This script reads Practice IDs from Practice_Locations.xlsx and fetches their location
details from the Zocdoc API, then writes the results back to the Excel file.
"""

import requests
import pandas as pd
import openpyxl
import json
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
import sys

# API Endpoints
PRACTICE_IDS_API = 'https://provider-reference-v1.east.zocdoccloud.com/provider-reference/v1/practice/ids-by-monolith-ids~batchGet'
LOCATION_API = 'https://provider-reference-v1.east.zocdoccloud.com/provider-reference/v1/practice/location~batchGet'

# Location fields to fetch
LOCATION_FIELDS = [
    'is_virtual', 'Location Type', 'address_1', 'address_2', 'address_3', 'city', 'state', 'zip',
    'monolith_location_id', 'location_id', 'virtual_visit_type',
    'software', 'software_id', 'hide_on_profile', 'phone', 'phone_extension', 'email_addresses',
    'fax', 'is_active', 'latitude', 'longitude', 'market_id', 'name', 'practice_facing_name',
    'practice_id', 'provider_location_mapping', 'sort_order', 'source', 'source_timestamp',
    'time_zone', 'timestamp', 'type', 'event_id', 'deleted_date', 'location_business_hours'
]


def convert_to_cloud_ids(practice_ids):
    """
    Convert monolith practice IDs to cloud practice IDs
    
    Args:
        practice_ids: List of monolith practice IDs (as strings or integers)
        
    Returns:
        Dictionary mapping monolith_id -> cloud_id
    """
    # Convert all IDs to strings
    practice_ids = [str(pid) for pid in practice_ids]
    
    url = PRACTICE_IDS_API
    headers = {
        'accept': 'application/json',
        'Content-Type': 'application/json'
    }
    data = {
        "monolith_practice_ids": practice_ids
    }
    
    print("Fetching Practice Cloud IDs...")
    try:
        response = requests.post(url, headers=headers, json=data, timeout=30)
        
        cloud_id_map = {}
        if response.status_code == 200:
            result = response.json()
            for item in result.get('practice_ids', []):
                monolith_id = str(item.get('monolith_practice_id'))
                cloud_id = item.get('practice_id')
                if cloud_id:
                    cloud_id_map[monolith_id] = cloud_id
            print(f"Successfully converted {len(cloud_id_map)} Practice IDs to Cloud IDs")
            return cloud_id_map
        else:
            print(f"Failed to fetch cloud IDs. Status code: {response.status_code}")
            print(response.text)
            return {}
    except Exception as e:
        print(f"Error converting Practice IDs to Cloud IDs: {str(e)}")
        return {}


def fetch_location_details(cloud_id):
    """
    Fetch location details for a cloud practice ID
    
    Args:
        cloud_id: Cloud practice ID
        
    Returns:
        List of location dictionaries
    """
    url = LOCATION_API
    headers = {
        'accept': 'application/json',
        'Content-Type': 'application/json'
    }
    data = {"practice_ids": [cloud_id]}
    
    try:
        response = requests.post(url, headers=headers, json=data, timeout=30)
        
        if response.status_code == 200:
            result = response.json()
            return result.get('practice_locations', [])
        else:
            print(f"Failed to fetch locations for Cloud ID {cloud_id}. Status code: {response.status_code}")
            return []
    except Exception as e:
        print(f"Error fetching locations for Cloud ID {cloud_id}: {str(e)}")
        return []


def process_location_data(location):
    """
    Process a single location dictionary and return row data
    
    Args:
        location: Location dictionary from API
        
    Returns:
        List of values for the row
    """
    row_data = []
    
    # is_virtual
    is_virtual_value = location.get('is_virtual', None)
    row_data.append(is_virtual_value)
    
    # Location Type (derived from is_virtual)
    if is_virtual_value is True or (isinstance(is_virtual_value, str) and str(is_virtual_value).upper() == 'TRUE'):
        row_data.append('Virtual')
    elif is_virtual_value is False or (isinstance(is_virtual_value, str) and str(is_virtual_value).upper() == 'FALSE'):
        row_data.append('In Person')
    else:
        row_data.append(None)
    
    # The rest of the fields
    for field in LOCATION_FIELDS[2:]:  # Skip is_virtual and Location Type (already processed)
        value = location.get(field, None)
        if isinstance(value, list):
            value = ', '.join(map(str, value))
        elif isinstance(value, dict):
            # Convert dict to JSON string for fields like location_business_hours
            value = json.dumps(value) if value else None
        row_data.append(value)
    
    return row_data


def fetch_and_write_locations(practice_locations_file):
    """
    Read Practice IDs from Practice_Locations.xlsx, fetch location details, and write to Excel
    
    This function handles both Practice IDs (monolith) and Practice Cloud IDs.
    It will try to convert Practice IDs to Cloud IDs, and if conversion fails,
    it will assume the values are already Cloud IDs.
    
    Args:
        practice_locations_file: Path to Practice_Locations.xlsx file
        
    Returns:
        Path to the updated file
    """
    practice_locations_path = Path(practice_locations_file)
    
    if not practice_locations_path.exists():
        print(f"Error: Practice_Locations.xlsx not found at {practice_locations_path}")
        return None
    
    # Read Practice IDs from the Practice sheet
    try:
        practice_df = pd.read_excel(practice_locations_path, sheet_name='Practice')
    except Exception as e:
        print(f"Error reading Practice sheet: {str(e)}")
        return None
    
    # Check for Practice ID or Practice Cloud ID columns
    has_practice_id = 'Practice ID' in practice_df.columns
    has_practice_cloud_id = 'Practice Cloud ID' in practice_df.columns
    
    if not has_practice_id and not has_practice_cloud_id:
        print("Error: Neither 'Practice ID' nor 'Practice Cloud ID' column found in Practice sheet")
        return None
    
    # Collect all IDs to process
    all_ids = []
    id_to_monolith_map = {}  # Map cloud_id -> monolith_id for tracking
    
    # Get Practice IDs (monolith) if column exists
    if has_practice_id:
        practice_ids = practice_df['Practice ID'].dropna().astype(str).str.strip()
        practice_ids = practice_ids[practice_ids != ''].unique().tolist()
        all_ids.extend(practice_ids)
        print(f"Found {len(practice_ids)} Practice ID(s) (monolith) from 'Practice ID' column")
    
    # Get Practice Cloud IDs if column exists
    if has_practice_cloud_id:
        cloud_ids = practice_df['Practice Cloud ID'].dropna().astype(str).str.strip()
        cloud_ids = cloud_ids[cloud_ids != ''].unique().tolist()
        # Track which cloud IDs came from the Cloud ID column (no monolith ID available)
        for cloud_id in cloud_ids:
            id_to_monolith_map[cloud_id] = None  # None means no monolith ID
        all_ids.extend(cloud_ids)
        print(f"Found {len(cloud_ids)} Practice Cloud ID(s) from 'Practice Cloud ID' column")
    
    if not all_ids:
        print("No Practice IDs or Practice Cloud IDs found in Practice_Locations.xlsx")
        return None
    
    print(f"Total unique IDs to process: {len(all_ids)}")
    
    cloud_id_map = {}
    failed_conversion_ids = []  # IDs that failed conversion (likely already Cloud IDs)
    
    # Step 1: Try to convert monolith IDs to cloud IDs (only if Practice ID column exists)
    # If only Practice Cloud ID is provided, skip conversion and use Cloud IDs directly
    if has_practice_id:
        # All IDs from 'Practice ID' column will be tried as monolith IDs
        practice_ids_to_convert = practice_df['Practice ID'].dropna().astype(str).str.strip()
        practice_ids_to_convert = practice_ids_to_convert[practice_ids_to_convert != ''].unique().tolist()
        
        if practice_ids_to_convert:
            cloud_id_map = convert_to_cloud_ids(practice_ids_to_convert)
            # Update id_to_monolith_map with converted IDs
            for monolith_id, cloud_id in cloud_id_map.items():
                id_to_monolith_map[cloud_id] = monolith_id
            
            # Find IDs that failed conversion (not in cloud_id_map)
            # These are likely already Cloud IDs from manual entry
            for pid in practice_ids_to_convert:
                if pid not in cloud_id_map:
                    # This ID failed conversion, assume it's already a Cloud ID
                    failed_conversion_ids.append(pid)
                    id_to_monolith_map[pid] = None  # No monolith ID available
                    print(f"ID '{pid}' not found in conversion result, treating as Practice Cloud ID")
    else:
        # Only Practice Cloud ID column exists, skip conversion step
        print("Only Practice Cloud ID column found. Skipping conversion step and using Cloud IDs directly.")
    
    # Step 2: Collect all Cloud IDs
    # If only Practice Cloud ID column exists, use those directly
    # Otherwise, combine converted IDs + direct Cloud IDs + failed conversions
    if has_practice_cloud_id and not has_practice_id:
        # Only Cloud IDs provided, use them directly
        all_cloud_ids = list(id_to_monolith_map.keys())
    else:
        # Mix of Practice IDs and/or Cloud IDs
        all_cloud_ids = list(set(list(cloud_id_map.values()) + [cid for cid in id_to_monolith_map.keys() if cid not in cloud_id_map.values()]))
    
    if not all_cloud_ids:
        print("No cloud IDs available. Cannot fetch location details.")
        return None
    
    print(f"Processing {len(all_cloud_ids)} unique Practice Cloud ID(s)")
    if failed_conversion_ids:
        print(f"Note: {len(failed_conversion_ids)} ID(s) were treated as Practice Cloud IDs (conversion failed or already Cloud IDs)")
    
    # Step 3: Create or load Excel file for locations
    wb = openpyxl.load_workbook(practice_locations_path)
    
    # Remove existing Locations sheet if it exists
    if 'Locations' in wb.sheetnames:
        wb.remove(wb['Locations'])
    
    # Create new Locations sheet
    ws = wb.create_sheet('Locations')
    
    # Create header
    header = ['Practice Monolith ID', 'Practice Cloud ID'] + LOCATION_FIELDS
    for idx, col in enumerate(header, 1):
        cell = ws.cell(row=1, column=idx, value=col)
        # Apply header styling
        cell.fill = PatternFill(start_color='000AD6', end_color='000AD6', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Step 4: Fetch location details for each practice Cloud ID
    processed_count = 0
    total_locations = 0
    
    for cloud_id in all_cloud_ids:
        monolith_id = id_to_monolith_map.get(cloud_id)  # Get monolith ID if available
        locations = fetch_location_details(cloud_id)
        
        if locations:
            for loc in locations:
                row_data = [monolith_id, cloud_id]  # monolith_id will be None if it came from Cloud ID column
                row_data.extend(process_location_data(loc))
                ws.append(row_data)
                total_locations += 1
        else:
            # Still add a row with practice info even if no locations found
            row_data = [monolith_id, cloud_id] + [None] * len(LOCATION_FIELDS)
            ws.append(row_data)
        
        processed_count += 1
        sys.stdout.write(f"\rProcessed Practice Cloud IDs: {processed_count}/{len(all_cloud_ids)} | Locations found: {total_locations}")
        sys.stdout.flush()
    
    # Add auto-filter
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{ws.max_row}"
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the file
    wb.save(practice_locations_path)
    
    print(f"\nSuccessfully fetched and wrote {total_locations} location(s) for {processed_count} Practice Cloud ID(s)")
    print(f"Location details saved to 'Locations' sheet in {practice_locations_path.name}")
    
    return str(practice_locations_path)


if __name__ == '__main__':
    # For testing purposes
    import sys
    from pathlib import Path
    
    # Get the project root directory
    LOGICS_DIR = Path(__file__).parent
    PROJECT_ROOT = LOGICS_DIR.parent
    BACKEND_DIR = PROJECT_ROOT / 'backend'
    EXCEL_FILES_DIR = BACKEND_DIR / 'Excel Files'
    
    practice_locations_file = EXCEL_FILES_DIR / 'Practice_Locations.xlsx'
    
    if not practice_locations_file.exists():
        print(f"Error: {practice_locations_file} does not exist")
        sys.exit(1)
    
    try:
        result = fetch_and_write_locations(practice_locations_file)
        if result:
            print(f"Successfully updated {result}")
        else:
            print("Failed to fetch location details")
            sys.exit(1)
    except Exception as e:
        print(f"Error: {str(e)}")
        sys.exit(1)

