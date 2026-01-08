"""
NPI - Extracts NPI Number column from _Mapped.xlsx and writes to Template copy.xlsx

This script reads the 'NPI Number' column from backend\Excel Files\_Mapped.xlsx
and writes it to the 'Provider' sheet in Template copy.xlsx under the 'NPI Number' column.
It also changes the header color to green.
"""

import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Get the project root directory (parent of Transposition Logics folder)
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
BACKEND_DIR = PROJECT_ROOT / 'backend'
EXCEL_FILES_DIR = BACKEND_DIR / 'Excel Files'

def extract_npi_to_template():
    """
    Extract NPI Number column from _Mapped.xlsx and write to Template copy.xlsx
    
    Returns:
        bool: True if successful, False otherwise
    """
    # File paths
    source_file = EXCEL_FILES_DIR / '_Mapped.xlsx'
    template_file = EXCEL_FILES_DIR / 'Template copy.xlsx'
    
    # Check if source file exists
    if not source_file.exists():
        return False
    
    # Check if template file exists
    if not template_file.exists():
        return False
    
    try:
        # Read the NPI Number column from _Mapped.xlsx
        mapped_df = pd.read_excel(source_file)
        
        # Check if 'NPI Number' column exists
        if 'NPI Number' not in mapped_df.columns:
            return False
        
        # Extract the NPI Number column
        npi_data = mapped_df['NPI Number'].tolist()
        
        # Load the template workbook
        wb = load_workbook(template_file)
        
        # Check if 'Provider' sheet exists
        if 'Provider' not in wb.sheetnames:
            return False
        
        # Get the Provider sheet
        provider_sheet = wb['Provider']
        
        # Find the 'NPI Number' column header
        header_row = 1
        npi_column_index = None
        npi_column_letter = None
        
        # Search for the header in the first row
        for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
            if cell.value and str(cell.value).strip().lower() == 'npi number':
                npi_column_index = col_idx
                npi_column_letter = get_column_letter(col_idx)
                break
        
        if npi_column_index is None:
            return False
        
        # Write the NPI data to the column (starting from row 2, as row 1 is the header)
        for row_idx, npi_value in enumerate(npi_data, start=2):
            cell = provider_sheet[f"{npi_column_letter}{row_idx}"]
            # Convert to string and handle NaN values
            if pd.isna(npi_value):
                cell.value = None
            else:
                # Convert float values like 1417469156.0 to integer string "1417469156"
                if isinstance(npi_value, float):
                    # Check if it's a whole number (no decimal part)
                    if npi_value.is_integer():
                        cell.value = str(int(npi_value))
                    else:
                        cell.value = str(npi_value).strip()
                elif isinstance(npi_value, (int, str)):
                    # For integers or strings, convert to string and remove .0 if present
                    value_str = str(npi_value).strip()
                    # Remove .0 at the end if it exists
                    if value_str.endswith('.0'):
                        cell.value = value_str[:-2]
                    else:
                        cell.value = value_str
                else:
                    cell.value = str(npi_value).strip() if npi_value else None
        
        # Color the header cell green
        header_cell = provider_sheet[f"{npi_column_letter}{header_row}"]
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        header_cell.fill = green_fill
        
        # Save the workbook
        wb.save(template_file)
        return True
        
    except Exception as e:
        return False

if __name__ == "__main__":
    extract_npi_to_template()

