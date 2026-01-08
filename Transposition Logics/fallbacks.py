"""
Fallbacks - Fallback logics for populating empty fields in Template copy.xlsx

This module contains fallback logics that populate fields in Template copy.xlsx
when the primary data source is empty or unavailable.
"""

import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Get the project root directory (parent of Transposition Logics folder)
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
BACKEND_DIR = PROJECT_ROOT / 'backend'
EXCEL_FILES_DIR = BACKEND_DIR / 'Excel Files'


def apply_specialty_fallback_from_npi_extracts():
    """
    Fallback logic for Specialty 1 column:
    - If 'Specialty 1' is empty for any row, get the NPI Number from that row
    - Search for the NPI in NPI-Extracts.xlsx (NPI Extracts sheet, 'NPI Number' column)
    - Get the 'Specialty Derived' value from the matched row
    - Split by ';' (semicolon) and populate 'Specialty 1' through 'Specialty 5'
    - If 6 or more values are found, highlight the Specialty columns for that row
    
    Returns:
        bool: True if successful, False otherwise
    """
    # File paths
    npi_extracts_file = EXCEL_FILES_DIR / 'NPI-Extracts.xlsx'
    template_file = EXCEL_FILES_DIR / 'Template copy.xlsx'
    
    # Check if NPI-Extracts file exists
    if not npi_extracts_file.exists():
        return False
    
    # Check if template file exists
    if not template_file.exists():
        return False
    
    try:
        # Load the template workbook
        wb = load_workbook(template_file)
        
        # Check if 'Provider' sheet exists
        if 'Provider' not in wb.sheetnames:
            return False
        
        # Get the Provider sheet
        provider_sheet = wb['Provider']
        
        # Find column headers
        header_row = 1
        npi_column_letter = None
        specialty_columns = {}
        
        # Search for the headers in the first row
        for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
            if cell.value:
                cell_value = str(cell.value).strip().lower()
                if cell_value == 'npi number':
                    npi_column_letter = get_column_letter(col_idx)
                elif cell_value == 'specialty 1':
                    specialty_columns[1] = get_column_letter(col_idx)
                elif cell_value == 'specialty 2':
                    specialty_columns[2] = get_column_letter(col_idx)
                elif cell_value == 'specialty 3':
                    specialty_columns[3] = get_column_letter(col_idx)
                elif cell_value == 'specialty 4':
                    specialty_columns[4] = get_column_letter(col_idx)
                elif cell_value == 'specialty 5':
                    specialty_columns[5] = get_column_letter(col_idx)
        
        if npi_column_letter is None or 1 not in specialty_columns:
            return False
        
        # Read NPI-Extracts.xlsx
        try:
            npi_extracts_df = pd.read_excel(npi_extracts_file, sheet_name='NPI Extracts')
        except Exception as e:
            # Try alternative sheet name without space
            try:
                npi_extracts_df = pd.read_excel(npi_extracts_file, sheet_name='NPI Extracts')
            except:
                return False
        
        # Check if required columns exist
        if 'NPI Number' not in npi_extracts_df.columns or 'Specialty Derived' not in npi_extracts_df.columns:
            return False
        
        # Create a lookup dictionary: NPI Number -> Specialty Derived
        npi_lookup = {}
        for idx, row in npi_extracts_df.iterrows():
            npi_value = row['NPI Number']
            specialty_derived = row['Specialty Derived']
            
            # Normalize NPI value (handle float/int/string formats)
            if pd.notna(npi_value):
                if isinstance(npi_value, float):
                    if npi_value.is_integer():
                        npi_key = str(int(npi_value))
                    else:
                        npi_key = str(npi_value)
                else:
                    npi_key = str(npi_value).strip()
                    # Remove .0 suffix if present
                    if npi_key.endswith('.0'):
                        npi_key = npi_key[:-2]
                
                # Only store non-empty specialty values (ignore blanks from database)
                if npi_key and pd.notna(specialty_derived) and str(specialty_derived).strip():
                    npi_lookup[npi_key] = specialty_derived
        
        # Highlight colors
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow for 6+ specialties
        fallback_fill = PatternFill(start_color="C1EAFF", end_color="C1EAFF", fill_type="solid")  # Blue for fallback cells
        
        # Process each row in the Provider sheet
        max_row = provider_sheet.max_row
        rows_updated = 0
        
        for row_idx in range(2, max_row + 1):  # Start from row 2 (skip header)
            # Check if Specialty 1 is empty
            specialty_1_cell = provider_sheet[f"{specialty_columns[1]}{row_idx}"]
            specialty_1_value = specialty_1_cell.value
            
            # Check if Specialty 1 is empty or None
            is_empty = False
            if specialty_1_value is None:
                is_empty = True
            elif isinstance(specialty_1_value, str) and not specialty_1_value.strip():
                is_empty = True
            
            if is_empty:
                # Get NPI Number from this row
                npi_cell = provider_sheet[f"{npi_column_letter}{row_idx}"]
                npi_value = npi_cell.value
                
                if npi_value is not None:
                    # Normalize NPI value for lookup
                    if isinstance(npi_value, float):
                        if npi_value.is_integer():
                            npi_key = str(int(npi_value))
                        else:
                            npi_key = str(npi_value)
                    else:
                        npi_key = str(npi_value).strip()
                        # Remove .0 suffix if present
                        if npi_key.endswith('.0'):
                            npi_key = npi_key[:-2]
                    
                    # Look up Specialty Derived value
                    if npi_key in npi_lookup:
                        specialty_derived_value = npi_lookup[npi_key]
                        
                        if pd.notna(specialty_derived_value) and str(specialty_derived_value).strip():
                            # Split by semicolon
                            specialty_parts = [part.strip() for part in str(specialty_derived_value).split(';')]
                            # Remove empty strings
                            specialty_parts = [part for part in specialty_parts if part]
                            
                            # Populate Specialty 1 through Specialty 5
                            num_specialties = len(specialty_parts)
                            should_highlight = num_specialties >= 6
                            
                            for i in range(1, 6):
                                if i in specialty_columns:
                                    cell = provider_sheet[f"{specialty_columns[i]}{row_idx}"]
                                    if i <= num_specialties:
                                        cell.value = specialty_parts[i - 1]
                                        # Highlight fallback cells in blue
                                        cell.fill = fallback_fill
                                    else:
                                        cell.value = None
                                    
                                    # Highlight if 6+ specialties found (yellow overrides blue)
                                    if should_highlight:
                                        cell.fill = highlight_fill
                            
                            rows_updated += 1
        
        # If any rows were updated, highlight the Specialty column headers green
        if rows_updated > 0:
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            for i in range(1, 6):
                if i in specialty_columns:
                    header_cell = provider_sheet[f"{specialty_columns[i]}{header_row}"]
                    header_cell.fill = green_fill
        
        # Save the workbook
        wb.save(template_file)
        return True
        
    except Exception as e:
        return False


def apply_professional_suffix_fallback_from_npi_extracts():
    """
    Fallback logic for Professional Suffix 1 column:
    - If 'Professional Suffix 1' is empty for any row, get the NPI Number from that row
    - Search for the NPI in NPI-Extracts.xlsx (NPI Extracts sheet, 'NPI Number' column)
    - Get the 'Suffix Derived' value from the matched row
    - Split by ';' (semicolon) and populate 'Professional Suffix 1' through 'Professional Suffix 3'
    - If 4 or more values are found, highlight the Professional Suffix columns for that row
    
    Returns:
        bool: True if successful, False otherwise
    """
    # File paths
    npi_extracts_file = EXCEL_FILES_DIR / 'NPI-Extracts.xlsx'
    template_file = EXCEL_FILES_DIR / 'Template copy.xlsx'
    
    # Check if NPI-Extracts file exists
    if not npi_extracts_file.exists():
        return False
    
    # Check if template file exists
    if not template_file.exists():
        return False
    
    try:
        # Load the template workbook
        wb = load_workbook(template_file)
        
        # Check if 'Provider' sheet exists
        if 'Provider' not in wb.sheetnames:
            return False
        
        # Get the Provider sheet
        provider_sheet = wb['Provider']
        
        # Find column headers
        header_row = 1
        npi_column_letter = None
        suffix_columns = {}
        
        # Search for the headers in the first row
        for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
            if cell.value:
                cell_value = str(cell.value).strip().lower()
                if cell_value == 'npi number':
                    npi_column_letter = get_column_letter(col_idx)
                elif cell_value == 'professional suffix 1':
                    suffix_columns[1] = get_column_letter(col_idx)
                elif cell_value == 'professional suffix 2':
                    suffix_columns[2] = get_column_letter(col_idx)
                elif cell_value == 'professional suffix 3':
                    suffix_columns[3] = get_column_letter(col_idx)
        
        if npi_column_letter is None or 1 not in suffix_columns:
            return False
        
        # Read NPI-Extracts.xlsx
        try:
            npi_extracts_df = pd.read_excel(npi_extracts_file, sheet_name='NPI Extracts')
        except Exception as e:
            # Try alternative sheet name without space
            try:
                npi_extracts_df = pd.read_excel(npi_extracts_file, sheet_name='NPI Extracts')
            except:
                return False
        
        # Check if required columns exist
        if 'NPI Number' not in npi_extracts_df.columns or 'Suffix Derived' not in npi_extracts_df.columns:
            return False
        
        # Create a lookup dictionary: NPI Number -> Suffix Derived
        npi_lookup = {}
        for idx, row in npi_extracts_df.iterrows():
            npi_value = row['NPI Number']
            suffix_derived = row['Suffix Derived']
            
            # Normalize NPI value (handle float/int/string formats)
            if pd.notna(npi_value):
                if isinstance(npi_value, float):
                    if npi_value.is_integer():
                        npi_key = str(int(npi_value))
                    else:
                        npi_key = str(npi_value)
                else:
                    npi_key = str(npi_value).strip()
                    # Remove .0 suffix if present
                    if npi_key.endswith('.0'):
                        npi_key = npi_key[:-2]
                
                # Only store non-empty suffix values (ignore blanks from database)
                if npi_key and pd.notna(suffix_derived) and str(suffix_derived).strip():
                    npi_lookup[npi_key] = suffix_derived
        
        # Highlight colors
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow for 4+ suffixes
        fallback_fill = PatternFill(start_color="C1EAFF", end_color="C1EAFF", fill_type="solid")  # Blue for fallback cells
        
        # Process each row in the Provider sheet
        max_row = provider_sheet.max_row
        rows_updated = 0
        
        for row_idx in range(2, max_row + 1):  # Start from row 2 (skip header)
            # Check if Professional Suffix 1 is empty
            suffix_1_cell = provider_sheet[f"{suffix_columns[1]}{row_idx}"]
            suffix_1_value = suffix_1_cell.value
            
            # Check if Professional Suffix 1 is empty or None
            is_empty = False
            if suffix_1_value is None:
                is_empty = True
            elif isinstance(suffix_1_value, str) and not suffix_1_value.strip():
                is_empty = True
            
            if is_empty:
                # Get NPI Number from this row
                npi_cell = provider_sheet[f"{npi_column_letter}{row_idx}"]
                npi_value = npi_cell.value
                
                if npi_value is not None:
                    # Normalize NPI value for lookup
                    if isinstance(npi_value, float):
                        if npi_value.is_integer():
                            npi_key = str(int(npi_value))
                        else:
                            npi_key = str(npi_value)
                    else:
                        npi_key = str(npi_value).strip()
                        # Remove .0 suffix if present
                        if npi_key.endswith('.0'):
                            npi_key = npi_key[:-2]
                    
                    # Look up Suffix Derived value
                    if npi_key in npi_lookup:
                        suffix_derived_value = npi_lookup[npi_key]
                        
                        if pd.notna(suffix_derived_value) and str(suffix_derived_value).strip():
                            # Split by semicolon
                            suffix_parts = [part.strip() for part in str(suffix_derived_value).split(';')]
                            # Remove empty strings
                            suffix_parts = [part for part in suffix_parts if part]
                            
                            # Populate Professional Suffix 1 through Professional Suffix 3
                            num_suffixes = len(suffix_parts)
                            should_highlight = num_suffixes >= 4
                            
                            for i in range(1, 4):
                                if i in suffix_columns:
                                    cell = provider_sheet[f"{suffix_columns[i]}{row_idx}"]
                                    if i <= num_suffixes:
                                        cell.value = suffix_parts[i - 1]
                                        # Highlight fallback cells in blue
                                        cell.fill = fallback_fill
                                    else:
                                        cell.value = None
                                    
                                    # Highlight if 4+ suffixes found (yellow overrides blue)
                                    if should_highlight:
                                        cell.fill = highlight_fill
                            
                            rows_updated += 1
        
        # If any rows were updated, highlight the Professional Suffix column headers green
        if rows_updated > 0:
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            for i in range(1, 4):
                if i in suffix_columns:
                    header_cell = provider_sheet[f"{suffix_columns[i]}{header_row}"]
                    header_cell.fill = green_fill
        
        # Save the workbook
        wb.save(template_file)
        return True
        
    except Exception as e:
        return False


def apply_gender_fallback_from_npi_extracts():
    """
    Fallback logic for Gender column:
    - If 'Gender' is empty for any row, get the NPI Number from that row
    - Search for the NPI in NPI-Extracts.xlsx (NPI Extracts sheet, 'NPI Number' column)
    - Get the 'GENDER' value from the matched row (format: 'F' or 'M')
    - Convert 'F' to 'Female' and 'M' to 'Male'
    - Put it in the Gender column of the Provider tab
    
    Returns:
        bool: True if successful, False otherwise
    """
    # File paths
    npi_extracts_file = EXCEL_FILES_DIR / 'NPI-Extracts.xlsx'
    template_file = EXCEL_FILES_DIR / 'Template copy.xlsx'
    
    # Check if NPI-Extracts file exists
    if not npi_extracts_file.exists():
        return False
    
    # Check if template file exists
    if not template_file.exists():
        return False
    
    try:
        # Load the template workbook
        wb = load_workbook(template_file)
        
        # Check if 'Provider' sheet exists
        if 'Provider' not in wb.sheetnames:
            return False
        
        # Get the Provider sheet
        provider_sheet = wb['Provider']
        
        # Find column headers
        header_row = 1
        npi_column_letter = None
        gender_column_letter = None
        
        # Search for the headers in the first row
        for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
            if cell.value:
                cell_value = str(cell.value).strip().lower()
                if cell_value == 'npi number':
                    npi_column_letter = get_column_letter(col_idx)
                elif cell_value == 'gender':
                    gender_column_letter = get_column_letter(col_idx)
        
        if npi_column_letter is None or gender_column_letter is None:
            return False
        
        # Read NPI-Extracts.xlsx
        try:
            npi_extracts_df = pd.read_excel(npi_extracts_file, sheet_name='NPI Extracts')
        except Exception as e:
            # Try alternative sheet name without space
            try:
                npi_extracts_df = pd.read_excel(npi_extracts_file, sheet_name='NPI Extracts')
            except:
                return False
        
        # Check if required columns exist
        if 'NPI Number' not in npi_extracts_df.columns or 'GENDER' not in npi_extracts_df.columns:
            return False
        
        # Create a lookup dictionary: NPI Number -> GENDER
        npi_lookup = {}
        for idx, row in npi_extracts_df.iterrows():
            npi_value = row['NPI Number']
            gender_value = row['GENDER']
            
            # Normalize NPI value (handle float/int/string formats)
            if pd.notna(npi_value):
                if isinstance(npi_value, float):
                    if npi_value.is_integer():
                        npi_key = str(int(npi_value))
                    else:
                        npi_key = str(npi_value)
                else:
                    npi_key = str(npi_value).strip()
                    # Remove .0 suffix if present
                    if npi_key.endswith('.0'):
                        npi_key = npi_key[:-2]
                
                # Only store non-empty gender values (ignore blanks from database)
                if npi_key and pd.notna(gender_value) and str(gender_value).strip():
                    npi_lookup[npi_key] = gender_value
        
        # Process each row in the Provider sheet
        max_row = provider_sheet.max_row
        rows_updated = 0
        
        for row_idx in range(2, max_row + 1):  # Start from row 2 (skip header)
            # Check if Gender is empty
            gender_cell = provider_sheet[f"{gender_column_letter}{row_idx}"]
            gender_value = gender_cell.value
            
            # Check if Gender is empty or None
            is_empty = False
            if gender_value is None:
                is_empty = True
            elif isinstance(gender_value, str) and not gender_value.strip():
                is_empty = True
            
            if is_empty:
                # Get NPI Number from this row
                npi_cell = provider_sheet[f"{npi_column_letter}{row_idx}"]
                npi_value = npi_cell.value
                
                if npi_value is not None:
                    # Normalize NPI value for lookup
                    if isinstance(npi_value, float):
                        if npi_value.is_integer():
                            npi_key = str(int(npi_value))
                        else:
                            npi_key = str(npi_value)
                    else:
                        npi_key = str(npi_value).strip()
                        # Remove .0 suffix if present
                        if npi_key.endswith('.0'):
                            npi_key = npi_key[:-2]
                    
                    # Look up GENDER value
                    if npi_key in npi_lookup:
                        gender_code = npi_lookup[npi_key]
                        
                        if pd.notna(gender_code) and str(gender_code).strip():
                            # Convert 'F' to 'Female' and 'M' to 'Male', otherwise use as-is
                            gender_code_upper = str(gender_code).strip().upper()
                            fallback_fill = PatternFill(start_color="C1EAFF", end_color="C1EAFF", fill_type="solid")  # Blue for fallback cells
                            
                            if gender_code_upper == 'F':
                                gender_cell.value = 'Female'
                                gender_cell.fill = fallback_fill
                                rows_updated += 1
                            elif gender_code_upper == 'M':
                                gender_cell.value = 'Male'
                                gender_cell.fill = fallback_fill
                                rows_updated += 1
                            else:
                                # For any other value, use it as-is
                                gender_cell.value = str(gender_code).strip()
                                gender_cell.fill = fallback_fill
                                rows_updated += 1
        
        # If any rows were updated, highlight the Gender column header green
        if rows_updated > 0:
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            header_cell = provider_sheet[f"{gender_column_letter}{header_row}"]
            header_cell.fill = green_fill
        
        # Save the workbook
        wb.save(template_file)
        return True
        
    except Exception as e:
        return False


def apply_additional_languages_fallback_from_npi_extracts():
    """
    Fallback logic for Additional Languages Spoken 1 column:
    - If 'Additional Languages Spoken 1' is empty for any row, get the NPI Number from that row
    - Search for the NPI in NPI-Extracts.xlsx (NPI Extracts sheet, 'NPI Number' column)
    - Get the 'LANGUAGES' value from the matched row
    - Split by comma (',') separator
    - Filter out 'ENGLISH' or 'English' (case-insensitive)
    - Populate 'Additional Languages Spoken 1' through 'Additional Languages Spoken 3' (max 3)
    
    Returns:
        bool: True if successful, False otherwise
    """
    # File paths
    npi_extracts_file = EXCEL_FILES_DIR / 'NPI-Extracts.xlsx'
    template_file = EXCEL_FILES_DIR / 'Template copy.xlsx'
    
    # Check if NPI-Extracts file exists
    if not npi_extracts_file.exists():
        return False
    
    # Check if template file exists
    if not template_file.exists():
        return False
    
    try:
        # Load the template workbook
        wb = load_workbook(template_file)
        
        # Check if 'Provider' sheet exists
        if 'Provider' not in wb.sheetnames:
            return False
        
        # Get the Provider sheet
        provider_sheet = wb['Provider']
        
        # Find column headers
        header_row = 1
        npi_column_letter = None
        lang_columns = {}
        
        # Search for the headers in the first row
        for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
            if cell.value:
                cell_value = str(cell.value).strip().lower()
                if cell_value == 'npi number':
                    npi_column_letter = get_column_letter(col_idx)
                elif cell_value == 'additional language spoken 1' or cell_value == 'additional languages spoken 1':
                    lang_columns[1] = get_column_letter(col_idx)
                elif cell_value == 'additional language spoken 2' or cell_value == 'additional languages spoken 2':
                    lang_columns[2] = get_column_letter(col_idx)
                elif cell_value == 'additional language spoken 3' or cell_value == 'additional languages spoken 3':
                    lang_columns[3] = get_column_letter(col_idx)
        
        if npi_column_letter is None or 1 not in lang_columns:
            return False
        
        # Read NPI-Extracts.xlsx
        try:
            npi_extracts_df = pd.read_excel(npi_extracts_file, sheet_name='NPI Extracts')
        except Exception as e:
            # Try alternative sheet name without space
            try:
                npi_extracts_df = pd.read_excel(npi_extracts_file, sheet_name='NPI Extracts')
            except:
                return False
        
        # Check if required columns exist
        if 'NPI Number' not in npi_extracts_df.columns or 'LANGUAGES' not in npi_extracts_df.columns:
            return False
        
        # Create a lookup dictionary: NPI Number -> LANGUAGES
        npi_lookup = {}
        for idx, row in npi_extracts_df.iterrows():
            npi_value = row['NPI Number']
            languages_value = row['LANGUAGES']
            
            # Normalize NPI value (handle float/int/string formats)
            if pd.notna(npi_value):
                if isinstance(npi_value, float):
                    if npi_value.is_integer():
                        npi_key = str(int(npi_value))
                    else:
                        npi_key = str(npi_value)
                else:
                    npi_key = str(npi_value).strip()
                    # Remove .0 suffix if present
                    if npi_key.endswith('.0'):
                        npi_key = npi_key[:-2]
                
                # Only store non-empty language values (ignore blanks from database)
                if npi_key and pd.notna(languages_value) and str(languages_value).strip():
                    npi_lookup[npi_key] = languages_value
        
        # Highlight color for fallback cells
        fallback_fill = PatternFill(start_color="C1EAFF", end_color="C1EAFF", fill_type="solid")  # Blue
        
        # Process each row in the Provider sheet
        max_row = provider_sheet.max_row
        rows_updated = 0
        
        for row_idx in range(2, max_row + 1):  # Start from row 2 (skip header)
            # Check if Additional Languages Spoken 1 is empty
            lang_1_cell = provider_sheet[f"{lang_columns[1]}{row_idx}"]
            lang_1_value = lang_1_cell.value
            
            # Check if Additional Languages Spoken 1 is empty or None
            is_empty = False
            if lang_1_value is None:
                is_empty = True
            elif isinstance(lang_1_value, str) and not lang_1_value.strip():
                is_empty = True
            
            if is_empty:
                # Get NPI Number from this row
                npi_cell = provider_sheet[f"{npi_column_letter}{row_idx}"]
                npi_value = npi_cell.value
                
                if npi_value is not None:
                    # Normalize NPI value for lookup
                    if isinstance(npi_value, float):
                        if npi_value.is_integer():
                            npi_key = str(int(npi_value))
                        else:
                            npi_key = str(npi_value)
                    else:
                        npi_key = str(npi_value).strip()
                        # Remove .0 suffix if present
                        if npi_key.endswith('.0'):
                            npi_key = npi_key[:-2]
                    
                    # Look up LANGUAGES value
                    if npi_key in npi_lookup:
                        languages_value = npi_lookup[npi_key]
                        
                        if pd.notna(languages_value) and str(languages_value).strip():
                            # Split by comma
                            lang_parts = [part.strip() for part in str(languages_value).split(',')]
                            # Remove empty strings
                            lang_parts = [part for part in lang_parts if part]
                            
                            # Filter out 'ENGLISH' or 'English' (case-insensitive)
                            lang_parts = [lang for lang in lang_parts if lang.strip().upper() != 'ENGLISH']
                            
                            # Convert to Camel case (first letter uppercase, rest lowercase)
                            lang_parts = [lang.strip().capitalize() if lang.strip() else lang for lang in lang_parts]
                            
                            # Take only first 3 languages (max 3)
                            lang_parts = lang_parts[:3]
                            
                            # Populate Additional Languages Spoken 1 through 3
                            for i in range(1, 4):
                                if i in lang_columns:
                                    cell = provider_sheet[f"{lang_columns[i]}{row_idx}"]
                                    if i <= len(lang_parts):
                                        cell.value = lang_parts[i - 1]
                                        # Highlight fallback cells in blue
                                        cell.fill = fallback_fill
                                    else:
                                        cell.value = None
                            
                            rows_updated += 1
        
        # If any rows were updated, highlight the Additional Languages Spoken column headers green
        if rows_updated > 0:
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            for i in range(1, 4):
                if i in lang_columns:
                    header_cell = provider_sheet[f"{lang_columns[i]}{header_row}"]
                    header_cell.fill = green_fill
        
        # Save the workbook
        wb.save(template_file)
        return True
        
    except Exception as e:
        return False


def apply_all_fallbacks():
    """
    Apply all fallback logics in sequence.
    
    Returns:
        bool: True if all fallbacks were successful, False otherwise
    """
    results = []
    
    # Apply Specialty fallback
    try:
        success = apply_specialty_fallback_from_npi_extracts()
        results.append(("Specialty Fallback", success))
    except Exception as e:
        results.append(("Specialty Fallback", False))
    
    # Apply Professional Suffix fallback
    try:
        success = apply_professional_suffix_fallback_from_npi_extracts()
        results.append(("Professional Suffix Fallback", success))
    except Exception as e:
        results.append(("Professional Suffix Fallback", False))
    
    # Apply Gender fallback
    try:
        success = apply_gender_fallback_from_npi_extracts()
        results.append(("Gender Fallback", success))
    except Exception as e:
        results.append(("Gender Fallback", False))
    
    # Apply Additional Languages fallback
    try:
        success = apply_additional_languages_fallback_from_npi_extracts()
        results.append(("Additional Languages Fallback", success))
    except Exception as e:
        results.append(("Additional Languages Fallback", False))
    
    # Add more fallback functions here as they are created
    
    return all(result[1] for result in results)


if __name__ == "__main__":
    apply_all_fallbacks()

