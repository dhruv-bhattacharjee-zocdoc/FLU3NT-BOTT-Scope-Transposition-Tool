"""
Patients Accepted - Applies data validation dropdowns and other logic for Patients Accepted columns

This script handles parsing of Patients Accepted data in two forms:
1. Age ranges (e.g., "0–5, 6–10, 11–13, 13–17") - determines category based on min/max age
2. Text variants (e.g., "Adults", "Pediatric", "Adult & Pediatric") - categorizes based on keywords

It applies a dropdown with options: Adult, Pediatric, Both
"""

import os
import re
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Get the project root directory (parent of Transposition Logics folder)
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
BACKEND_DIR = PROJECT_ROOT / 'backend'
EXCEL_FILES_DIR = BACKEND_DIR / 'Excel Files'

def parse_age_ranges(value):
    """
    Parse age ranges from a string and determine if it's Adult, Pediatric, or Both.
    
    Examples:
        "0–5, 6–10, 11–13, 13–17" -> Pediatric (min=0, max=17)
        "18–21, 22–26, 27–40, 41–64, 65–74, 75+" -> Adult (min=18, max=200)
        "13–17, 18–21, 22–26" -> Both (min=13, max=26)
        "6–10, 11–13, 27–40" -> Both (min=6, max=40)
    
    Returns:
        str: 'Adult', 'Pediatric', or 'Both', or None if no age ranges found
    """
    if not value or not isinstance(value, str):
        return None
    
    # Check if the value contains age ranges (numbers with dashes or plus signs)
    # Pattern: matches things like "0–5", "6-10", "11-13", "75+", etc.
    # Handles en dash (–), em dash (—), and regular dash (-)
    age_pattern = r'(\d+)\s*[–—\-]\s*(\d+)|(\d+)\+'
    
    matches = re.findall(age_pattern, value)
    if not matches:
        return None
    
    min_age = float('inf')
    max_age = float('-inf')
    
    for match in matches:
        if match[2]:  # Matched "75+" pattern
            age = int(match[2])
            min_age = min(min_age, age)
            max_age = max(max_age, 200)  # 75+ means up to 200
        else:  # Matched "0–5" pattern
            start_age = int(match[0])
            end_age = int(match[1])
            min_age = min(min_age, start_age)
            max_age = max(max_age, end_age)
    
    if min_age == float('inf') or max_age == float('-inf'):
        return None
    
    # Determine category based on age ranges
    # Pediatric: min_age = 0, max_age <= 17
    # Adult: min_age >= 18, max_age >= 18
    # Both: spans across both ranges
    
    if min_age == 0 and max_age <= 17:
        return 'Pediatric'
    elif min_age >= 18:
        return 'Adult'
    else:
        # Spans across both ranges (e.g., min=6, max=40 or min=13, max=26)
        return 'Both'

def parse_text_variants(value):
    """
    Parse text variants and determine if it's Adult, Pediatric, or Both.
    
    Adult variants: Adults, Adult only, Adult patients, Adults (18+), Adult behavioral health, etc.
    Pediatric variants: Pediatrics, Pediatric, Children, Kids, Child, Adolescent, Teen, 
                       Child & Adolescent (when no adult wording is present)
    Both variants: Adult & Pediatric, Child & Adult, Children and Adults, All ages, All patients, etc.
    
    Returns:
        str: 'Adult', 'Pediatric', or 'Both', or None if no match found
    """
    if not value or not isinstance(value, str):
        return None
    
    value_lower = value.lower().strip()
    
    # Check for "Both" variants first (these contain both adult and pediatric keywords)
    both_patterns = [
        r'adult.*pediatric|pediatric.*adult',
        r'child.*adult|adult.*child',
        r'children.*adult|adult.*children',
        r'all\s+ages',
        r'all\s+patients',
        r'both',
    ]
    
    for pattern in both_patterns:
        if re.search(pattern, value_lower):
            return 'Both'
    
    # Check for Adult variants
    adult_patterns = [
        r'\badult\b',
        r'\badults\b',
        r'adult\s+only',
        r'adult\s+patients',
        r'adults\s*\(18\+\)',
        r'adult\s+behavioral',
    ]
    
    has_adult = False
    for pattern in adult_patterns:
        if re.search(pattern, value_lower):
            has_adult = True
            break
    
    # Check for Pediatric variants
    pediatric_patterns = [
        r'\bpediatric\b',
        r'\bpediatrics\b',
        r'\bchildren\b',
        r'\bkids\b',
        r'\bchild\b',
        r'\badolescent\b',
        r'\bteen\b',
        r'child\s*&\s*adolescent',
    ]
    
    has_pediatric = False
    for pattern in pediatric_patterns:
        if re.search(pattern, value_lower):
            has_pediatric = True
            break
    
    # If both adult and pediatric keywords found, it's Both
    if has_adult and has_pediatric:
        return 'Both'
    elif has_adult:
        return 'Adult'
    elif has_pediatric:
        return 'Pediatric'
    
    return None

def categorize_patients_accepted(value):
    """
    Categorize a Patients Accepted value into 'Adult', 'Pediatric', or 'Both'.
    
    First tries to parse as age ranges, then as text variants.
    If neither works, returns None.
    
    Returns:
        str: 'Adult', 'Pediatric', 'Both', or None
    """
    if not value:
        return None
    
    # Convert to string if not already
    value_str = str(value).strip()
    
    if not value_str:
        return None
    
    # Check if it's already one of the standard values
    value_lower = value_str.lower()
    if value_lower == 'adult':
        return 'Adult'
    elif value_lower == 'pediatric':
        return 'Pediatric'
    elif value_lower == 'both':
        return 'Both'
    
    # Try parsing as age ranges first
    result = parse_age_ranges(value_str)
    if result:
        return result
    
    # Try parsing as text variants
    result = parse_text_variants(value_str)
    if result:
        return result
    
    return None

def apply_patients_accepted_to_template():
    """
    Apply Patients Accepted dropdowns and logic to Template copy.xlsx
    
    Reads 'Patients Accepted' column from _Mapped.xlsx if available, parses age ranges or text variants,
    categorizes them, and writes to the Provider sheet in Template copy.xlsx.
    If source data is not available, applies dropdown and sets default value 'Adult' for all rows.
    
    Returns:
        bool: True if successful, False otherwise
    """
    # File paths
    source_file = EXCEL_FILES_DIR / '_Mapped.xlsx'
    template_file = EXCEL_FILES_DIR / 'Template copy.xlsx'
    
    # Check if template file exists
    if not template_file.exists():
        return False
    
    try:
        # Try to read the Patients Accepted column from _Mapped.xlsx (if available)
        patients_accepted_data = None
        if source_file.exists():
            try:
                mapped_df = pd.read_excel(source_file)
                if 'Patients Accepted' in mapped_df.columns:
                    patients_accepted_data = mapped_df['Patients Accepted'].tolist()
            except:
                pass  # If reading fails, continue without source data
        
        # Load the template workbook
        wb = load_workbook(template_file)
        
        # Check if 'Provider' sheet exists
        if 'Provider' not in wb.sheetnames:
            return False
        
        # Get the Provider sheet
        provider_sheet = wb['Provider']
        header_row = 1
        max_row = provider_sheet.max_row
        
        # Find Patients Accepted column (or similar column name)
        patients_accepted_column_letter = None
        for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
            if cell.value:
                cell_value = str(cell.value).strip().lower()
                # Check for various possible column names
                if cell_value == 'patients accepted' or cell_value == 'patient accepted' or \
                   cell_value == 'patients accepted?' or cell_value == 'patient accepted?':
                    patients_accepted_column_letter = get_column_letter(col_idx)
                    break
        
        if patients_accepted_column_letter is None:
            return False
        
        # Apply dropdown: Adult, Pediatric, Both
        patients_accepted_validation = DataValidation(type="list", formula1='"Adult,Pediatric,Both"')
        patients_accepted_validation.add(f"{patients_accepted_column_letter}2:{patients_accepted_column_letter}{max_row}")
        provider_sheet.add_data_validation(patients_accepted_validation)
        
        # Process values: if source data exists, parse and categorize; otherwise set default to 'Adult'
        if patients_accepted_data is not None and len(patients_accepted_data) > 0:
            # Process each value from _Mapped.xlsx: parse age ranges or text variants, categorize, and write
            for row_idx, value in enumerate(patients_accepted_data, start=2):
                if row_idx > max_row:
                    break
                cell = provider_sheet[f"{patients_accepted_column_letter}{row_idx}"]
                
                if pd.isna(value) or (isinstance(value, str) and value.strip() == ''):
                    # Set default value 'Adult' for empty cells
                    cell.value = 'Adult'
                else:
                    # Try to categorize the value (age ranges or text variants)
                    categorized = categorize_patients_accepted(value)
                    if categorized:
                        # Write the categorized value
                        cell.value = categorized
                    else:
                        # If we can't categorize, set default to 'Adult'
                        cell.value = 'Adult'
        else:
            # No source data available - set default 'Adult' for all rows
            for row_idx in range(2, max_row + 1):
                cell = provider_sheet[f"{patients_accepted_column_letter}{row_idx}"]
                if cell.value is None or str(cell.value).strip() == '':
                    cell.value = 'Adult'
        
        # Color the header cell green
        header_cell = provider_sheet[f"{patients_accepted_column_letter}{header_row}"]
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        header_cell.fill = green_fill
        
        # Save the workbook
        wb.save(template_file)
        return True
        
    except Exception as e:
        print(f"Error in apply_patients_accepted_to_template: {e}")
        return False

if __name__ == "__main__":
    apply_patients_accepted_to_template()

