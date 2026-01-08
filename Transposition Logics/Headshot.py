"""
Headshot - Extracts Headshot Link column from _Mapped.xlsx and writes to Template copy.xlsx

This script reads the 'Headshot Link' column from backend\Excel Files\_Mapped.xlsx
and writes it to the 'Provider' sheet in Template copy.xlsx under the 'Headshot Link' column.
It also changes the header color to green.
"""

import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Get the project root directory (parent of Transposition Logics folder)
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
BACKEND_DIR = PROJECT_ROOT / 'backend'
EXCEL_FILES_DIR = BACKEND_DIR / 'Excel Files'

def extract_headshot_to_template():
    """
    Extract Headshot Link column from _Mapped.xlsx and write to Template copy.xlsx
    
    Returns:
        bool: True if successful, False otherwise
    """
    # File paths
    source_file = EXCEL_FILES_DIR / '_Mapped.xlsx'
    template_file = EXCEL_FILES_DIR / 'Template copy.xlsx'
    
    # Check if source file exists
    if not source_file.exists():
        return False
    
    # Check if template file exists
    if not template_file.exists():
        return False
    
    try:
        # Read the Headshot Link column from _Mapped.xlsx
        mapped_df = pd.read_excel(source_file)
        
        # Check if 'Headshot Link' column exists
        if 'Headshot Link' not in mapped_df.columns:
            return False
        
        # Extract the Headshot Link column
        headshot_data = mapped_df['Headshot Link'].tolist()
        
        # Load the template workbook
        wb = load_workbook(template_file)
        
        # Check if 'Provider' sheet exists
        if 'Provider' not in wb.sheetnames:
            return False
        
        # Get the Provider sheet
        provider_sheet = wb['Provider']
        
        # Find the 'Headshot Link' column header
        header_row = 1
        headshot_column_index = None
        headshot_column_letter = None
        
        # Search for the header in the first row
        for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
            if cell.value and str(cell.value).strip().lower() == 'headshot link':
                headshot_column_index = col_idx
                headshot_column_letter = get_column_letter(col_idx)
                break
        
        if headshot_column_index is None:
            return False
        
        # Write the Headshot Link data to the column (starting from row 2, as row 1 is the header)
        for row_idx, headshot_value in enumerate(headshot_data, start=2):
            cell = provider_sheet[f"{headshot_column_letter}{row_idx}"]
            # Convert to string and handle NaN values
            if pd.isna(headshot_value):
                cell.value = None
            else:
                cell.value = str(headshot_value).strip() if headshot_value else None
        
        # Color the header cell green
        header_cell = provider_sheet[f"{headshot_column_letter}{header_row}"]
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        header_cell.fill = green_fill
        
        # Save the workbook
        wb.save(template_file)
        return True
        
    except Exception as e:
        return False

if __name__ == "__main__":
    extract_headshot_to_template()

