"""
NPI Snowflake - Extracts NPI Number column from _Mapped.xlsx and queries Snowflake

This script reads _Mapped.xlsx, extracts the 'NPI Number' column, queries Snowflake
for each NPI to get provider data, and creates NPI-Extracts.xlsx with NPI numbers
and extracted data in new columns.
"""

import os
import sys
from pathlib import Path
import pandas as pd
import snowflake.connector
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Get the project root directory (parent of Logics folder)
LOGICS_DIR = Path(__file__).parent
PROJECT_ROOT = LOGICS_DIR.parent
BACKEND_DIR = PROJECT_ROOT / 'backend'
EXCEL_FILES_DIR = BACKEND_DIR / 'Excel Files'

def format_location(location_obj):
    """Format a location object as 'line_1 city state zip'"""
    if not isinstance(location_obj, dict):
        return None
    
    # Try to get address fields from nested 'value' first
    value = location_obj.get('value', {})
    
    # If 'value' is a dict, use it; otherwise check if fields are at top level
    if isinstance(value, dict) and (value.get('city') or value.get('line_1') or value.get('state') or value.get('zip')):
        # Fields are in nested 'value'
        line_1 = value.get('line_1') or ''
        city = value.get('city') or ''
        state = value.get('state') or ''
        zip_code = value.get('zip') or ''
    else:
        # Fields might be at top level of location_obj
        line_1 = location_obj.get('line_1') or ''
        city = location_obj.get('city') or ''
        state = location_obj.get('state') or ''
        zip_code = location_obj.get('zip') or ''
    
    # Convert to strings and strip
    line_1 = str(line_1).strip() if line_1 else ''
    city = str(city).strip() if city else ''
    state = str(state).strip() if state else ''
    zip_code = str(zip_code).strip() if zip_code else ''
    
    # Build address string with commas
    parts = []
    if line_1 and line_1.lower() != 'null':
        parts.append(line_1)
    if city and city.lower() != 'null':
        parts.append(city)
    if state and state.lower() != 'null':
        parts.append(state)
    if zip_code and zip_code.lower() != 'null':
        parts.append(zip_code)
    
    # Join with commas and spaces: "line_1, city, state, zip"
    return ', '.join(parts) if parts else None

def extract_value(val, colname):
    """Extract 'value' from JSON strings in each cell, with special handling for SPECIALTIES, LANGUAGES, and LOCATIONS"""
    if isinstance(val, str):
        try:
            parsed = json.loads(val)
            
            # Handle LOCATIONS column - format addresses
            if colname == 'LOCATIONS':
                locations = []
                # Handle list of location objects
                if isinstance(parsed, list) and len(parsed) > 0:
                    for location_obj in parsed:
                        formatted = format_location(location_obj)
                        if formatted:
                            locations.append(formatted)
                # Handle single location object
                elif isinstance(parsed, dict):
                    formatted = format_location(parsed)
                    if formatted:
                        locations.append(formatted)
                
                # Join multiple locations with newlines
                if locations:
                    return '\n'.join(locations)
                return None
            
            # Handle list-based columns (SPECIALTIES, LANGUAGES)
            if colname in ['SPECIALTIES', 'LANGUAGES'] and isinstance(parsed, list) and len(parsed) > 0:
                # Extract ALL items from the list, not just the first one
                items = []
                for item in parsed:
                    if isinstance(item, dict) and 'value' in item:
                        items.append(item['value'])
                    elif isinstance(item, str):
                        items.append(item)
                # Return as list to preserve all items
                return items if items else None
            if isinstance(parsed, dict) and 'value' in parsed:
                return parsed['value']
        except Exception:
            pass
    return val

def query_snowflake_for_npis(cursor, npi_list):
    """Query Snowflake for a list of NPIs in a single batch query and return extracted data"""
    try:
        # Create a list of quoted NPI strings for the IN clause
        npi_strings = [f"'{npi}'" for npi in npi_list]
        npi_list_str = ', '.join(npi_strings)
        
        # Single query for all NPIs
        query = f"""
        SELECT * FROM merged_provider
        WHERE NPI:value::string IN ({npi_list_str})
        """
        
        cursor.execute(query)
        results = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        
        if not results:
            return {}
        
        df = pd.DataFrame(results, columns=columns)
        
        # Remove timezone info from all datetime columns
        for col in df.select_dtypes(include=['datetimetz']).columns:
            df[col] = df[col].dt.tz_localize(None)
        for col in df.columns:
            if df[col].dtype == 'object':
                if df[col].apply(lambda x: hasattr(x, 'tzinfo') and x.tzinfo is not None).any():
                    df[col] = df[col].apply(lambda x: x.tz_localize(None) if hasattr(x, 'tzinfo') and x.tzinfo is not None else x)
        
        # Select all required columns
        selected_columns = [
            'EVENT_ID', 'EVENT_NAME', 'EVENT_SOURCE', 'FIRST_NAME', 'LAST_NAME', 'SUFFIX',
            'GENDER', 'PHONE_NUMBER', 'LANGUAGES', 'SPECIALTIES', 'LOCATIONS',
            'DISCIPLINARY_ACTIONS', 'SUPERVISING_PROVIDER', 'NPI', 'LICENSE', 'OIG',
            'VERSION', 'APPROXIMATE_CREATION_DATE_TIME', 'RECORD_FORMAT', 'TABLE_NAME',
            'MODEL_TYPE', '_ZDDE_DEDUPE_KEY', '_ZDDE_EVENT_TIMESTAMP',
            '_ZDDE_LANDED_SILVER_DATALAKE_TIMESTAMP', '_ZDDE_CPRA_DELETED_FIELDS',
            '_ZDDE_IS_CPRA_DELETED', '_ZDDE_LANDED_BRONZE_DATALAKE_TIMESTAMP',
            '_ZDDE_LANDED_DATALAKE_TIMESTAMP', '_ZDDE_MODEL_TYPE',
            '_ZDDE_ORIGINATING_FILE_PATH', '_ZDDE_ROW_HASH'
        ]
        
        # Only select columns that exist in the DataFrame
        available_columns = [col for col in selected_columns if col in df.columns]
        df_selected = df[available_columns].copy()
        
        # Extract 'value' from JSON strings in each cell
        for col in available_columns:
            df_selected[col] = df_selected[col].apply(lambda x: extract_value(x, col))
        
        # For SPECIALTIES and LANGUAGES, expand rows where they are lists (one row per item)
        # This ensures we capture all items even if they're in a single row
        expanded_rows = []
        for idx, row in df_selected.iterrows():
            row_dict = row.to_dict()
            specialties = row_dict.get('SPECIALTIES')
            languages = row_dict.get('LANGUAGES')
            
            # Handle SPECIALTIES expansion
            if isinstance(specialties, list) and len(specialties) > 0:
                # If SPECIALTIES is a list, create one row per specialty
                for specialty in specialties:
                    if specialty:  # Only add non-empty specialties
                        new_row = row_dict.copy()
                        new_row['SPECIALTIES'] = specialty
                        expanded_rows.append(new_row)
            elif specialties and pd.notna(specialties) and str(specialties).strip() != '':
                # If it's a single specialty value, keep it as is
                expanded_rows.append(row_dict.copy())
            else:
                # If no specialties, still add the row
                expanded_rows.append(row_dict.copy())
        
        # Handle LANGUAGES expansion (similar to SPECIALTIES)
        final_rows = []
        for row_dict in expanded_rows:
            languages = row_dict.get('LANGUAGES')
            
            if isinstance(languages, list) and len(languages) > 0:
                # If LANGUAGES is a list, create one row per language
                for language in languages:
                    if language:  # Only add non-empty languages
                        new_row = row_dict.copy()
                        new_row['LANGUAGES'] = language
                        final_rows.append(new_row)
            elif languages and pd.notna(languages) and str(languages).strip() != '':
                # If it's a single language value, keep it as is
                final_rows.append(row_dict.copy())
            else:
                # If no languages, still add the row
                final_rows.append(row_dict.copy())
        
        # Create new DataFrame with expanded SPECIALTIES and LANGUAGES
        if final_rows:
            df_expanded = pd.DataFrame(final_rows)
        else:
            df_expanded = pd.DataFrame(columns=available_columns)
        
        # Remove duplicate rows based on NPI, SPECIALTIES, and LANGUAGES
        # This ensures we don't have duplicate combinations
        df_expanded = df_expanded.drop_duplicates(
            subset=['NPI', 'SPECIALTIES', 'LANGUAGES'] if 'SPECIALTIES' in df_expanded.columns and 'LANGUAGES' in df_expanded.columns 
            else ['NPI'] if 'NPI' in df_expanded.columns 
            else [],
            keep='first'
        )
        
        # Group by NPI and aggregate SPECIALTIES and LANGUAGES
        results_dict = {}
        for npi in npi_list:
            npi_data = df_expanded[df_expanded['NPI'].astype(str) == str(npi)]
            
            if len(npi_data) > 0:
                # Get first row for all other columns
                first_row = npi_data.iloc[0].to_dict()
                
                # Collect ALL unique specialties for this NPI
                if 'SPECIALTIES' in npi_data.columns:
                    all_specialties = npi_data['SPECIALTIES'].dropna().unique().tolist()
                    all_specialties = [str(s).strip() for s in all_specialties if s and str(s).strip()]
                    if all_specialties:
                        first_row['SPECIALTIES'] = ', '.join(all_specialties)
                    else:
                        first_row['SPECIALTIES'] = None
                
                # Collect ALL unique languages for this NPI
                if 'LANGUAGES' in npi_data.columns:
                    all_languages = npi_data['LANGUAGES'].dropna().unique().tolist()
                    all_languages = [str(s).strip() for s in all_languages if s and str(s).strip()]
                    if all_languages:
                        first_row['LANGUAGES'] = ', '.join(all_languages)
                    else:
                        first_row['LANGUAGES'] = None
                
                # Collect ALL unique locations for this NPI
                # LOCATIONS are already formatted as strings with newlines, so we need to combine them
                if 'LOCATIONS' in npi_data.columns:
                    all_locations = npi_data['LOCATIONS'].dropna().unique().tolist()
                    # Filter out empty strings and None values
                    all_locations = [str(s).strip() for s in all_locations if s and str(s).strip()]
                    if all_locations:
                        # Join all location strings, splitting by newlines and recombining
                        # This handles cases where one location might have multiple addresses (already with newlines)
                        # and we want to combine locations from different rows
                        combined_locations = []
                        for loc_str in all_locations:
                            # Split by newline to get individual addresses
                            addresses = loc_str.split('\n')
                            combined_locations.extend([addr.strip() for addr in addresses if addr.strip()])
                        # Remove duplicates while preserving order
                        seen = set()
                        unique_locations = []
                        for loc in combined_locations:
                            if loc not in seen:
                                seen.add(loc)
                                unique_locations.append(loc)
                        # Join with newlines
                        first_row['LOCATIONS'] = '\n'.join(unique_locations)
                    else:
                        first_row['LOCATIONS'] = None
                
                results_dict[npi] = first_row
            else:
                results_dict[npi] = None
        
        return results_dict
    except Exception as e:
        print(f"Error querying NPIs: {str(e)}")
        return {}

def create_npi_extracts(user_email, role):
    """
    Extract NPI Number column from _Mapped.xlsx, query Snowflake for each NPI,
    and create NPI-Extracts.xlsx with NPI numbers and extracted data
    
    Args:
        user_email: User email address for Snowflake connection
        role: Snowflake role for connection
    
    Returns:
        Path to the created NPI-Extracts.xlsx file
    """
    # Ensure Excel Files directory exists
    EXCEL_FILES_DIR.mkdir(parents=True, exist_ok=True)
    
    # Path to _Mapped.xlsx
    mapped_file = EXCEL_FILES_DIR / '_Mapped.xlsx'
    
    if not mapped_file.exists():
        raise FileNotFoundError(f"_Mapped.xlsx not found at {mapped_file}")
    
    # Read _Mapped.xlsx
    try:
        df = pd.read_excel(mapped_file, engine='openpyxl')
    except Exception as e:
        raise Exception(f"Failed to read _Mapped.xlsx: {str(e)}")
    
    # Check if 'NPI Number' column exists
    if 'NPI Number' not in df.columns:
        # NPI column is optional - return None to indicate it was skipped
        print("'NPI Number' column not found in _Mapped.xlsx. Skipping NPI-Extracts.xlsx creation.")
        return None, {'status': 'skipped', 'reason': 'NPI Number column not found'}
    
    # Extract NPI Number column and remove NaN/empty values
    npi_series = df['NPI Number'].dropna()
    npi_series = npi_series[npi_series.astype(str).str.strip() != '']
    
    if len(npi_series) == 0:
        # No valid NPIs found - return None to indicate it was skipped
        print("No valid NPI numbers found in 'NPI Number' column. Skipping NPI-Extracts.xlsx creation.")
        return None, {'status': 'skipped', 'reason': 'No valid NPI numbers found'}
    
    # Convert NPIs to string, removing .0 suffix if present, and get unique values
    # Handle both float and string formats
    npi_list = []
    for npi in npi_series.unique():
        # Convert to string and remove .0 suffix if present
        npi_str = str(npi).replace('.0', '') if isinstance(npi, float) else str(npi)
        npi_str = npi_str.strip()
        if npi_str and npi_str not in npi_list:
            npi_list.append(npi_str)
    
    total_npis = len(npi_list)
    print(f"Processing {total_npis} unique NPIs...")
    
    # Connect to Snowflake
    try:
        conn = snowflake.connector.connect(
            user=user_email,
            account="OLIKNSY-ZOCDOC_001",
            warehouse="USER_QUERY_WH",
            database='CISTERN',
            schema='PROVIDER_PREFILL',
            role=role,
            authenticator='externalbrowser'
        )
    except Exception as e:
        raise Exception(f"Failed to connect to Snowflake: {str(e)}")
    
    # Query Snowflake for all NPIs in a single batch query
    results_list = []
    try:
        cs = conn.cursor()
        
        total_npis = len(npi_list)
        print(f"Querying Snowflake for {total_npis} NPIs in batch...")
        print(f"NPI extraction progress: 0/{total_npis} NPIs processed")
        
        npi_results = query_snowflake_for_npis(cs, npi_list)
        
        # Process results for each NPI
        processed_count = 0
        for i, npi in enumerate(npi_list, 1):
            result = npi_results.get(npi)
            
            if result:
                # Ensure NPI Number is set to the input NPI (not from Snowflake)
                # Convert to string to avoid .0 suffix
                result['NPI Number'] = str(npi).replace('.0', '') if isinstance(npi, float) else str(npi)
                # Keep 'NPI' column as well (it's in the required columns list)
                results_list.append(result)
            else:
                # Add row with NPI but no data - create empty dict with all expected columns
                # Convert to string to avoid .0 suffix
                npi_str = str(npi).replace('.0', '') if isinstance(npi, float) else str(npi)
                empty_row = {'NPI Number': npi_str}
                # Add all other columns as None
                for col in ['EVENT_ID', 'EVENT_NAME', 'EVENT_SOURCE', 'FIRST_NAME', 'LAST_NAME', 'SUFFIX',
                           'GENDER', 'PHONE_NUMBER', 'LANGUAGES', 'SPECIALTIES', 'LOCATIONS',
                           'DISCIPLINARY_ACTIONS', 'SUPERVISING_PROVIDER', 'NPI', 'LICENSE', 'OIG',
                           'VERSION', 'APPROXIMATE_CREATION_DATE_TIME', 'RECORD_FORMAT', 'TABLE_NAME',
                           'MODEL_TYPE', '_ZDDE_DEDUPE_KEY', '_ZDDE_EVENT_TIMESTAMP',
                           '_ZDDE_LANDED_SILVER_DATALAKE_TIMESTAMP', '_ZDDE_CPRA_DELETED_FIELDS',
                           '_ZDDE_IS_CPRA_DELETED', '_ZDDE_LANDED_BRONZE_DATALAKE_TIMESTAMP',
                           '_ZDDE_LANDED_DATALAKE_TIMESTAMP', '_ZDDE_MODEL_TYPE',
                           '_ZDDE_ORIGINATING_FILE_PATH', '_ZDDE_ROW_HASH']:
                    empty_row[col] = None
                results_list.append(empty_row)
            
            processed_count = i
            # Log progress every 10 NPIs or on the last one
            if processed_count % 10 == 0 or processed_count == total_npis:
                print(f"NPI extraction progress: {processed_count}/{total_npis} NPIs processed")
        
        cs.close()
        print(f"Successfully queried {total_npis} NPIs")
    except Exception as e:
        conn.close()
        raise Exception(f"Error querying Snowflake: {str(e)}")
    finally:
        conn.close()
    
    # Create DataFrame from results
    if results_list:
        results_df = pd.DataFrame(results_list)
        
        # Reorder columns to have NPI Number first, then all other columns in specified order
        column_order = [
            'NPI Number', 'EVENT_ID', 'EVENT_NAME', 'EVENT_SOURCE', 'FIRST_NAME', 'LAST_NAME', 'SUFFIX',
            'GENDER', 'PHONE_NUMBER', 'LANGUAGES', 'SPECIALTIES', 'LOCATIONS',
            'DISCIPLINARY_ACTIONS', 'SUPERVISING_PROVIDER', 'NPI', 'LICENSE', 'OIG',
            'VERSION', 'APPROXIMATE_CREATION_DATE_TIME', 'RECORD_FORMAT', 'TABLE_NAME',
            'MODEL_TYPE', '_ZDDE_DEDUPE_KEY', '_ZDDE_EVENT_TIMESTAMP',
            '_ZDDE_LANDED_SILVER_DATALAKE_TIMESTAMP', '_ZDDE_CPRA_DELETED_FIELDS',
            '_ZDDE_IS_CPRA_DELETED', '_ZDDE_LANDED_BRONZE_DATALAKE_TIMESTAMP',
            '_ZDDE_LANDED_DATALAKE_TIMESTAMP', '_ZDDE_MODEL_TYPE',
            '_ZDDE_ORIGINATING_FILE_PATH', '_ZDDE_ROW_HASH'
        ]
        # Add any additional columns that might exist
        for col in results_df.columns:
            if col not in column_order:
                column_order.append(col)
        
        results_df = results_df[[col for col in column_order if col in results_df.columns]]
        
        # Convert NPI Number column to string to avoid .0 suffix
        if 'NPI Number' in results_df.columns:
            # Convert to string, removing .0 if present
            results_df['NPI Number'] = results_df['NPI Number'].astype(str).str.replace(r'\.0$', '', regex=True)
            # Also handle the 'NPI' column if it exists
            if 'NPI' in results_df.columns:
                results_df['NPI'] = results_df['NPI'].astype(str).str.replace(r'\.0$', '', regex=True)
    else:
        # Create empty DataFrame with expected columns
        results_df = pd.DataFrame(columns=[
            'NPI Number', 'EVENT_ID', 'EVENT_NAME', 'EVENT_SOURCE', 'FIRST_NAME', 'LAST_NAME', 'SUFFIX',
            'GENDER', 'PHONE_NUMBER', 'LANGUAGES', 'SPECIALTIES', 'LOCATIONS',
            'DISCIPLINARY_ACTIONS', 'SUPERVISING_PROVIDER', 'NPI', 'LICENSE', 'OIG',
            'VERSION', 'APPROXIMATE_CREATION_DATE_TIME', 'RECORD_FORMAT', 'TABLE_NAME',
            'MODEL_TYPE', '_ZDDE_DEDUPE_KEY', '_ZDDE_EVENT_TIMESTAMP',
            '_ZDDE_LANDED_SILVER_DATALAKE_TIMESTAMP', '_ZDDE_CPRA_DELETED_FIELDS',
            '_ZDDE_IS_CPRA_DELETED', '_ZDDE_LANDED_BRONZE_DATALAKE_TIMESTAMP',
            '_ZDDE_LANDED_DATALAKE_TIMESTAMP', '_ZDDE_MODEL_TYPE',
            '_ZDDE_ORIGINATING_FILE_PATH', '_ZDDE_ROW_HASH'
        ])
    
    # Path to output file
    output_file = EXCEL_FILES_DIR / 'NPI-Extracts.xlsx'
    
    # Path to reference file with ValidationAndReference sheet
    reference_file = BACKEND_DIR / 'New Business Scope Sheet - Practice Locations and Providers.xlsx'
    
    # Read ValidationAndReference sheet and create lookup dictionaries for 'Suffix Derived' and 'Specialty Derived' columns
    lookup_dict = {}  # For Suffix: ProfSuffixID -> Suffix
    specialty_lookup_dict = {}  # For Specialty: Specialty ID -> Specialty Name
    validation_df = None
    if reference_file.exists():
        try:
            # Read the ValidationAndReference sheet from the reference file
            validation_df = pd.read_excel(reference_file, sheet_name='ValidationAndReference', engine='openpyxl')
            print(f"Successfully read 'ValidationAndReference' sheet from reference file")
            
            # Create lookup dictionary: 'ProfSuffixID' -> 'Suffix'
            # Lookup: Take value from 'SUFFIX' in NPI Extracts, search in 'ProfSuffixID' in ValidationAndReference,
            # then get corresponding value from 'Suffix' in ValidationAndReference
            if validation_df is not None:
                print(f"ValidationAndReference sheet has {len(validation_df)} rows and {len(validation_df.columns)} columns")
                print(f"Column names in ValidationAndReference: {list(validation_df.columns)}")
                
                # Use specific column names: 'ProfSuffixID' (search column) -> 'Suffix' (return column)
                search_col_name = 'ProfSuffixID'  # Column to search in ValidationAndReference
                return_col_name = 'Suffix'  # Column to get value from ValidationAndReference
                
                # Check if columns exist
                if search_col_name in validation_df.columns and return_col_name in validation_df.columns:
                    print(f"Using column names: Search column = '{search_col_name}', Return column = '{return_col_name}'")
                    for idx, row in validation_df.iterrows():
                        search_value = row[search_col_name] if search_col_name in row else None
                        return_value = row[return_col_name] if return_col_name in row else None
                        if pd.notna(search_value) and pd.notna(return_value):
                            # Convert to string for comparison, normalize whitespace
                            # Handle both text and numeric values
                            if isinstance(search_value, (int, float)):
                                search_key = str(int(search_value)) if isinstance(search_value, float) and search_value.is_integer() else str(search_value)
                            else:
                                search_key = str(search_value).strip()
                            # Normalize to uppercase for case-insensitive matching
                            search_key = search_key.upper()
                            
                            return_val = str(return_value).strip()
                            if search_key and return_val:
                                # Store both string and numeric versions if it's a number
                                lookup_dict[search_key] = return_val
                                # Also store as numeric key if it's a number (for reverse lookup)
                                try:
                                    if '.' in search_key:
                                        lookup_dict[float(search_key)] = return_val
                                    else:
                                        lookup_dict[int(search_key)] = return_val
                                except (ValueError, TypeError):
                                    pass  # Not a number, just use string key
                    
                    print(f"Created lookup dictionary for Suffix with {len(lookup_dict)} entries")
                    if len(lookup_dict) > 0:
                        print(f"Sample Suffix lookup entries (first 5): {list(lookup_dict.items())[:5]}")
                    else:
                        print(f"WARNING: Suffix lookup dictionary is empty! Check if '{search_col_name}' and '{return_col_name}' have data.")
                        # Show sample values from search and return columns for debugging
                        if len(validation_df) > 0:
                            print(f"Sample values from '{search_col_name}' (first 5 rows): {validation_df[search_col_name].dropna().head(5).tolist()}")
                            print(f"Sample values from '{return_col_name}' (first 5 rows): {validation_df[return_col_name].dropna().head(5).tolist()}")
                
                # Create lookup dictionary for Specialty: 'Specialty ID' -> 'Specialty Name'
                specialty_search_col = 'Specialty ID'  # Column to search in ValidationAndReference
                specialty_return_col = 'Specialty Name'  # Column to get value from ValidationAndReference
                
                if specialty_search_col in validation_df.columns and specialty_return_col in validation_df.columns:
                    print(f"Using column names for Specialty: Search column = '{specialty_search_col}', Return column = '{specialty_return_col}'")
                    for idx, row in validation_df.iterrows():
                        search_value = row[specialty_search_col] if specialty_search_col in row else None
                        return_value = row[specialty_return_col] if specialty_return_col in row else None
                        if pd.notna(search_value) and pd.notna(return_value):
                            # Handle both text and numeric values
                            if isinstance(search_value, (int, float)):
                                search_key = str(int(search_value)) if isinstance(search_value, float) and search_value.is_integer() else str(search_value)
                            else:
                                search_key = str(search_value).strip()
                            # Normalize to uppercase for case-insensitive matching
                            search_key = search_key.upper()
                            
                            return_val = str(return_value).strip()
                            if search_key and return_val:
                                # Store both string and numeric versions if it's a number
                                specialty_lookup_dict[search_key] = return_val
                                # Also store as numeric key if it's a number
                                try:
                                    if '.' in search_key:
                                        specialty_lookup_dict[float(search_key)] = return_val
                                    else:
                                        specialty_lookup_dict[int(search_key)] = return_val
                                except (ValueError, TypeError):
                                    pass  # Not a number, just use string key
                    
                    print(f"Created lookup dictionary for Specialty with {len(specialty_lookup_dict)} entries")
                    if len(specialty_lookup_dict) > 0:
                        print(f"Sample Specialty lookup entries (first 5): {list(specialty_lookup_dict.items())[:5]}")
                    else:
                        print(f"WARNING: Specialty lookup dictionary is empty! Check if '{specialty_search_col}' and '{specialty_return_col}' have data.")
                        if len(validation_df) > 0:
                            print(f"Sample values from '{specialty_search_col}' (first 5 rows): {validation_df[specialty_search_col].dropna().head(5).tolist()}")
                            print(f"Sample values from '{specialty_return_col}' (first 5 rows): {validation_df[specialty_return_col].dropna().head(5).tolist()}")
                else:
                    print(f"Warning: Specialty columns not found in ValidationAndReference sheet")
                    print(f"Looking for: '{specialty_search_col}' and '{specialty_return_col}'")
                    print(f"Available columns: {list(validation_df.columns)}")
                
                if search_col_name not in validation_df.columns or return_col_name not in validation_df.columns:
                    print(f"Warning: Required Suffix columns not found in ValidationAndReference sheet")
                    print(f"Looking for: '{search_col_name}' and '{return_col_name}'")
                    print(f"Available columns: {list(validation_df.columns)}")
            else:
                print(f"Warning: ValidationAndReference sheet is None or empty")
        except Exception as e:
            print(f"Warning: Could not read 'ValidationAndReference' sheet: {str(e)}")
    else:
        print(f"Warning: Reference file not found at {reference_file}")
    
    # Add 'Suffix Derived' column to results_df by looking up values from 'SUFFIX' column
    # Process: Take value from 'SUFFIX' in NPI Extracts -> Search in 'ProfSuffixID' in ValidationAndReference -> Get 'Suffix' from ValidationAndReference
    col_suffix_name = 'SUFFIX'  # Source column in NPI Extracts
    
    if col_suffix_name in results_df.columns:
        print(f"Processing column '{col_suffix_name}' from NPI Extracts sheet")
        print(f"Column names in results_df: {list(results_df.columns)}")
        
        # Show sample values from SUFFIX column
        sample_values = results_df[col_suffix_name].dropna().head(5).tolist()
        print(f"Sample values from '{col_suffix_name}' (first 5 non-null): {sample_values}")
        
        # Create 'Suffix Derived' column by looking up values
        suffix_derived = []
        matched_count = 0
        unmatched_samples = []
        
        for idx in results_df.index:
            original_value = results_df.loc[idx, col_suffix_name]
            if pd.notna(original_value):
                # Normalize the search key - handle both text and numeric values
                # Try multiple formats for matching
                search_keys = []
                
                # Convert to string and normalize
                if isinstance(original_value, (int, float)):
                    # If it's a number, try both string and numeric formats
                    if isinstance(original_value, float) and original_value.is_integer():
                        search_keys.append(str(int(original_value)).upper())
                        search_keys.append(int(original_value))
                    else:
                        search_keys.append(str(original_value).upper())
                        search_keys.append(float(original_value))
                else:
                    # If it's text, try string format
                    search_key_str = str(original_value).strip().upper()
                    search_keys.append(search_key_str)
                    # Also try converting to number if possible
                    try:
                        if '.' in search_key_str:
                            search_keys.append(float(search_key_str))
                        else:
                            search_keys.append(int(search_key_str))
                    except (ValueError, TypeError):
                        pass  # Not a number, just use string
                
                # Try to find a match using any of the search key formats
                found_match = False
                for search_key in search_keys:
                    if search_key in lookup_dict:
                        suffix_derived.append(lookup_dict[search_key])
                        matched_count += 1
                        found_match = True
                        break
                
                if not found_match:
                    suffix_derived.append(None)
                    # Collect some unmatched samples for debugging
                    if len(unmatched_samples) < 5:
                        unmatched_samples.append(str(original_value).strip())
            else:
                suffix_derived.append(None)
        
        # Add 'Suffix Derived' column to results_df
        results_df['Suffix Derived'] = suffix_derived
        print(f"Added 'Suffix Derived' column with {matched_count} matched values out of {len(results_df)} rows")
        
        if unmatched_samples and len(lookup_dict) > 0:
            print(f"Sample unmatched values from '{col_suffix_name}': {unmatched_samples}")
            print(f"Sample lookup keys from ValidationAndReference 'ProfSuffixID' (first 5): {list(lookup_dict.keys())[:5]}")
    else:
        print(f"Warning: Column '{col_suffix_name}' not found in results_df")
        print(f"Available columns: {list(results_df.columns)}")
        results_df['Suffix Derived'] = None
    
    # Add 'Specialty Derived' column to results_df by looking up values from 'SPECIALTIES' column
    # Process: Take value(s) from 'SPECIALTIES' in NPI Extracts -> Search in 'Specialty ID' in ValidationAndReference -> Get 'Specialty Name' from ValidationAndReference
    # Handle comma-separated values (e.g., "405, 406, 407") and join results with semicolon
    col_specialties_name = 'SPECIALTIES'  # Source column in NPI Extracts
    
    if col_specialties_name in results_df.columns:
        print(f"Processing column '{col_specialties_name}' from NPI Extracts sheet for Specialty Derived")
        
        # Show sample values from SPECIALTIES column
        sample_values = results_df[col_specialties_name].dropna().head(5).tolist()
        print(f"Sample values from '{col_specialties_name}' (first 5 non-null): {sample_values}")
        
        # Create 'Specialty Derived' column by looking up values
        specialty_derived = []
        matched_count = 0
        unmatched_samples = []
        
        for idx in results_df.index:
            original_value = results_df.loc[idx, col_specialties_name]
            if pd.notna(original_value):
                # Handle comma-separated values (e.g., "405, 406, 407, 531, 401")
                specialty_str = str(original_value).strip()
                # Split by comma to get individual specialty IDs
                specialty_ids = [s.strip() for s in specialty_str.split(',') if s.strip()]
                
                # Look up each specialty ID
                matched_specialties = []
                for specialty_id in specialty_ids:
                    if not specialty_id:
                        continue
                    
                    # Try multiple formats for matching (handle text/numeric)
                    search_keys = []
                    search_key_str = specialty_id.upper()
                    search_keys.append(search_key_str)
                    # Also try converting to number if possible
                    try:
                        if '.' in search_key_str:
                            search_keys.append(float(search_key_str))
                        else:
                            search_keys.append(int(search_key_str))
                    except (ValueError, TypeError):
                        pass  # Not a number, just use string
                    
                    # Try to find a match using any of the search key formats
                    found_match = False
                    for search_key in search_keys:
                        if search_key in specialty_lookup_dict:
                            matched_specialties.append(specialty_lookup_dict[search_key])
                            found_match = True
                            break
                    
                    if not found_match and len(unmatched_samples) < 5:
                        unmatched_samples.append(specialty_id)
                
                # Join matched specialties with semicolon
                if matched_specialties:
                    specialty_derived.append('; '.join(matched_specialties))
                    matched_count += 1
                else:
                    specialty_derived.append(None)
            else:
                specialty_derived.append(None)
        
        # Add 'Specialty Derived' column to results_df
        results_df['Specialty Derived'] = specialty_derived
        print(f"Added 'Specialty Derived' column with {matched_count} matched rows out of {len(results_df)} rows")
        
        if unmatched_samples and len(specialty_lookup_dict) > 0:
            print(f"Sample unmatched specialty IDs: {unmatched_samples}")
            print(f"Sample lookup keys from ValidationAndReference 'Specialty ID' (first 5): {list(specialty_lookup_dict.keys())[:5]}")
    else:
        print(f"Warning: Column '{col_specialties_name}' not found in results_df")
        print(f"Available columns: {list(results_df.columns)}")
        results_df['Specialty Derived'] = None
    
    # Create Excel file with formatting
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Write the NPI Extracts sheet (with 'Suffix derived' column)
            results_df.to_excel(writer, sheet_name='NPI Extracts', index=False)
            
            # Copy ValidationAndReference sheet from reference file if it exists
            if validation_df is not None:
                try:
                    # Write the ValidationAndReference sheet to the output file
                    validation_df.to_excel(writer, sheet_name='ValidationAndReference', index=False)
                    print(f"Successfully added 'ValidationAndReference' sheet to output file")
                except Exception as e:
                    print(f"Warning: Could not write 'ValidationAndReference' sheet: {str(e)}")
            elif reference_file.exists():
                try:
                    # Read the ValidationAndReference sheet from the reference file
                    validation_df = pd.read_excel(reference_file, sheet_name='ValidationAndReference', engine='openpyxl')
                    # Write it to the output file
                    validation_df.to_excel(writer, sheet_name='ValidationAndReference', index=False)
                    print(f"Successfully added 'ValidationAndReference' sheet from reference file")
                except Exception as e:
                    print(f"Warning: Could not copy 'ValidationAndReference' sheet: {str(e)}")
            else:
                print(f"Warning: Reference file not found at {reference_file}")
            
            # Get the workbook and worksheet for styling
            workbook = writer.book
            worksheet = writer.sheets['NPI Extracts']
            
            # Style the header row
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Adjust column widths (set reasonable defaults)
            for col_idx, col_name in enumerate(results_df.columns, start=1):
                col_letter = get_column_letter(col_idx)
                if col_name in ['SPECIALTIES', 'LANGUAGES', 'LOCATIONS', 'DISCIPLINARY_ACTIONS']:
                    worksheet.column_dimensions[col_letter].width = 30
                elif col_name in ['APPROXIMATE_CREATION_DATE_TIME', '_ZDDE_ORIGINATING_FILE_PATH']:
                    worksheet.column_dimensions[col_letter].width = 25
                elif col_name.startswith('_ZDDE_'):
                    worksheet.column_dimensions[col_letter].width = 20
                else:
                    worksheet.column_dimensions[col_letter].width = 18
            
            # Define column colors
            # Light green columns: EVENT_ID, EVENT_NAME, EVENT_SOURCE, FIRST_NAME, LAST_NAME, SUFFIX, GENDER, LANGUAGES, SPECIALTIES, LOCATIONS, NPI
            light_green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            # Light grey columns: All _ZDDE_ columns (columns 22-31)
            light_grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            # Get column names to find indices
            column_names = list(results_df.columns)
            light_green_columns = ['EVENT_ID', 'EVENT_NAME', 'EVENT_SOURCE', 'FIRST_NAME', 'LAST_NAME', 
                                  'SUFFIX', 'GENDER', 'LANGUAGES', 'SPECIALTIES', 'LOCATIONS', 'NPI']
            
            # Apply colors to data cells
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                for col_idx, cell in enumerate(row, start=1):
                    if col_idx <= len(column_names):
                        col_name = column_names[col_idx - 1]
                        
                        # Apply light green to specified columns
                        if col_name in light_green_columns:
                            cell.fill = light_green_fill
                        # Apply light grey to _ZDDE_ columns
                        elif col_name.startswith('_ZDDE_'):
                            cell.fill = light_grey_fill
                        
                        # Enable text wrapping for LOCATIONS column to display newlines
                        if col_name == 'LOCATIONS':
                            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                        else:
                            # Center align all other data cells
                            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        print(f"Successfully created NPI-Extracts.xlsx with {len(results_df)} rows")
        
        # Return file path and extraction info
        extraction_info = {
            'total_npis': total_npis,
            'processed_count': len(results_list),
            'successful_count': len([r for r in results_list if r.get('FIRST_NAME')]),
            'status': 'completed'
        }
        
        return str(output_file), extraction_info
    except Exception as e:
        raise Exception(f"Failed to create NPI-Extracts.xlsx: {str(e)}")

if __name__ == '__main__':
    # For testing purposes - requires email and role
    if len(sys.argv) >= 3:
        user_email = sys.argv[1]
        role = sys.argv[2]
        try:
            output_path = create_npi_extracts(user_email, role)
            print(f"Successfully created NPI-Extracts.xlsx at: {output_path}")
        except Exception as e:
            print(f"Error: {str(e)}")
            sys.exit(1)
    else:
        print("Usage: python NPIsnowflake.py <user_email> <role>")
        print("Example: python NPIsnowflake.py abc.xyz@zocdoc.com PROD_OPS_PUNE_ROLE")

