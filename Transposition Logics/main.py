import shutil
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Import column extraction functions
from NPI import extract_npi_to_template
from Name import extract_names_to_template
from Gender import extract_gender_to_template
from Professionalsuffix import extract_professional_suffix_to_template
from specialty import extract_specialty_to_template
from AdditionalLanguages import extract_additional_languages_to_template
from Headshot import extract_headshot_to_template
from PFS import extract_pfs_to_template
from Locationcloudid import extract_location_cloud_id_to_template
from LocationDetails import extract_location_details_to_template
from practicemngmt import extract_practice_cloud_id_to_template
from Dropdown import apply_dropdowns_to_template
from Formulas import apply_formulas_to_template
from PatientsAccepted import apply_patients_accepted_to_template
from fallbacks import apply_specialty_fallback_from_npi_extracts, apply_professional_suffix_fallback_from_npi_extracts, apply_gender_fallback_from_npi_extracts, apply_additional_languages_fallback_from_npi_extracts
from similarprovider import merge_duplicate_providers
from checkdiff import check_differences

# Get the absolute path of the source file relative to the script location
script_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(script_dir)

# Source template file path
source_file = os.path.join(project_root, 'backend', 'New Business Scope Sheet - Practice Locations and Providers.xlsx')

# Destination file path (use dynamic path based on project root)
destination_dir = os.path.join(project_root, 'backend', 'Excel Files')
destination_file = os.path.join(destination_dir, 'Template copy.xlsx')

# Ensure the destination directory exists
os.makedirs(destination_dir, exist_ok=True)

# Check if source file exists
if not os.path.exists(source_file):
    print(f"✗ Error: Source template file not found at {source_file}")
    exit(1)

# Copy the source template file to destination
try:
    shutil.copy2(source_file, destination_file)
    print(f"✓ Successfully copied template file to {destination_file}")
except FileNotFoundError as e:
    print(f"✗ Error: Source file not found: {source_file}")
    print(f"  Details: {str(e)}")
    exit(1)
except Exception as e:
    print(f"✗ Error: Failed to copy template file")
    print(f"  Details: {str(e)}")
    exit(1)

# Verify the destination file was created
if not os.path.exists(destination_file):
    print(f"✗ Error: Template copy.xlsx was not created at {destination_file}")
    exit(1)


# List of column extraction functions to call
column_extraction_functions = [
    ("NPI Number", extract_npi_to_template),
    ("First Name & Last Name", extract_names_to_template),
    ("Gender", extract_gender_to_template),
    ("Professional Suffix", extract_professional_suffix_to_template),
    ("Specialty", extract_specialty_to_template),
    ("Additional Languages", extract_additional_languages_to_template),
    ("Headshot Link", extract_headshot_to_template),
    ("Professional Statement", extract_pfs_to_template),
    ("Location ID", extract_location_cloud_id_to_template),
    ("Location Details", extract_location_details_to_template),
    ("Practice Cloud ID", extract_practice_cloud_id_to_template),
    # Add more column extraction functions here as they are created
]

# Call each column extraction function
for column_name, extraction_func in column_extraction_functions:
    try:
        success = extraction_func()
        if success:
            print(f"✓ Successfully extracted {column_name}")
    except Exception as e:
        pass

# Apply fallback logics
try:
    success = apply_specialty_fallback_from_npi_extracts()
    if success:
        print(f"✓ Successfully Applied Specialty Fallback")
    else:
        print(f"✗ Failed to apply Specialty Fallback")
except Exception as e:
    print(f"✗ Failed to apply Specialty Fallback")

try:
    success = apply_professional_suffix_fallback_from_npi_extracts()
    if success:
        print(f"✓ Successfully Applied Professional Suffix Fallback")
    else:
        print(f"✗ Failed to apply Professional Suffix Fallback")
except Exception as e:
    print(f"✗ Failed to apply Professional Suffix Fallback")

try:
    success = apply_gender_fallback_from_npi_extracts()
    if success:
        print(f"✓ Successfully Applied Gender Fallback")
    else:
        print(f"✗ Failed to apply Gender Fallback")
except Exception as e:
    print(f"✗ Failed to apply Gender Fallback")

try:
    success = apply_additional_languages_fallback_from_npi_extracts()
    if success:
        print(f"✓ Successfully Applied Additional Languages Fallback")
    else:
        print(f"✗ Failed to apply Additional Languages Fallback")
except Exception as e:
    print(f"✗ Failed to apply Additional Languages Fallback")

# Merge duplicate providers based on NPI Number
try:
    success = merge_duplicate_providers()
    if success:
        print(f"✓ Successfully Merged Duplicate Providers")
    else:
        print(f"✗ Failed to merge duplicate providers")
except Exception as e:
    print(f"✗ Failed to merge duplicate providers")

# Check differences between Template copy.xlsx and NPI-Extracts.xlsx
try:
    success = check_differences()
    if success:
        print(f"✓ Successfully Checked Differences")
    else:
        print(f"✗ Failed to check differences")
except Exception as e:
    print(f"✗ Failed to check differences")

# Auto-fit column widths and freeze header rows for all sheets in Template copy.xlsx
try:
    if not os.path.exists(destination_file):
        print(f"✗ Error: Template copy.xlsx not found at {destination_file}")
        raise FileNotFoundError(f"Template copy.xlsx not found at {destination_file}")
    wb = load_workbook(destination_file)
    
    # Sheets that need header row frozen
    sheets_to_freeze = ['Provider', 'Location', 'ValidationAndReference']
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # Auto-fit columns by finding the maximum width needed for each column
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            # Check header cell
            if column[0].value:
                max_length = len(str(column[0].value))
            
            # Check data cells
            for cell in column[1:]:
                if cell.value:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
            
            # Set column width (add some padding)
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row (row 1) for specified sheets
        if sheet_name in sheets_to_freeze:
            sheet.freeze_panes = 'A2'  # Freezes row 1 and column A
    
    wb.save(destination_file)
except Exception as e:
    pass

# Apply formulas to Template copy.xlsx
try:
    success = apply_formulas_to_template()
    if success:
        print(f"✓ Successfully Applied Formulas")
    else:
        print(f"✗ Failed to apply formulas")
except Exception as e:
    print(f"✗ Failed to apply formulas")

# Apply dropdowns to Template copy.xlsx (should be called after formulas)
try:
    success = apply_dropdowns_to_template()
    if success:
        print(f"✓ Successfully Applied Dropdowns")
    else:
        print(f"✗ Failed to apply dropdowns")
except Exception as e:
    print(f"✗ Failed to apply dropdowns")

# Apply Patients Accepted dropdowns and logic
try:
    success = apply_patients_accepted_to_template()
    if success:
        print(f"✓ Successfully Applied Patients Accepted")
    else:
        print(f"✗ Failed to apply Patients Accepted")
except Exception as e:
    print(f"✗ Failed to apply Patients Accepted")

