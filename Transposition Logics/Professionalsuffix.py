"""
Professional Suffix - Extracts Professional Suffix columns from _Mapped.xlsx and writes to Template copy.xlsx

This script reads the 'Professional Suffix 1-3' column from backend\Excel Files\_Mapped.xlsx
and writes it to the 'Professional Suffix 1', 'Professional Suffix 2', and 'Professional Suffix 3' columns
in the 'Provider' sheet in Template copy.xlsx. Comma-separated values are split across the columns.
It also changes the header colors to green.
"""

import os
import re
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Get the project root directory (parent of Transposition Logics folder)
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
BACKEND_DIR = PROJECT_ROOT / 'backend'
EXCEL_FILES_DIR = BACKEND_DIR / 'Excel Files'

def normalize_suffix(suffix_str):
    """
    Normalize a suffix string by removing punctuation, spaces, and converting to uppercase.
    Example: "M.D." -> "MD", "M D" -> "MD"
    """
    if not suffix_str:
        return ""
    # Remove all punctuation and spaces, convert to uppercase
    normalized = re.sub(r'[^\w]', '', str(suffix_str).upper())
    return normalized

def find_close_match(value, valid_suffixes):
    """
    Find a close match for a value in the valid suffixes list.
    Returns the original valid suffix if a close match is found, None otherwise.
    Checks for:
    1. Exact match (case-insensitive) - no replacement needed
    2. Normalized match (without punctuation/spaces) - e.g., "M.D." matches "MD"
    3. Prefix match - e.g., "ABC" matches "ABCd"
    """
    if not value or not valid_suffixes:
        return None
    
    normalized_value = normalize_suffix(value)
    
    # First check exact match (case-insensitive)
    for valid_suffix in valid_suffixes:
        if str(value).strip().lower() == str(valid_suffix).strip().lower():
            return None  # Exact match, no need to replace
    
    # Check normalized match (without punctuation/spaces)
    for valid_suffix in valid_suffixes:
        normalized_valid = normalize_suffix(valid_suffix)
        if normalized_value == normalized_valid:
            return valid_suffix  # Close match found, return the valid version
    
    # Check if normalized input is a prefix of any normalized valid suffix
    # Only match if the input is at least 2 characters to avoid too broad matches
    if len(normalized_value) >= 2:
        for valid_suffix in valid_suffixes:
            normalized_valid = normalize_suffix(valid_suffix)
            # Check if input is a prefix of valid suffix
            if normalized_valid.startswith(normalized_value):
                return valid_suffix  # Prefix match found, return the valid version
    
    return None

def extract_professional_suffix_to_template():
    """
    Extract Professional Suffix 1-3 column from _Mapped.xlsx and write to Template copy.xlsx
    Comma-separated values are split across Professional Suffix 1, 2, and 3 columns
    
    Returns:
        bool: True if successful, False otherwise
    """
    # File paths
    source_file = EXCEL_FILES_DIR / '_Mapped.xlsx'
    template_file = EXCEL_FILES_DIR / 'Template copy.xlsx'
    
    # Check if source file exists
    if not source_file.exists():
        return False
    
    # Check if template file exists
    if not template_file.exists():
        return False
    
    try:
        # Read the Professional Suffix 1-3 column from _Mapped.xlsx
        mapped_df = pd.read_excel(source_file)
        
        # Check if 'Professional Suffix 1-3' column exists
        if 'Professional Suffix 1-3' not in mapped_df.columns:
            return False
        
        # Extract the Professional Suffix 1-3 column
        suffix_data = mapped_df['Professional Suffix 1-3'].tolist()
        
        # Load the template workbook
        wb = load_workbook(template_file)
        
        # Check if 'Provider' sheet exists
        if 'Provider' not in wb.sheetnames:
            return False
        
        # Get the Provider sheet
        provider_sheet = wb['Provider']
        
        # Find the Professional Suffix column headers
        header_row = 1
        suffix1_column_letter = None
        suffix2_column_letter = None
        suffix3_column_letter = None
        
        # Search for the headers in the first row
        for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
            if cell.value:
                cell_value = str(cell.value).strip().lower()
                if cell_value == 'professional suffix 1':
                    suffix1_column_letter = get_column_letter(col_idx)
                elif cell_value == 'professional suffix 2':
                    suffix2_column_letter = get_column_letter(col_idx)
                elif cell_value == 'professional suffix 3':
                    suffix3_column_letter = get_column_letter(col_idx)
        
        if suffix1_column_letter is None:
            return False
        
        # Load valid suffixes from ValidationAndReference sheet
        valid_suffixes = set()
        if 'ValidationAndReference' in wb.sheetnames:
            validation_sheet = wb['ValidationAndReference']
            suffix_column_letter = None
            
            # Find the 'Suffix' column in ValidationAndReference sheet
            for col_idx, cell in enumerate(validation_sheet[header_row], start=1):
                if cell.value and str(cell.value).strip().lower() == 'suffix':
                    suffix_column_letter = get_column_letter(col_idx)
                    break
            
            # Read all valid suffix values
            if suffix_column_letter:
                max_row = validation_sheet.max_row
                for row_idx in range(2, max_row + 1):  # Start from row 2 (skip header)
                    cell = validation_sheet[f"{suffix_column_letter}{row_idx}"]
                    if cell.value:
                        valid_suffixes.add(str(cell.value).strip())
        
        # Grey fill for invalid values
        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Process and write the Professional Suffix data
        for row_idx, suffix_value in enumerate(suffix_data, start=2):
            # Split comma-separated values
            if pd.isna(suffix_value):
                suffix_parts = [None, None, None]
            else:
                suffix_str = str(suffix_value).strip()
                if not suffix_str:
                    suffix_parts = [None, None, None]
                else:
                    # Split by comma and strip whitespace from each part
                    # Convert empty strings to None
                    parts = [part.strip() if part.strip() else None for part in suffix_str.split(',')]
                    # Take only first 3 parts and pad with None if needed
                    suffix_parts = parts[:3]
                    while len(suffix_parts) < 3:
                        suffix_parts.append(None)
            
            # Write to Professional Suffix 1
            if suffix1_column_letter:
                cell1 = provider_sheet[f"{suffix1_column_letter}{row_idx}"]
                value1 = suffix_parts[0] if suffix_parts[0] else None
                
                # Check for close match and replace if found
                if value1 and valid_suffixes:
                    close_match = find_close_match(value1, valid_suffixes)
                    if close_match:
                        cell1.value = close_match
                        cell1.fill = grey_fill  # Highlight corrected values
                    elif value1 not in valid_suffixes:
                        cell1.value = value1
                        cell1.fill = grey_fill  # Highlight invalid values
                    else:
                        cell1.value = value1
                else:
                    cell1.value = value1
            
            # Write to Professional Suffix 2
            if suffix2_column_letter:
                cell2 = provider_sheet[f"{suffix2_column_letter}{row_idx}"]
                value2 = suffix_parts[1] if suffix_parts[1] else None
                
                # Check for close match and replace if found
                if value2 and valid_suffixes:
                    close_match = find_close_match(value2, valid_suffixes)
                    if close_match:
                        cell2.value = close_match
                        cell2.fill = grey_fill  # Highlight corrected values
                    elif value2 not in valid_suffixes:
                        cell2.value = value2
                        cell2.fill = grey_fill  # Highlight invalid values
                    else:
                        cell2.value = value2
                else:
                    cell2.value = value2
            
            # Write to Professional Suffix 3
            if suffix3_column_letter:
                cell3 = provider_sheet[f"{suffix3_column_letter}{row_idx}"]
                value3 = suffix_parts[2] if suffix_parts[2] else None
                
                # Check for close match and replace if found
                if value3 and valid_suffixes:
                    close_match = find_close_match(value3, valid_suffixes)
                    if close_match:
                        cell3.value = close_match
                        cell3.fill = grey_fill  # Highlight corrected values
                    elif value3 not in valid_suffixes:
                        cell3.value = value3
                        cell3.fill = grey_fill  # Highlight invalid values
                    else:
                        cell3.value = value3
                else:
                    cell3.value = value3
        
        # Color the header cells green
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        
        if suffix1_column_letter:
            header1 = provider_sheet[f"{suffix1_column_letter}{header_row}"]
            header1.fill = green_fill
        
        if suffix2_column_letter:
            header2 = provider_sheet[f"{suffix2_column_letter}{header_row}"]
            header2.fill = green_fill
        
        if suffix3_column_letter:
            header3 = provider_sheet[f"{suffix3_column_letter}{header_row}"]
            header3.fill = green_fill
        
        # Save the workbook
        wb.save(template_file)
        return True
        
    except Exception as e:
        return False

if __name__ == "__main__":
    extract_professional_suffix_to_template()

