"""
Gender - Extracts Gender column from _Mapped.xlsx and writes to Template copy.xlsx

This script reads the 'Gender' column from backend\Excel Files\_Mapped.xlsx
and writes it to the 'Provider' sheet in Template copy.xlsx under the 'Gender' column.
It also changes the header color to green.
"""

import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Get the project root directory (parent of Transposition Logics folder)
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
BACKEND_DIR = PROJECT_ROOT / 'backend'
EXCEL_FILES_DIR = BACKEND_DIR / 'Excel Files'

def extract_gender_to_template():
    """
    Extract Gender column from _Mapped.xlsx and write to Template copy.xlsx
    
    Returns:
        bool: True if successful, False otherwise
    """
    # File paths
    source_file = EXCEL_FILES_DIR / '_Mapped.xlsx'
    template_file = EXCEL_FILES_DIR / 'Template copy.xlsx'
    
    # Check if source file exists
    if not source_file.exists():
        return False
    
    # Check if template file exists
    if not template_file.exists():
        return False
    
    try:
        # Read the Gender column from _Mapped.xlsx
        mapped_df = pd.read_excel(source_file)
        
        # Check if 'Gender' column exists
        if 'Gender' not in mapped_df.columns:
            return False
        
        # Extract the Gender column
        gender_data = mapped_df['Gender'].tolist()
        
        # Load the template workbook
        wb = load_workbook(template_file)
        
        # Check if 'Provider' sheet exists
        if 'Provider' not in wb.sheetnames:
            return False
        
        # Get the Provider sheet
        provider_sheet = wb['Provider']
        
        # Find the 'Gender' column header
        header_row = 1
        gender_column_index = None
        gender_column_letter = None
        
        # Search for the header in the first row
        for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
            if cell.value and str(cell.value).strip().lower() == 'gender':
                gender_column_index = col_idx
                gender_column_letter = get_column_letter(col_idx)
                break
        
        if gender_column_index is None:
            return False
        
        # Valid gender values
        valid_genders = ['Male', 'Female', 'NonBinary', 'Not Applicable']
        
        # Grey fill for invalid values
        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Write the Gender data to the column (starting from row 2, as row 1 is the header)
        for row_idx, gender_value in enumerate(gender_data, start=2):
            cell = provider_sheet[f"{gender_column_letter}{row_idx}"]
            # Convert to string and handle NaN values
            if pd.isna(gender_value):
                cell.value = None
            else:
                gender_str = str(gender_value).strip() if gender_value else ""
                
                # Convert M to Male and F to Female (case-insensitive)
                if gender_str.upper() == 'M':
                    gender_str = 'Male'
                elif gender_str.upper() == 'F':
                    gender_str = 'Female'
                
                cell.value = gender_str if gender_str else None
                
                # Check if the value is not one of the valid genders (case-insensitive)
                if cell.value:
                    # Case-insensitive check against valid genders
                    if cell.value not in valid_genders and cell.value.lower() not in [v.lower() for v in valid_genders]:
                        cell.fill = grey_fill
        
        # Color the header cell green
        header_cell = provider_sheet[f"{gender_column_letter}{header_row}"]
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        header_cell.fill = green_fill
        
        # Save the workbook
        wb.save(template_file)
        return True
        
    except Exception as e:
        return False

if __name__ == "__main__":
    extract_gender_to_template()

