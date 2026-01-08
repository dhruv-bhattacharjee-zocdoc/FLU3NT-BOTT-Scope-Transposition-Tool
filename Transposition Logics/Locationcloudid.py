"""
Location Cloud ID - Extracts Location Cloud ID and Practice Cloud ID values from Locations_input.xlsx and writes to Template copy.xlsx

This script reads the 'NPI Number' column from the 'Provider' sheet in Template copy.xlsx,
searches for matching NPI values in Locations_input.xlsx (Locations sheet),
and extracts 'Location Cloud ID 1' through 'Location Cloud ID 5' values,
writing them to 'Location ID 1' through 'Location ID 5' columns in the Provider sheet.
It also extracts 'Practice Cloud ID 1' through 'Practice Cloud ID 5' values,
writing them to 'Practice Cloud ID 1' through 'Practice Cloud ID 5' columns in the Provider sheet.
It also changes the header colors to green.
"""

import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Get the project root directory (parent of Transposition Logics folder)
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
BACKEND_DIR = PROJECT_ROOT / 'backend'
EXCEL_FILES_DIR = BACKEND_DIR / 'Excel Files'

def extract_location_cloud_id_to_template():
    """
    Extract Location Cloud ID and Practice Cloud ID values from Locations_input.xlsx and write to Template copy.xlsx
    Uses NPI Number to match rows between the two files.
    
    Returns:
        bool: True if successful, False otherwise
    """
    # File paths
    locations_input_file = EXCEL_FILES_DIR / 'Locations_input.xlsx'
    template_file = EXCEL_FILES_DIR / 'Template copy.xlsx'
    
    # Check if Locations_input file exists
    if not locations_input_file.exists():
        return False
    
    # Check if template file exists
    if not template_file.exists():
        return False
    
    try:
        # Read Locations_input.xlsx (Locations sheet)
        locations_df = pd.read_excel(locations_input_file, sheet_name='Locations')
        
        # Check if 'NPI Number' column exists in Locations_input
        if 'NPI Number' not in locations_df.columns:
            return False
        
        # Find all Location Cloud ID columns (Location Cloud ID 1, 2, 3, etc.)
        location_cloud_id_columns = []
        for col in locations_df.columns:
            if col.startswith('Location Cloud ID'):
                # Extract the number if it exists (e.g., "Location Cloud ID 1" -> 1)
                if col == 'Location Cloud ID':
                    location_cloud_id_columns.append((col, 1))
                else:
                    # Extract number from "Location Cloud ID 1", "Location Cloud ID 2", etc.
                    try:
                        number = int(col.replace('Location Cloud ID', '').strip())
                        if number <= 5:  # Maximum is 5
                            location_cloud_id_columns.append((col, number))
                    except:
                        # If we can't parse the number, skip it
                        pass
        
        # Sort by number
        location_cloud_id_columns.sort(key=lambda x: x[1])
        
        # Find all Practice Cloud ID columns (Practice Cloud ID 1, 2, 3, etc.)
        practice_cloud_id_columns = []
        for col in locations_df.columns:
            if col.startswith('Practice Cloud ID'):
                # Extract the number if it exists (e.g., "Practice Cloud ID 1" -> 1)
                if col == 'Practice Cloud ID':
                    practice_cloud_id_columns.append((col, 1))
                else:
                    # Extract number from "Practice Cloud ID 1", "Practice Cloud ID 2", etc.
                    try:
                        number = int(col.replace('Practice Cloud ID', '').strip())
                        if number <= 5:  # Maximum is 5
                            practice_cloud_id_columns.append((col, number))
                    except:
                        # If we can't parse the number, skip it
                        pass
        
        # Sort by number
        practice_cloud_id_columns.sort(key=lambda x: x[1])
        
        # Find all Location Type columns (Location Type 1, 2, 3, etc.)
        location_type_columns = []
        for col in locations_df.columns:
            if col.startswith('Location Type') and col != 'Location Type':
                # Extract number from "Location Type 1", "Location Type 2", etc.
                try:
                    number = int(col.replace('Location Type', '').strip())
                    if number <= 5:  # Maximum is 5
                        location_type_columns.append((col, number))
                except:
                    # If we can't parse the number, skip it
                    pass
        
        # Sort by number
        location_type_columns.sort(key=lambda x: x[1])
        
        if not location_cloud_id_columns:
            return False
        
        # Load the template workbook
        wb = load_workbook(template_file)
        
        # Check if 'Provider' sheet exists
        if 'Provider' not in wb.sheetnames:
            return False
        
        # Check if 'Location' sheet exists
        if 'Location' not in wb.sheetnames:
            return False
        
        # Get the Provider sheet
        provider_sheet = wb['Provider']
        
        # Get the Location sheet
        location_sheet = wb['Location']
        
        # Find the 'NPI Number' column in Provider sheet
        header_row = 1
        npi_column_letter = None
        
        for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
            if cell.value and str(cell.value).strip().lower() == 'npi number':
                npi_column_letter = get_column_letter(col_idx)
                break
        
        if npi_column_letter is None:
            return False
        
        # Find Location ID columns in Provider sheet (Location ID 1, 2, 3, 4, 5)
        location_id_columns = {}
        for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
            if cell.value:
                cell_value = str(cell.value).strip().lower()
                if cell_value == 'location id 1':
                    location_id_columns[1] = get_column_letter(col_idx)
                elif cell_value == 'location id 2':
                    location_id_columns[2] = get_column_letter(col_idx)
                elif cell_value == 'location id 3':
                    location_id_columns[3] = get_column_letter(col_idx)
                elif cell_value == 'location id 4':
                    location_id_columns[4] = get_column_letter(col_idx)
                elif cell_value == 'location id 5':
                    location_id_columns[5] = get_column_letter(col_idx)
        
        # Find Practice Cloud ID column in Provider sheet (single column, not numbered)
        practice_cloud_id_column_letter = None
        for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
            if cell.value:
                cell_value = str(cell.value).strip().lower()
                if cell_value == 'practice cloud id':
                    practice_cloud_id_column_letter = get_column_letter(col_idx)
                    break
        
        if not location_id_columns:
            return False
        
        # Create mappings from NPI to Location Cloud IDs, Location Types, and Practice Cloud IDs
        # Also create mapping from Location Cloud ID to Location Type and Practice Cloud ID
        npi_to_location_cloud_ids = {}
        npi_to_practice_cloud_ids = {}
        location_cloud_id_to_location_type = {}  # Map Location Cloud ID to Location Type
        location_cloud_id_to_practice_cloud_id = {}  # Map Location Cloud ID to Practice Cloud ID
        
        for _, row in locations_df.iterrows():
            npi_value = row.get('NPI Number', '')
            if pd.notna(npi_value) and npi_value != '':
                # Normalize NPI (convert to string, remove .0 if present)
                npi_str = str(npi_value).strip()
                if npi_str.endswith('.0'):
                    npi_str = npi_str[:-2]
                
                # Get the main 'Location Type' value for this row (In Person, Virtual, or Both)
                main_location_type = row.get('Location Type', '')
                if pd.notna(main_location_type) and main_location_type != '':
                    main_location_type = str(main_location_type).strip()
                else:
                    main_location_type = None
                
                # Collect Location Cloud ID values for this NPI from this row
                # Filter based on main Location Type and Location Type n
                for col_name, col_number in location_cloud_id_columns:
                    cloud_id_value = row.get(col_name, '')
                    if pd.notna(cloud_id_value) and cloud_id_value != '':
                        cloud_id_str = str(cloud_id_value).strip()
                        
                        # Get corresponding Location Type n for this Location Cloud ID n
                        location_type_n_value = None
                        for loc_type_col, loc_type_num in location_type_columns:
                            if loc_type_num == col_number:
                                location_type_n_value = row.get(loc_type_col, '')
                                if pd.notna(location_type_n_value) and location_type_n_value != '':
                                    location_type_n_value = str(location_type_n_value).strip()
                                else:
                                    location_type_n_value = None
                                break
                        
                        # Filter based on main Location Type
                        should_include = False
                        if main_location_type:
                            main_location_type_lower = main_location_type.lower()
                            if main_location_type_lower == 'both':
                                # Include all Location Cloud IDs (both In Person and Virtual)
                                should_include = True
                            elif main_location_type_lower == 'in person':
                                # Only include if Location Type n is 'In Person'
                                if location_type_n_value and location_type_n_value.lower() == 'in person':
                                    should_include = True
                            elif main_location_type_lower == 'virtual':
                                # Only include if Location Type n is 'Virtual'
                                if location_type_n_value and location_type_n_value.lower() == 'virtual':
                                    should_include = True
                        else:
                            # If main Location Type is empty, include all (default behavior)
                            should_include = True
                        
                        # Only add if it passes the filter
                        if should_include:
                            # Store in mapping (if multiple rows have same NPI, we'll combine them)
                            if npi_str not in npi_to_location_cloud_ids:
                                npi_to_location_cloud_ids[npi_str] = []
                            # Add the Location Cloud ID (avoid duplicates)
                            if cloud_id_str not in [x[1] for x in npi_to_location_cloud_ids[npi_str]]:
                                npi_to_location_cloud_ids[npi_str].append((col_number, cloud_id_str))
                                
                                # Find Practice Cloud ID n column with same number
                                practice_cloud_id_value = None
                                for prac_cloud_id_col, prac_cloud_id_num in practice_cloud_id_columns:
                                    if prac_cloud_id_num == col_number:
                                        practice_cloud_id_value = row.get(prac_cloud_id_col, '')
                                        if pd.notna(practice_cloud_id_value) and practice_cloud_id_value != '':
                                            practice_cloud_id_value = str(practice_cloud_id_value).strip()
                                        else:
                                            practice_cloud_id_value = None
                                        break
                                
                                # Store mappings (use the first value found if multiple rows have same Location Cloud ID)
                                if cloud_id_str not in location_cloud_id_to_location_type:
                                    location_cloud_id_to_location_type[cloud_id_str] = location_type_n_value
                                if cloud_id_str not in location_cloud_id_to_practice_cloud_id:
                                    location_cloud_id_to_practice_cloud_id[cloud_id_str] = practice_cloud_id_value
                
                # Collect all unique Practice Cloud ID values for this NPI from this row
                # Process columns in order (1, 2, 3, etc.)
                for col_name, col_number in practice_cloud_id_columns:
                    practice_cloud_id_value = row.get(col_name, '')
                    if pd.notna(practice_cloud_id_value) and practice_cloud_id_value != '':
                        practice_cloud_id_str = str(practice_cloud_id_value).strip()
                        # Store in mapping (if multiple rows have same NPI, we'll combine them)
                        if npi_str not in npi_to_practice_cloud_ids:
                            npi_to_practice_cloud_ids[npi_str] = set()  # Use set to automatically handle uniqueness
                        # Add the Practice Cloud ID (set automatically handles duplicates)
                        npi_to_practice_cloud_ids[npi_str].add(practice_cloud_id_str)
        
        # Read NPI values from Provider sheet
        max_row = provider_sheet.max_row
        npi_values = []
        for row_idx in range(2, max_row + 1):  # Start from row 2 (skip header)
            npi_cell = provider_sheet[f"{npi_column_letter}{row_idx}"]
            npi_value = npi_cell.value
            if npi_value:
                # Normalize NPI (convert to string, remove .0 if present)
                npi_str = str(npi_value).strip()
                if npi_str.endswith('.0'):
                    npi_str = npi_str[:-2]
                npi_values.append((row_idx, npi_str))
            else:
                npi_values.append((row_idx, None))
        
        # Find columns in Location sheet
        location_sheet_location_cloud_id_column_letter = None
        location_sheet_location_type_column_letter = None
        location_sheet_practice_cloud_id_column_letter = None
        
        for col_idx, cell in enumerate(location_sheet[header_row], start=1):
            if cell.value:
                cell_value = str(cell.value).strip().lower()
                if cell_value == 'location cloud id':
                    location_sheet_location_cloud_id_column_letter = get_column_letter(col_idx)
                elif cell_value == 'location type':
                    location_sheet_location_type_column_letter = get_column_letter(col_idx)
                elif cell_value == 'practice cloud id':
                    location_sheet_practice_cloud_id_column_letter = get_column_letter(col_idx)
        
        # Collect all unique Location Cloud IDs in the order they appear (for Location sheet)
        all_location_cloud_ids_ordered = []  # List to preserve order
        all_location_cloud_ids_set = set()  # Set to check for duplicates quickly
        
        # Write Location Cloud ID values to Location ID columns and Practice Cloud ID values to Practice Cloud ID columns
        for row_idx, npi_str in npi_values:
            if npi_str and npi_str in npi_to_location_cloud_ids:
                location_cloud_ids = npi_to_location_cloud_ids[npi_str]
                # Sort by column number to maintain order (Location Cloud ID 1, 2, 3, etc.)
                location_cloud_ids.sort(key=lambda x: x[0])
                
                # Extract just the values (ignore source column numbers)
                cloud_id_values = [cloud_id_value for _, cloud_id_value in location_cloud_ids]
                
                # Write up to 5 Location IDs sequentially (Location ID 1, 2, 3, 4, 5)
                # Map sequentially regardless of source column number
                # e.g., Location Cloud ID 2, 3, 4 -> Location ID 1, 2, 3
                for i, cloud_id_value in enumerate(cloud_id_values[:5], start=1):
                    if i in location_id_columns:
                        cell = provider_sheet[f"{location_id_columns[i]}{row_idx}"]
                        cell.value = cloud_id_value
                        # Add to ordered list for Location sheet (preserve order of appearance)
                        if cloud_id_value and cloud_id_value not in all_location_cloud_ids_set:
                            all_location_cloud_ids_ordered.append(cloud_id_value)
                            all_location_cloud_ids_set.add(cloud_id_value)
            
            # Write Practice Cloud ID value (single column, unique values only)
            if npi_str and npi_str in npi_to_practice_cloud_ids and practice_cloud_id_column_letter:
                practice_cloud_id_set = npi_to_practice_cloud_ids[npi_str]
                # Convert set to sorted list for consistent ordering
                practice_cloud_id_values = sorted(list(practice_cloud_id_set))
                
                # If there's only one unique value, write it
                # If there are multiple unique values, write the first one
                if len(practice_cloud_id_values) > 0:
                    cell = provider_sheet[f"{practice_cloud_id_column_letter}{row_idx}"]
                    # Write the first unique value
                    cell.value = practice_cloud_id_values[0]
        
        # Color the header cells green
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        
        # Color Location ID headers
        for i in range(1, 6):
            if i in location_id_columns:
                header = provider_sheet[f"{location_id_columns[i]}{header_row}"]
                header.fill = green_fill
        
        # Color Practice Cloud ID header
        if practice_cloud_id_column_letter:
            header = provider_sheet[f"{practice_cloud_id_column_letter}{header_row}"]
            header.fill = green_fill
        
        # Write all unique Location Cloud IDs to Location sheet (avoid duplicates, preserve order)
        # Also write corresponding Location Type and Practice Cloud ID
        if location_sheet_location_cloud_id_column_letter and all_location_cloud_ids_ordered:
            # Get existing Location Cloud IDs from Location sheet to avoid duplicates
            existing_location_cloud_ids = set()
            location_sheet_max_row = location_sheet.max_row
            if location_sheet_max_row > 1:
                for row_idx in range(2, location_sheet_max_row + 1):
                    cell = location_sheet[f"{location_sheet_location_cloud_id_column_letter}{row_idx}"]
                    if cell.value:
                        existing_location_cloud_ids.add(str(cell.value).strip())
            
            # Filter out Location Cloud IDs that already exist, preserving order
            new_location_cloud_ids_ordered = [
                loc_id for loc_id in all_location_cloud_ids_ordered 
                if loc_id not in existing_location_cloud_ids
            ]
            
            if new_location_cloud_ids_ordered:
                # Find the next empty row in Location sheet
                next_row = location_sheet_max_row + 1
                
                # Check if there's an empty row before the end
                for row_idx in range(2, location_sheet_max_row + 1):
                    cell = location_sheet[f"{location_sheet_location_cloud_id_column_letter}{row_idx}"]
                    if cell.value is None or str(cell.value).strip() == '':
                        next_row = row_idx
                        break
                
                # Write new Location Cloud IDs to Location sheet in the order they appeared in Provider sheet
                # Also write corresponding Location Type and Practice Cloud ID
                for i, location_cloud_id in enumerate(new_location_cloud_ids_ordered):
                    row_idx = next_row + i
                    
                    # Write Location Cloud ID
                    cell = location_sheet[f"{location_sheet_location_cloud_id_column_letter}{row_idx}"]
                    cell.value = location_cloud_id
                    
                    # Write Location Type if column exists and we have a mapping
                    if location_sheet_location_type_column_letter and location_cloud_id in location_cloud_id_to_location_type:
                        location_type_value = location_cloud_id_to_location_type[location_cloud_id]
                        if location_type_value:
                            cell = location_sheet[f"{location_sheet_location_type_column_letter}{row_idx}"]
                            cell.value = location_type_value
                    
                    # Write Practice Cloud ID if column exists and we have a mapping
                    if location_sheet_practice_cloud_id_column_letter and location_cloud_id in location_cloud_id_to_practice_cloud_id:
                        practice_cloud_id_value = location_cloud_id_to_practice_cloud_id[location_cloud_id]
                        if practice_cloud_id_value:
                            cell = location_sheet[f"{location_sheet_practice_cloud_id_column_letter}{row_idx}"]
                            cell.value = practice_cloud_id_value
            
            # Color the headers in Location sheet green
            if location_sheet_location_cloud_id_column_letter:
                header = location_sheet[f"{location_sheet_location_cloud_id_column_letter}{header_row}"]
                header.fill = green_fill
            if location_sheet_location_type_column_letter:
                header = location_sheet[f"{location_sheet_location_type_column_letter}{header_row}"]
                header.fill = green_fill
            if location_sheet_practice_cloud_id_column_letter:
                header = location_sheet[f"{location_sheet_practice_cloud_id_column_letter}{header_row}"]
                header.fill = green_fill
        
        # Save the workbook
        wb.save(template_file)
        return True
        
    except Exception as e:
        return False

if __name__ == "__main__":
    extract_location_cloud_id_to_template()

