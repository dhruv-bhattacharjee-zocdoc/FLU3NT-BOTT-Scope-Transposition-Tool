"""
Specialty - Extracts Specialty column from _Mapped.xlsx and writes to Template copy.xlsx

This script reads the 'Specialty' column from backend\Excel Files\_Mapped.xlsx
and writes it to the 'Specialty 1', 'Specialty 2', 'Specialty 3', 'Specialty 4', and 'Specialty 5' columns
in the 'Provider' sheet in Template copy.xlsx. Values separated by '+', ';', or ',' are split across the columns.
Note: Comma (',') is a separator EXCEPT for 'Ear, Nose & Throat Doctor' and 
'Ophthalmic Plastic, Orbital & Reconstructive Surgeon' where commas are preserved.
Slash ('/') is part of specialty names.
It also changes the header colors to green.
"""

import os
import re
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Get the project root directory (parent of Transposition Logics folder)
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
BACKEND_DIR = PROJECT_ROOT / 'backend'
EXCEL_FILES_DIR = BACKEND_DIR / 'Excel Files'

def normalize_specialty(specialty_str):
    """
    Normalize a specialty string by removing punctuation, spaces, and converting to uppercase.
    Example: "Cardiology" -> "CARDIOLOGY", "Cardio logy" -> "CARDIOLOGY"
    """
    if not specialty_str:
        return ""
    # Remove all punctuation and spaces, convert to uppercase
    normalized = re.sub(r'[^\w]', '', str(specialty_str).upper())
    return normalized

def find_close_match(value, valid_specialties):
    """
    Find a close match for a value in the valid specialties list.
    Returns the original valid specialty if a close match is found, None otherwise.
    Checks for:
    1. Exact match (case-insensitive) - no replacement needed
    2. Normalized match (without punctuation/spaces) - e.g., "Cardio logy" matches "Cardiology"
    3. Prefix match - e.g., "ABC" matches "ABCd"
    """
    if not value or not valid_specialties:
        return None
    
    normalized_value = normalize_specialty(value)
    
    # First check exact match (case-insensitive)
    for valid_specialty in valid_specialties:
        if str(value).strip().lower() == str(valid_specialty).strip().lower():
            return None  # Exact match, no need to replace
    
    # Check normalized match (without punctuation/spaces)
    for valid_specialty in valid_specialties:
        normalized_valid = normalize_specialty(valid_specialty)
        if normalized_value == normalized_valid:
            return valid_specialty  # Close match found, return the valid version
    
    # Check if normalized input is a prefix of any normalized valid specialty
    # Only match if the input is at least 2 characters to avoid too broad matches
    if len(normalized_value) >= 2:
        for valid_specialty in valid_specialties:
            normalized_valid = normalize_specialty(valid_specialty)
            # Check if input is a prefix of valid specialty
            if normalized_valid.startswith(normalized_value):
                return valid_specialty  # Prefix match found, return the valid version
    
    return None

def extract_specialty_to_template():
    """
    Extract Specialty column from _Mapped.xlsx and write to Template copy.xlsx
    Values separated by '+', ';', or ',' are split across Specialty 1, 2, 3, 4, and 5 columns.
    Note: Comma (',') is a separator EXCEPT for 'Ear, Nose & Throat Doctor' and 
    'Ophthalmic Plastic, Orbital & Reconstructive Surgeon' where commas are preserved.
    
    Returns:
        bool: True if successful, False otherwise
    """
    # File paths
    source_file = EXCEL_FILES_DIR / '_Mapped.xlsx'
    npi_extracts_file = EXCEL_FILES_DIR / 'NPI-Extracts.xlsx'
    template_file = EXCEL_FILES_DIR / 'Template copy.xlsx'
    
    # Check if source file exists
    if not source_file.exists():
        return False
    
    # Check if template file exists
    if not template_file.exists():
        return False
    
    try:
        # Read the Specialty column from _Mapped.xlsx
        mapped_df = pd.read_excel(source_file)
        
        # Check if 'Specialty' column exists
        if 'Specialty' not in mapped_df.columns:
            return False
        
        # Extract the Specialty column
        specialty_data = mapped_df['Specialty'].tolist()
        
        # Read Specialty Derived column from NPI-Extracts.xlsx as fallback
        specialty_derived_data = []
        if npi_extracts_file.exists():
            try:
                npi_extracts_df = pd.read_excel(npi_extracts_file)
                if 'Specialty Derived' in npi_extracts_df.columns:
                    specialty_derived_data = npi_extracts_df['Specialty Derived'].tolist()
            except:
                pass  # If file doesn't exist or column not found, continue without it
        
        # Load the template workbook
        wb = load_workbook(template_file)
        
        # Check if 'Provider' sheet exists
        if 'Provider' not in wb.sheetnames:
            return False
        
        # Get the Provider sheet
        provider_sheet = wb['Provider']
        
        # Find the Specialty column headers
        header_row = 1
        specialty_columns = {}
        
        # Search for the headers in the first row
        for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
            if cell.value:
                cell_value = str(cell.value).strip().lower()
                if cell_value == 'specialty 1':
                    specialty_columns[1] = get_column_letter(col_idx)
                elif cell_value == 'specialty 2':
                    specialty_columns[2] = get_column_letter(col_idx)
                elif cell_value == 'specialty 3':
                    specialty_columns[3] = get_column_letter(col_idx)
                elif cell_value == 'specialty 4':
                    specialty_columns[4] = get_column_letter(col_idx)
                elif cell_value == 'specialty 5':
                    specialty_columns[5] = get_column_letter(col_idx)
        
        if 1 not in specialty_columns:
            return False
        
        # Load valid specialties from ValidationAndReference sheet
        valid_specialties = set()
        if 'ValidationAndReference' in wb.sheetnames:
            validation_sheet = wb['ValidationAndReference']
            specialty_column_letter = None
            
            # Find the 'Specialty Name' column in ValidationAndReference sheet
            for col_idx, cell in enumerate(validation_sheet[header_row], start=1):
                if cell.value and str(cell.value).strip().lower() == 'specialty name':
                    specialty_column_letter = get_column_letter(col_idx)
                    break
            
            # Read all valid specialty values
            if specialty_column_letter:
                max_row = validation_sheet.max_row
                for row_idx in range(2, max_row + 1):  # Start from row 2 (skip header)
                    cell = validation_sheet[f"{specialty_column_letter}{row_idx}"]
                    if cell.value:
                        valid_specialties.add(str(cell.value).strip())
        
        # Grey fill for invalid values
        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Process and write the Specialty data
        for row_idx, specialty_value in enumerate(specialty_data, start=2):
            # Check if Specialty value is empty, if so use Specialty Derived as fallback
            data_index = row_idx - 2  # Convert to 0-based index for accessing specialty_derived_data
            # Helper function to split specialty string by '+', ';', or ',' (with exceptions)
            def split_specialty_string(specialty_str):
                """
                Split specialty string by '+', ';', or ',' 
                Exceptions: Preserve commas in 'Ear, Nose & Throat Doctor' and 
                'Ophthalmic Plastic, Orbital & Reconstructive Surgeon'
                """
                if not specialty_str:
                    return [None, None, None, None, None]
                specialty_str = str(specialty_str).strip()
                
                # Exception phrases that should preserve commas
                exception_phrases = [
                    'Ear, Nose & Throat Doctor',
                    'Ophthalmic Plastic, Orbital & Reconstructive Surgeon'
                ]
                
                # Replace exception phrases with placeholders to protect commas
                placeholder_map = {}
                protected_str = specialty_str
                for i, phrase in enumerate(exception_phrases):
                    placeholder = f'__EXCEPTION_{i}__'
                    if phrase in protected_str:
                        protected_str = protected_str.replace(phrase, placeholder)
                        placeholder_map[placeholder] = phrase
                
                # Now split by separators
                parts = []
                if '+' in protected_str:
                    parts = [part.strip() if part.strip() else None for part in protected_str.split('+')]
                elif ';' in protected_str:
                    parts = [part.strip() if part.strip() else None for part in protected_str.split(';')]
                elif ',' in protected_str:
                    parts = [part.strip() if part.strip() else None for part in protected_str.split(',')]
                else:
                    # No separator found, treat entire string as one specialty
                    parts = [protected_str] if protected_str else [None]
                
                # Restore exception phrases from placeholders
                restored_parts = []
                for part in parts:
                    if part:
                        # Check if this part contains a placeholder
                        for placeholder, original_phrase in placeholder_map.items():
                            if placeholder in part:
                                part = part.replace(placeholder, original_phrase)
                        restored_parts.append(part)
                    else:
                        restored_parts.append(None)
                
                # Take only first 5 parts and pad with None if needed
                specialty_parts = restored_parts[:5]
                while len(specialty_parts) < 5:
                    specialty_parts.append(None)
                return specialty_parts
            
            # Track if we're using Specialty Derived or if values are split
            using_specialty_derived = False
            has_multiple_values = False
            
            if pd.isna(specialty_value) or (isinstance(specialty_value, str) and not specialty_value.strip()):
                # Use Specialty Derived if available
                if specialty_derived_data and data_index < len(specialty_derived_data):
                    specialty_derived_value = specialty_derived_data[data_index]
                    if not pd.isna(specialty_derived_value) and specialty_derived_value:
                        # Split Specialty Derived value (can contain '+' or ';' separators)
                        specialty_parts = split_specialty_string(specialty_derived_value)
                        using_specialty_derived = True
                        # Check if there are multiple values
                        non_empty_count = sum(1 for p in specialty_parts if p is not None)
                        has_multiple_values = non_empty_count > 1
                    else:
                        specialty_parts = [None, None, None, None, None]
                else:
                    specialty_parts = [None, None, None, None, None]
            else:
                specialty_str = str(specialty_value).strip()
                if not specialty_str:
                    # Check Specialty Derived as fallback
                    if specialty_derived_data and data_index < len(specialty_derived_data):
                        specialty_derived_value = specialty_derived_data[data_index]
                        if not pd.isna(specialty_derived_value) and specialty_derived_value:
                            # Split Specialty Derived value (can contain '+' or ';' separators)
                            specialty_parts = split_specialty_string(specialty_derived_value)
                            using_specialty_derived = True
                            # Check if there are multiple values
                            non_empty_count = sum(1 for p in specialty_parts if p is not None)
                            has_multiple_values = non_empty_count > 1
                        else:
                            specialty_parts = [None, None, None, None, None]
                    else:
                        specialty_parts = [None, None, None, None, None]
                else:
                    # Split by '+', ';', or ',' (with exceptions for certain specialty names)
                    specialty_parts = split_specialty_string(specialty_str)
                    # Check if there are multiple values
                    non_empty_count = sum(1 for p in specialty_parts if p is not None)
                    has_multiple_values = non_empty_count > 1
            
            # Write to Specialty 1 through 5
            for i in range(1, 6):
                if i in specialty_columns:
                    cell = provider_sheet[f"{specialty_columns[i]}{row_idx}"]
                    value = specialty_parts[i-1] if specialty_parts[i-1] else None
                    
                    # Check for close match and replace if found
                    if value and valid_specialties:
                        close_match = find_close_match(value, valid_specialties)
                        if close_match:
                            cell.value = close_match
                            cell.fill = grey_fill  # Highlight corrected values
                        elif value not in valid_specialties:
                            cell.value = value
                            cell.fill = grey_fill  # Highlight invalid values
                        else:
                            cell.value = value
                            # Highlight if using Specialty Derived or has multiple values
                            if using_specialty_derived or has_multiple_values:
                                cell.fill = grey_fill
                    else:
                        cell.value = value
                        # Highlight if using Specialty Derived or has multiple values
                        if using_specialty_derived or has_multiple_values:
                            cell.fill = grey_fill
        
        # Color the header cells green
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        
        for i in range(1, 6):
            if i in specialty_columns:
                header = provider_sheet[f"{specialty_columns[i]}{header_row}"]
                header.fill = green_fill
        
        # Save the workbook
        wb.save(template_file)
        return True
        
    except Exception as e:
        return False

if __name__ == "__main__":
    extract_specialty_to_template()

