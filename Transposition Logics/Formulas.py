"""
Formulas - Applies Excel formulas to specific columns in Template copy.xlsx

This script applies formulas to columns in the Provider and Location sheets
of Template copy.xlsx. Formulas are organized by section (Provider and Locations).
"""

import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Get the project root directory (parent of Transposition Logics folder)
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
BACKEND_DIR = PROJECT_ROOT / 'backend'
EXCEL_FILES_DIR = BACKEND_DIR / 'Excel Files'

def apply_formulas_to_template():
    """
    Apply formulas to columns in Template copy.xlsx
    
    Returns:
        bool: True if successful, False otherwise
    """
    template_file = EXCEL_FILES_DIR / 'Template copy.xlsx'
    
    # Check if template file exists
    if not template_file.exists():
        return False
    
    try:
        # Load the template workbook
        wb = load_workbook(template_file)
        
        # ========== PROVIDER SECTION ==========
        if 'Provider' in wb.sheetnames:
            provider_sheet = wb['Provider']
            header_row = 1
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            
            # Find Language ID columns in Provider sheet
            language_id_columns = {}
            for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip()
                    if cell_value == 'Language ID 1':
                        language_id_columns[1] = get_column_letter(col_idx)
                    elif cell_value == 'Language ID 2':
                        language_id_columns[2] = get_column_letter(col_idx)
                    elif cell_value == 'Language ID 3':
                        language_id_columns[3] = get_column_letter(col_idx)
            
            # Find Additional Language Spoken columns to determine reference columns (AZ, BA, BB)
            # These are the columns that contain the language names we're matching
            additional_language_columns = {}
            for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip().lower()
                    # Check for both singular and plural variations
                    if cell_value == 'additional language spoken 1' or cell_value == 'additional languages spoken 1':
                        additional_language_columns[1] = get_column_letter(col_idx)
                    elif cell_value == 'additional language spoken 2' or cell_value == 'additional languages spoken 2':
                        additional_language_columns[2] = get_column_letter(col_idx)
                    elif cell_value == 'additional language spoken 3' or cell_value == 'additional languages spoken 3':
                        additional_language_columns[3] = get_column_letter(col_idx)
            
            # Apply formulas to Language ID columns
            # Formula: IFERROR(INDEX(ValidationAndReference!V:V, MATCH(reference_cell, ValidationAndReference!W:W, 0)), "")
            max_row = provider_sheet.max_row
            
            for lang_id_num in [1, 2, 3]:
                if lang_id_num in language_id_columns and lang_id_num in additional_language_columns:
                    lang_id_col = language_id_columns[lang_id_num]
                    ref_col = additional_language_columns[lang_id_num]
                    
                    # Apply formula to all data rows (starting from row 2)
                    for row_idx in range(2, max_row + 1):
                        cell = provider_sheet[f"{lang_id_col}{row_idx}"]
                        # Formula: IFERROR(INDEX(ValidationAndReference!V:V, MATCH(ref_cell, ValidationAndReference!W:W, 0)), "")
                        formula = f'=IFERROR(INDEX(ValidationAndReference!V:V, MATCH({ref_col}{row_idx}, ValidationAndReference!W:W, 0)), "")'
                        cell.value = formula
                    
                    # Color the Language ID column header with green (#00FF00)
                    header_cell = provider_sheet[f"{lang_id_col}{header_row}"]
                    header_cell.fill = green_fill
            
            # Find Provider Type (Substatus) ID column and Provider Type reference column
            provider_type_substatus_id_column_letter = None
            provider_type_ref_column_letter = None
            
            for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip()
                    if cell_value == 'Provider Type (Substatus) ID':
                        provider_type_substatus_id_column_letter = get_column_letter(col_idx)
                    elif cell_value == 'Provider Type':
                        provider_type_ref_column_letter = get_column_letter(col_idx)
            
            # Apply formula to Provider Type (Substatus) ID
            # Formula: IFERROR(INDEX(ValidationAndReference!P:P, MATCH(ref_cell, ValidationAndReference!Q:Q, 0)), "")
            if provider_type_substatus_id_column_letter and provider_type_ref_column_letter:
                for row_idx in range(2, max_row + 1):
                    cell = provider_sheet[f"{provider_type_substatus_id_column_letter}{row_idx}"]
                    formula = f'=IFERROR(INDEX(ValidationAndReference!P:P, MATCH({provider_type_ref_column_letter}{row_idx}, ValidationAndReference!Q:Q, 0)), "")'
                    cell.value = formula
                
                # Color the Provider Type (Substatus) ID column header with green (#00FF00)
                header_cell = provider_sheet[f"{provider_type_substatus_id_column_letter}{header_row}"]
                header_cell.fill = green_fill
            
            # Find Professional Suffix ID columns and Professional Suffix reference columns
            professional_suffix_id_columns = {}
            professional_suffix_ref_columns = {}
            
            for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip()
                    if cell_value == 'Professional Suffix ID 1':
                        professional_suffix_id_columns[1] = get_column_letter(col_idx)
                    elif cell_value == 'Professional Suffix ID 2':
                        professional_suffix_id_columns[2] = get_column_letter(col_idx)
                    elif cell_value == 'Professional Suffix ID 3':
                        professional_suffix_id_columns[3] = get_column_letter(col_idx)
                    elif cell_value == 'Professional Suffix 1':
                        professional_suffix_ref_columns[1] = get_column_letter(col_idx)
                    elif cell_value == 'Professional Suffix 2':
                        professional_suffix_ref_columns[2] = get_column_letter(col_idx)
                    elif cell_value == 'Professional Suffix 3':
                        professional_suffix_ref_columns[3] = get_column_letter(col_idx)
            
            # Apply formulas to Professional Suffix ID columns
            # Formula: IFERROR(INDEX(ValidationAndReference!F:F, MATCH(ref_cell, ValidationAndReference!G:G, 0)), "")
            for suffix_id_num in [1, 2, 3]:
                if suffix_id_num in professional_suffix_id_columns and suffix_id_num in professional_suffix_ref_columns:
                    suffix_id_col = professional_suffix_id_columns[suffix_id_num]
                    ref_col = professional_suffix_ref_columns[suffix_id_num]
                    
                    # Apply formula to all data rows (starting from row 2)
                    for row_idx in range(2, max_row + 1):
                        cell = provider_sheet[f"{suffix_id_col}{row_idx}"]
                        formula = f'=IFERROR(INDEX(ValidationAndReference!F:F, MATCH({ref_col}{row_idx}, ValidationAndReference!G:G, 0)), "")'
                        cell.value = formula
                    
                    # Color the Professional Suffix ID column header with green (#00FF00)
                    header_cell = provider_sheet[f"{suffix_id_col}{header_row}"]
                    header_cell.fill = green_fill
            
            # Find Specialty ID columns and Specialty reference columns
            specialty_id_columns = {}
            specialty_ref_columns = {}
            
            for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip()
                    if cell_value == 'Specialty ID 1':
                        specialty_id_columns[1] = get_column_letter(col_idx)
                    elif cell_value == 'Specialty ID 2':
                        specialty_id_columns[2] = get_column_letter(col_idx)
                    elif cell_value == 'Specialty ID 3':
                        specialty_id_columns[3] = get_column_letter(col_idx)
                    elif cell_value == 'Specialty ID 4':
                        specialty_id_columns[4] = get_column_letter(col_idx)
                    elif cell_value == 'Specialty ID 5':
                        specialty_id_columns[5] = get_column_letter(col_idx)
                    elif cell_value == 'Specialty 1':
                        specialty_ref_columns[1] = get_column_letter(col_idx)
                    elif cell_value == 'Specialty 2':
                        specialty_ref_columns[2] = get_column_letter(col_idx)
                    elif cell_value == 'Specialty 3':
                        specialty_ref_columns[3] = get_column_letter(col_idx)
                    elif cell_value == 'Specialty 4':
                        specialty_ref_columns[4] = get_column_letter(col_idx)
                    elif cell_value == 'Specialty 5':
                        specialty_ref_columns[5] = get_column_letter(col_idx)
            
            # Apply formulas to Specialty ID columns
            # Formula: IF(ISBLANK(ref_cell),"",INDEX(ValidationAndReference!J:J,MATCH(ref_cell,ValidationAndReference!K:K,0)))
            for specialty_id_num in [1, 2, 3, 4, 5]:
                if specialty_id_num in specialty_id_columns and specialty_id_num in specialty_ref_columns:
                    specialty_id_col = specialty_id_columns[specialty_id_num]
                    ref_col = specialty_ref_columns[specialty_id_num]
                    
                    # Apply formula to all data rows (starting from row 2)
                    for row_idx in range(2, max_row + 1):
                        cell = provider_sheet[f"{specialty_id_col}{row_idx}"]
                        formula = f'=IF(ISBLANK({ref_col}{row_idx}),"",INDEX(ValidationAndReference!J:J,MATCH({ref_col}{row_idx},ValidationAndReference!K:K,0)))'
                        cell.value = formula
                    
                    # Color the Specialty ID column header with green (#00FF00)
                    header_cell = provider_sheet[f"{specialty_id_col}{header_row}"]
                    header_cell.fill = green_fill
            
            # Find Board Cert ID columns and Board Certification reference columns
            board_cert_id_columns = {}
            board_cert_ref_columns = {}
            
            for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip()
                    if cell_value == 'Board Cert ID 1':
                        board_cert_id_columns[1] = get_column_letter(col_idx)
                    elif cell_value == 'Board Cert ID 2':
                        board_cert_id_columns[2] = get_column_letter(col_idx)
                    elif cell_value == 'Board Cert ID 3':
                        board_cert_id_columns[3] = get_column_letter(col_idx)
                    elif cell_value == 'Board Cert ID 4':
                        board_cert_id_columns[4] = get_column_letter(col_idx)
                    elif cell_value == 'Board Cert ID 5':
                        board_cert_id_columns[5] = get_column_letter(col_idx)
                    elif cell_value == 'Board Certification 1':
                        board_cert_ref_columns[1] = get_column_letter(col_idx)
                    elif cell_value == 'Board Certification 2':
                        board_cert_ref_columns[2] = get_column_letter(col_idx)
                    elif cell_value == 'Board Certification 3':
                        board_cert_ref_columns[3] = get_column_letter(col_idx)
                    elif cell_value == 'Board Certification 4':
                        board_cert_ref_columns[4] = get_column_letter(col_idx)
                    elif cell_value == 'Board Certification 5':
                        board_cert_ref_columns[5] = get_column_letter(col_idx)
            
            # Apply formulas to Board Cert ID columns
            # Formula: IF(ISBLANK(ref_cell),"",INDEX(ValidationAndReference!$AA:$AA,MATCH(ref_cell,ValidationAndReference!$AB:$AB,0)))
            for cert_id_num in [1, 2, 3, 4, 5]:
                if cert_id_num in board_cert_id_columns and cert_id_num in board_cert_ref_columns:
                    cert_id_col = board_cert_id_columns[cert_id_num]
                    ref_col = board_cert_ref_columns[cert_id_num]
                    
                    # Apply formula to all data rows (starting from row 2)
                    for row_idx in range(2, max_row + 1):
                        cell = provider_sheet[f"{cert_id_col}{row_idx}"]
                        formula = f'=IF(ISBLANK({ref_col}{row_idx}),"",INDEX(ValidationAndReference!$AA:$AA,MATCH({ref_col}{row_idx},ValidationAndReference!$AB:$AB,0)))'
                        cell.value = formula
            
            # Find Sub Board Cert ID columns and Sub Board Certification reference columns
            sub_board_cert_id_columns = {}
            sub_board_cert_ref_columns = {}
            
            for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip()
                    if cell_value == 'Sub Board Cert ID 1':
                        sub_board_cert_id_columns[1] = get_column_letter(col_idx)
                    elif cell_value == 'Sub Board Cert ID 2':
                        sub_board_cert_id_columns[2] = get_column_letter(col_idx)
                    elif cell_value == 'Sub Board Cert ID 3':
                        sub_board_cert_id_columns[3] = get_column_letter(col_idx)
                    elif cell_value == 'Sub Board Cert ID 4':
                        sub_board_cert_id_columns[4] = get_column_letter(col_idx)
                    elif cell_value == 'Sub Board Cert ID 5':
                        sub_board_cert_id_columns[5] = get_column_letter(col_idx)
                    elif cell_value == 'Sub Board Certification 1':
                        sub_board_cert_ref_columns[1] = get_column_letter(col_idx)
                    elif cell_value == 'Sub Board Certification 2':
                        sub_board_cert_ref_columns[2] = get_column_letter(col_idx)
                    elif cell_value == 'Sub Board Certification 3':
                        sub_board_cert_ref_columns[3] = get_column_letter(col_idx)
                    elif cell_value == 'Sub Board Certification 4':
                        sub_board_cert_ref_columns[4] = get_column_letter(col_idx)
                    elif cell_value == 'Sub Board Certification 5':
                        sub_board_cert_ref_columns[5] = get_column_letter(col_idx)
            
            # Apply formulas to Sub Board Cert ID columns
            # Formula: IF(ISBLANK(ref_cell),"",INDEX(ValidationAndReference!$M:$M,MATCH(ref_cell,ValidationAndReference!$N:$N,0)))
            for sub_cert_id_num in [1, 2, 3, 4, 5]:
                if sub_cert_id_num in sub_board_cert_id_columns and sub_cert_id_num in sub_board_cert_ref_columns:
                    sub_cert_id_col = sub_board_cert_id_columns[sub_cert_id_num]
                    ref_col = sub_board_cert_ref_columns[sub_cert_id_num]
                    
                    # Apply formula to all data rows (starting from row 2)
                    for row_idx in range(2, max_row + 1):
                        cell = provider_sheet[f"{sub_cert_id_col}{row_idx}"]
                        formula = f'=IF(ISBLANK({ref_col}{row_idx}),"",INDEX(ValidationAndReference!$M:$M,MATCH({ref_col}{row_idx},ValidationAndReference!$N:$N,0)))'
                        cell.value = formula
            
            # Find Hospital Affiliation ID columns and Hospital Affiliation reference columns
            hospital_affiliation_id_columns = {}
            hospital_affiliation_ref_columns = {}
            
            for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip()
                    if cell_value == 'Hospital Affiliation ID 1':
                        hospital_affiliation_id_columns[1] = get_column_letter(col_idx)
                    elif cell_value == 'Hospital Affiliation ID 2':
                        hospital_affiliation_id_columns[2] = get_column_letter(col_idx)
                    elif cell_value == 'Hospital Affiliation ID 3':
                        hospital_affiliation_id_columns[3] = get_column_letter(col_idx)
                    elif cell_value == 'Hospital Affiliation ID 4':
                        hospital_affiliation_id_columns[4] = get_column_letter(col_idx)
                    elif cell_value == 'Hospital Affiliation ID 5':
                        hospital_affiliation_id_columns[5] = get_column_letter(col_idx)
                    elif cell_value == 'Hospital Affiliation 1':
                        hospital_affiliation_ref_columns[1] = get_column_letter(col_idx)
                    elif cell_value == 'Hospital Affiliation 2':
                        hospital_affiliation_ref_columns[2] = get_column_letter(col_idx)
                    elif cell_value == 'Hospital Affiliation 3':
                        hospital_affiliation_ref_columns[3] = get_column_letter(col_idx)
                    elif cell_value == 'Hospital Affiliation 4':
                        hospital_affiliation_ref_columns[4] = get_column_letter(col_idx)
                    elif cell_value == 'Hospital Affiliation 5':
                        hospital_affiliation_ref_columns[5] = get_column_letter(col_idx)
            
            # Apply formulas to Hospital Affiliation ID columns
            # Formula: IF(ISBLANK(ref_cell),"",INDEX(ValidationAndReference!$S:$S,MATCH(ref_cell,ValidationAndReference!$T:$T,0)))
            for hosp_id_num in [1, 2, 3, 4, 5]:
                if hosp_id_num in hospital_affiliation_id_columns and hosp_id_num in hospital_affiliation_ref_columns:
                    hosp_id_col = hospital_affiliation_id_columns[hosp_id_num]
                    ref_col = hospital_affiliation_ref_columns[hosp_id_num]
                    
                    # Apply formula to all data rows (starting from row 2)
                    for row_idx in range(2, max_row + 1):
                        cell = provider_sheet[f"{hosp_id_col}{row_idx}"]
                        formula = f'=IF(ISBLANK({ref_col}{row_idx}),"",INDEX(ValidationAndReference!$S:$S,MATCH({ref_col}{row_idx},ValidationAndReference!$T:$T,0)))'
                        cell.value = formula
            
            # Find Location 1-5 columns and Location ID 1-5 reference columns
            location_columns = {}
            location_id_ref_columns = {}
            
            for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip()
                    if cell_value == 'Location 1':
                        location_columns[1] = get_column_letter(col_idx)
                    elif cell_value == 'Location 2':
                        location_columns[2] = get_column_letter(col_idx)
                    elif cell_value == 'Location 3':
                        location_columns[3] = get_column_letter(col_idx)
                    elif cell_value == 'Location 4':
                        location_columns[4] = get_column_letter(col_idx)
                    elif cell_value == 'Location 5':
                        location_columns[5] = get_column_letter(col_idx)
                    elif cell_value == 'Location ID 1':
                        location_id_ref_columns[1] = get_column_letter(col_idx)
                    elif cell_value == 'Location ID 2':
                        location_id_ref_columns[2] = get_column_letter(col_idx)
                    elif cell_value == 'Location ID 3':
                        location_id_ref_columns[3] = get_column_letter(col_idx)
                    elif cell_value == 'Location ID 4':
                        location_id_ref_columns[4] = get_column_letter(col_idx)
                    elif cell_value == 'Location ID 5':
                        location_id_ref_columns[5] = get_column_letter(col_idx)
            
            # Apply formulas to Location 1-5 columns
            # Formula: IFERROR(INDEX(Location!X:X, MATCH(ref_cell, Location!W:W, 0)), "")
            for location_num in [1, 2, 3, 4, 5]:
                if location_num in location_columns and location_num in location_id_ref_columns:
                    location_col = location_columns[location_num]
                    ref_col = location_id_ref_columns[location_num]
                    
                    # Apply formula to all data rows (starting from row 2)
                    for row_idx in range(2, max_row + 1):
                        cell = provider_sheet[f"{location_col}{row_idx}"]
                        formula = f'=IFERROR(INDEX(Location!X:X, MATCH({ref_col}{row_idx}, Location!W:W, 0)), "")'
                        cell.value = formula
                    
                    # Color the Location column header with green (#00FF00)
                    header_cell = provider_sheet[f"{location_col}{header_row}"]
                    header_cell.fill = green_fill
        
        # ========== LOCATIONS SECTION ==========
        if 'Location' in wb.sheetnames:
            location_sheet = wb['Location']
            header_row = 1
            max_row = location_sheet.max_row
            
            # Find Scheduling Software ID column and Scheduling Software reference column
            scheduling_software_id_column_letter = None
            scheduling_software_ref_column_letter = None
            
            for col_idx, cell in enumerate(location_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip().lower()
                    if cell_value == 'scheduling software id':
                        scheduling_software_id_column_letter = get_column_letter(col_idx)
                    elif cell_value == 'scheduling software':
                        scheduling_software_ref_column_letter = get_column_letter(col_idx)
            
            # Apply formula to Scheduling Software ID
            # Formula: IF(ISBLANK(ref_cell),"",INDEX(ValidationAndReference!C:C,MATCH(ref_cell,ValidationAndReference!D:D,0)))
            if scheduling_software_id_column_letter and scheduling_software_ref_column_letter:
                for row_idx in range(2, max_row + 1):
                    cell = location_sheet[f"{scheduling_software_id_column_letter}{row_idx}"]
                    formula = f'=IF(ISBLANK({scheduling_software_ref_column_letter}{row_idx}),"",INDEX(ValidationAndReference!C:C,MATCH({scheduling_software_ref_column_letter}{row_idx},ValidationAndReference!D:D,0)))'
                    cell.value = formula
                
                # Color the Scheduling Software ID column header with green (#00FF00)
                header_cell = location_sheet[f"{scheduling_software_id_column_letter}{header_row}"]
                green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                header_cell.fill = green_fill
            
            # Find Combined address column and reference columns
            combined_address_column_letter = None
            address_line1_column_letter = None
            address_line2_column_letter = None
            city_column_letter = None
            state_column_letter = None
            zip_code_column_letter = None
            
            for col_idx, cell in enumerate(location_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip()
                    if cell_value == 'Combined address':
                        combined_address_column_letter = get_column_letter(col_idx)
                    elif cell_value == 'Address line 1':
                        address_line1_column_letter = get_column_letter(col_idx)
                    elif cell_value == 'Address line 2 (Office/Suite #)':
                        address_line2_column_letter = get_column_letter(col_idx)
                    elif cell_value == 'City':
                        city_column_letter = get_column_letter(col_idx)
                    elif cell_value == 'State':
                        state_column_letter = get_column_letter(col_idx)
                    elif cell_value == 'ZIP Code':
                        zip_code_column_letter = get_column_letter(col_idx)
            
            # Apply formula to Combined address
            # Formula: Concatenate Address line 1, Address line 2 (if not empty), City, State, ZIP Code separated by commas
            if (combined_address_column_letter and address_line1_column_letter and 
                city_column_letter and state_column_letter and zip_code_column_letter):
                for row_idx in range(2, max_row + 1):
                    cell = location_sheet[f"{combined_address_column_letter}{row_idx}"]
                    # Build formula: Address line 1, (Address line 2 if not empty), City, State, ZIP Code
                    # Format: A2, (B2 if not empty), C2, D2, E2
                    if address_line2_column_letter:
                        # Include Address line 2 only if not empty
                        formula = f'={address_line1_column_letter}{row_idx}&IF({address_line2_column_letter}{row_idx}<>"",", "&{address_line2_column_letter}{row_idx},"")&", "&{city_column_letter}{row_idx}&", "&{state_column_letter}{row_idx}&", "&{zip_code_column_letter}{row_idx}'
                    else:
                        # If Address line 2 column doesn't exist, skip it
                        formula = f'={address_line1_column_letter}{row_idx}&", "&{city_column_letter}{row_idx}&", "&{state_column_letter}{row_idx}&", "&{zip_code_column_letter}{row_idx}'
                    cell.value = formula
                
                # Color the Combined address column header with green (#00FF00)
                header_cell = location_sheet[f"{combined_address_column_letter}{header_row}"]
                green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                header_cell.fill = green_fill
            
            # Find Complete Location column and reference columns (A, B, C, D, E, F, G)
            # Based on the formula, these are likely: Address Line 1, Address Line 2, City, State, ZIP Code, etc.
            complete_location_column_letter = None
            ref_columns = {}  # Will store A, B, C, D, E, F, G column letters
            
            for col_idx, cell in enumerate(location_sheet[header_row], start=1):
                if cell.value:
                    cell_value = str(cell.value).strip()
                    if cell_value == 'Complete Location':
                        complete_location_column_letter = get_column_letter(col_idx)
                    # Find columns by their actual position (A=1, B=2, C=3, D=4, E=5, F=6, G=7)
                    # Or find by header names if we can identify them
                    col_letter = get_column_letter(col_idx)
                    if col_letter == 'A':
                        ref_columns['A'] = col_letter
                    elif col_letter == 'B':
                        ref_columns['B'] = col_letter
                    elif col_letter == 'C':
                        ref_columns['C'] = col_letter
                    elif col_letter == 'D':
                        ref_columns['D'] = col_letter
                    elif col_letter == 'E':
                        ref_columns['E'] = col_letter
                    elif col_letter == 'F':
                        ref_columns['F'] = col_letter
                    elif col_letter == 'G':
                        ref_columns['G'] = col_letter
            
            # Apply formula to Complete Location
            # Formula: IF(A2<>"",CONCATENATE(A2," ",B2," ",D2," ",E2," ",F2," ",G2," ","(",C2,")"),"")
            if complete_location_column_letter and all(key in ref_columns for key in ['A', 'B', 'C', 'D', 'E', 'F', 'G']):
                for row_idx in range(2, max_row + 1):
                    cell = location_sheet[f"{complete_location_column_letter}{row_idx}"]
                    formula = f'=IF({ref_columns["A"]}{row_idx}<>"",CONCATENATE({ref_columns["A"]}{row_idx}," ",{ref_columns["B"]}{row_idx}," ",{ref_columns["D"]}{row_idx}," ",{ref_columns["E"]}{row_idx}," ",{ref_columns["F"]}{row_idx}," ",{ref_columns["G"]}{row_idx}," ","(",{ref_columns["C"]}{row_idx},")"),"")'
                    cell.value = formula
                
                # Color the Complete Location column header with green (#00FF00)
                header_cell = location_sheet[f"{complete_location_column_letter}{header_row}"]
                green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                header_cell.fill = green_fill
        
        # Save the workbook
        wb.save(template_file)
        return True
        
    except Exception as e:
        return False

if __name__ == "__main__":
    apply_formulas_to_template()

