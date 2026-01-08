"""
Location Details - Extracts location details from Practice_Locations.xlsx and writes to Template copy.xlsx

This script reads the 'Location Cloud ID' column from the 'Location' sheet in Template copy.xlsx,
searches for matching location_id values in Practice_Locations.xlsx (Locations sheet),
and extracts associated location details (address, city, state, zip, email, phone, etc.),
writing them to corresponding columns in the Location sheet.
It also changes the header colors to green.
"""

import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Get the project root directory (parent of Transposition Logics folder)
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
BACKEND_DIR = PROJECT_ROOT / 'backend'
EXCEL_FILES_DIR = BACKEND_DIR / 'Excel Files'

def extract_location_details_to_template():
    """
    Extract location details from Practice_Locations.xlsx and write to Template copy.xlsx
    Uses Location Cloud ID to match rows between the two files.
    
    Returns:
        bool: True if successful, False otherwise
    """
    # File paths
    practice_locations_file = EXCEL_FILES_DIR / 'Practice_Locations.xlsx'
    template_file = EXCEL_FILES_DIR / 'Template copy.xlsx'
    
    # Check if Practice_Locations file exists
    if not practice_locations_file.exists():
        return False
    
    # Check if template file exists
    if not template_file.exists():
        return False
    
    try:
        # Read Practice_Locations.xlsx (Locations sheet)
        # Read ZIP column as string to preserve leading zeros
        dtype_dict = {'zip': str}
        practice_locations_df = pd.read_excel(practice_locations_file, sheet_name='Locations', dtype=dtype_dict)
        
        # Check if 'location_id' column exists
        if 'location_id' not in practice_locations_df.columns:
            return False
        
        # Create a mapping from location_id to all the details we need
        location_id_to_details = {}
        
        for _, row in practice_locations_df.iterrows():
            location_id_value = row.get('location_id', '')
            if pd.notna(location_id_value) and location_id_value != '':
                location_id_str = str(location_id_value).strip()
                
                # Extract all the details for this location_id
                details = {
                    'address_1': row.get('address_1', ''),
                    'address_2': row.get('address_2', ''),
                    'city': row.get('city', ''),
                    'state': row.get('state', ''),
                    'zip': row.get('zip', ''),
                    'email_addresses': row.get('email_addresses', ''),
                    'phone': row.get('phone', ''),
                    'virtual_visit_type': row.get('virtual_visit_type', ''),
                    'software': row.get('software', ''),
                    'practice_facing_name': row.get('practice_facing_name', '')
                }
                
                # Convert to strings and strip, handle NaN
                # Special handling for ZIP code to preserve leading zeros
                for key in details:
                    if pd.notna(details[key]) and details[key] != '':
                        if key == 'zip':
                            # For ZIP code, preserve leading zeros
                            zip_value = details[key]
                            if isinstance(zip_value, (int, float)):
                                # If it's a number, convert to string and pad with leading zeros (up to 5 digits)
                                # This handles cases where Excel converted "01234" to 1234
                                details[key] = str(int(zip_value)).zfill(5)
                            else:
                                # If it's already a string, just strip it
                                zip_str = str(zip_value).strip()
                                # If it's a numeric string without leading zeros, pad it (but this shouldn't happen if read as string)
                                if zip_str.isdigit() and len(zip_str) < 5:
                                    details[key] = zip_str.zfill(5)
                                else:
                                    details[key] = zip_str
                        else:
                            details[key] = str(details[key]).strip()
                    else:
                        details[key] = None
                
                # Store in mapping (use first value if duplicate location_id)
                if location_id_str not in location_id_to_details:
                    location_id_to_details[location_id_str] = details
        
        # Load the template workbook
        wb = load_workbook(template_file)
        
        # Check if 'Location' sheet exists
        if 'Location' not in wb.sheetnames:
            return False
        
        # Get the Location sheet
        location_sheet = wb['Location']
        
        # Find the 'Location Cloud ID' column in Location sheet
        header_row = 1
        location_cloud_id_column_letter = None
        
        for col_idx, cell in enumerate(location_sheet[header_row], start=1):
            if cell.value and str(cell.value).strip().lower() == 'location cloud id':
                location_cloud_id_column_letter = get_column_letter(col_idx)
                break
        
        if location_cloud_id_column_letter is None:
            return False
        
        # Find target columns in Location sheet
        target_columns = {
            'Address line 1': None,
            'Address line 2 (Office/Suite #)': None,
            'City': None,
            'State': None,
            'ZIP Code': None,
            'Email for appointment notifications 1': None,
            'Phone': None,
            'Virtual Visit Type': None,
            'Scheduling Software': None,
            'Location Name': None
        }
        
        for col_idx, cell in enumerate(location_sheet[header_row], start=1):
            if cell.value:
                cell_value = str(cell.value).strip()
                if cell_value in target_columns:
                    target_columns[cell_value] = get_column_letter(col_idx)
        
        # Column mapping from Practice_Locations to Template copy
        column_mapping = {
            'address_1': 'Address line 1',
            'address_2': 'Address line 2 (Office/Suite #)',
            'city': 'City',
            'state': 'State',
            'zip': 'ZIP Code',
            'email_addresses': 'Email for appointment notifications 1',
            'phone': 'Phone',
            'virtual_visit_type': 'Virtual Visit Type',
            'software': 'Scheduling Software',
            'practice_facing_name': 'Location Name'
        }
        
        # Read Location Cloud ID values from Location sheet
        max_row = location_sheet.max_row
        location_cloud_ids = []
        
        for row_idx in range(2, max_row + 1):  # Start from row 2 (skip header)
            cell = location_sheet[f"{location_cloud_id_column_letter}{row_idx}"]
            location_cloud_id_value = cell.value
            if location_cloud_id_value:
                location_cloud_id_str = str(location_cloud_id_value).strip()
                location_cloud_ids.append((row_idx, location_cloud_id_str))
        
        # Get ZIP Code column letter for special handling
        zip_code_column_letter = target_columns.get('ZIP Code')
        
        # Write location details to Location sheet
        for row_idx, location_cloud_id_str in location_cloud_ids:
            if location_cloud_id_str in location_id_to_details:
                details = location_id_to_details[location_cloud_id_str]
                
                # Write each detail to its corresponding column
                for source_key, target_column_name in column_mapping.items():
                    target_column_letter = target_columns.get(target_column_name)
                    if target_column_letter and source_key in details:
                        value = details[source_key]
                        # For practice_facing_name, only write if NOT empty
                        if source_key == 'practice_facing_name':
                            if value and str(value).strip() != '':
                                cell = location_sheet[f"{target_column_letter}{row_idx}"]
                                cell.value = str(value).strip()
                        elif value:
                            cell = location_sheet[f"{target_column_letter}{row_idx}"]
                            
                            # Special handling for ZIP Code to preserve leading zeros
                            if source_key == 'zip' and target_column_name == 'ZIP Code':
                                # Set as text format to preserve leading zeros
                                cell.number_format = '@'  # Text format
                                # Ensure value is written as string with leading zeros preserved
                                cell.value = str(value)
                            else:
                                cell.value = value
        
        # Color the header cells green
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        
        for target_column_name, target_column_letter in target_columns.items():
            if target_column_letter:
                header = location_sheet[f"{target_column_letter}{header_row}"]
                header.fill = green_fill
        
        # Save the workbook
        wb.save(template_file)
        return True
        
    except Exception as e:
        return False

if __name__ == "__main__":
    extract_location_details_to_template()

