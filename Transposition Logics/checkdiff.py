"""
CheckDiff - Compare data between Template copy.xlsx and NPI-Extracts.xlsx

This script compares data from the Provider tab of Template copy.xlsx with
the NPI Extracts sheet of NPI-Extracts.xlsx and highlights differences in yellow.
"""

import os
import re
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Get the project root directory (parent of Transposition Logics folder)
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
BACKEND_DIR = PROJECT_ROOT / 'backend'
EXCEL_FILES_DIR = BACKEND_DIR / 'Excel Files'


def normalize_value(value):
    """
    Normalize a value for comparison (handle None, NaN, strings, numbers)
    
    Args:
        value: The value to normalize
        
    Returns:
        str: Normalized string value or empty string
    """
    if value is None or pd.isna(value):
        return ""
    
    # Convert to string and strip whitespace
    normalized = str(value).strip()
    
    # Handle float values (e.g., 1234567890.0 -> 1234567890)
    if isinstance(value, float) and value.is_integer():
        normalized = str(int(value))
    
    # Remove .0 suffix if present
    if normalized.endswith('.0'):
        normalized = normalized[:-2]
    
    return normalized


def compare_values(val1, val2):
    """
    Compare two values (case-insensitive, whitespace-insensitive)
    
    Args:
        val1: First value to compare
        val2: Second value to compare
        
    Returns:
        bool: True if values are different, False if they match
    """
    norm1 = normalize_value(val1).lower()
    norm2 = normalize_value(val2).lower()
    return norm1 != norm2


def has_invalid_symbols(text):
    """
    Check if text contains symbols other than allowed ones (',', '.', '&')
    
    Args:
        text: The text to check
        
    Returns:
        bool: True if invalid symbols are found, False otherwise
    """
    if text is None or pd.isna(text):
        return False
    
    text_str = str(text)
    
    # Allowed characters: letters, numbers, spaces, and the symbols: ',', '.', '&'
    # Create a regex pattern that matches anything NOT in the allowed set
    # Allowed: a-z, A-Z, 0-9, space, comma, period, ampersand
    allowed_pattern = r'^[a-zA-Z0-9\s,\.&]+$'
    
    # If the text doesn't match the allowed pattern, it contains invalid symbols
    return not bool(re.match(allowed_pattern, text_str))


def check_differences():
    """
    Compare data between Template copy.xlsx (Provider tab) and NPI-Extracts.xlsx (NPI Extracts sheet)
    and highlight differences in yellow (#FFFF00)
    
    Column comparisons:
    - First Name vs FIRST_NAME
    - Last Name vs LAST_NAME
    - Gender vs GENDER
    - Professional Suffix 1 vs Suffix Derived
    
    Also highlights Professional Statement cells if:
    - Cell has more than 2000 characters, OR
    - Cell contains symbols except ',', '.', '&'
    
    Also highlights Headshot Link cells in red (#FFB3B3) if they are empty.
    Also highlights Location ID 1 cells in red (#FFB3B3) if they are empty.
    
    Returns:
        bool: True if successful, False otherwise
    """
    # File paths
    template_file = EXCEL_FILES_DIR / 'Template copy.xlsx'
    npi_extracts_file = EXCEL_FILES_DIR / 'NPI-Extracts.xlsx'
    
    # Check if files exist
    if not template_file.exists():
        print(f"Error: Template copy.xlsx not found at {template_file}")
        return False
    
    if not npi_extracts_file.exists():
        print(f"Error: NPI-Extracts.xlsx not found at {npi_extracts_file}")
        return False
    
    try:
        # Load the template workbook
        wb = load_workbook(template_file)
        
        # Check if 'Provider' sheet exists
        if 'Provider' not in wb.sheetnames:
            print("Error: 'Provider' sheet not found in Template copy.xlsx")
            return False
        
        # Get the Provider sheet
        provider_sheet = wb['Provider']
        
        # Find column headers in Provider sheet
        header_row = 1
        npi_column_letter = None
        first_name_column_letter = None
        last_name_column_letter = None
        gender_column_letter = None
        professional_suffix_1_column_letter = None
        professional_statement_column_letter = None
        headshot_link_column_letter = None
        location_id_1_column_letter = None
        
        # Search for the headers in the first row
        for col_idx, cell in enumerate(provider_sheet[header_row], start=1):
            if cell.value:
                cell_value = str(cell.value).strip().lower()
                if cell_value == 'npi number':
                    npi_column_letter = get_column_letter(col_idx)
                elif cell_value == 'first name':
                    first_name_column_letter = get_column_letter(col_idx)
                elif cell_value == 'last name':
                    last_name_column_letter = get_column_letter(col_idx)
                elif cell_value == 'gender':
                    gender_column_letter = get_column_letter(col_idx)
                elif cell_value == 'professional suffix 1':
                    professional_suffix_1_column_letter = get_column_letter(col_idx)
                elif cell_value == 'professional statement':
                    professional_statement_column_letter = get_column_letter(col_idx)
                elif cell_value == 'headshot link':
                    headshot_link_column_letter = get_column_letter(col_idx)
                elif cell_value == 'location id 1':
                    location_id_1_column_letter = get_column_letter(col_idx)
        
        # Check if all required columns are found
        if npi_column_letter is None:
            print("Error: 'NPI Number' column not found in Provider sheet")
            return False
        
        if first_name_column_letter is None:
            print("Warning: 'First Name' column not found in Provider sheet")
        
        if last_name_column_letter is None:
            print("Warning: 'Last Name' column not found in Provider sheet")
        
        if gender_column_letter is None:
            print("Warning: 'Gender' column not found in Provider sheet")
        
        if professional_suffix_1_column_letter is None:
            print("Warning: 'Professional Suffix 1' column not found in Provider sheet")
        
        if professional_statement_column_letter is None:
            print("Warning: 'Professional Statement' column not found in Provider sheet")
        
        if headshot_link_column_letter is None:
            print("Warning: 'Headshot Link' column not found in Provider sheet")
        
        if location_id_1_column_letter is None:
            print("Warning: 'Location ID 1' column not found in Provider sheet")
        
        # Read NPI-Extracts.xlsx
        try:
            npi_extracts_df = pd.read_excel(npi_extracts_file, sheet_name='NPI Extracts')
        except Exception as e:
            print(f"Error reading NPI-Extracts.xlsx: {str(e)}")
            return False
        
        # Check if required columns exist in NPI Extracts
        required_npi_columns = ['NPI Number']
        missing_columns = [col for col in required_npi_columns if col not in npi_extracts_df.columns]
        if missing_columns:
            print(f"Error: Required columns not found in NPI Extracts sheet: {', '.join(missing_columns)}")
            return False
        
        # Create a lookup dictionary: NPI Number -> row data
        npi_lookup = {}
        for idx, row in npi_extracts_df.iterrows():
            npi_value = row['NPI Number']
            npi_key = normalize_value(npi_value)
            
            if npi_key:  # Only add if NPI is not empty
                npi_lookup[npi_key] = {
                    'FIRST_NAME': row.get('FIRST_NAME', None),
                    'LAST_NAME': row.get('LAST_NAME', None),
                    'GENDER': row.get('GENDER', None),
                    'Suffix Derived': row.get('Suffix Derived', None)
                }
        
        # Yellow fill for differences
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        # Red fill for empty headshot links
        red_fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
        
        # Process each row in the Provider sheet
        max_row = provider_sheet.max_row
        differences_found = 0
        
        for row_idx in range(2, max_row + 1):  # Start from row 2 (skip header)
            # Get NPI Number from this row
            npi_cell = provider_sheet[f"{npi_column_letter}{row_idx}"]
            npi_value = npi_cell.value
            npi_key = normalize_value(npi_value)
            
            if not npi_key:
                continue  # Skip rows without NPI
            
            # Look up the NPI in NPI Extracts
            if npi_key not in npi_lookup:
                continue  # Skip if NPI not found in NPI Extracts
            
            npi_extract_data = npi_lookup[npi_key]
            
            # Compare First Name
            if first_name_column_letter:
                first_name_cell = provider_sheet[f"{first_name_column_letter}{row_idx}"]
                template_first_name = first_name_cell.value
                extract_first_name = npi_extract_data.get('FIRST_NAME', None)
                
                if compare_values(template_first_name, extract_first_name):
                    first_name_cell.fill = yellow_fill
                    differences_found += 1
            
            # Compare Last Name
            if last_name_column_letter:
                last_name_cell = provider_sheet[f"{last_name_column_letter}{row_idx}"]
                template_last_name = last_name_cell.value
                extract_last_name = npi_extract_data.get('LAST_NAME', None)
                
                if compare_values(template_last_name, extract_last_name):
                    last_name_cell.fill = yellow_fill
                    differences_found += 1
            
            # Compare Gender
            if gender_column_letter:
                gender_cell = provider_sheet[f"{gender_column_letter}{row_idx}"]
                template_gender = gender_cell.value
                extract_gender = npi_extract_data.get('GENDER', None)
                
                if compare_values(template_gender, extract_gender):
                    gender_cell.fill = yellow_fill
                    differences_found += 1
            
            # Compare Professional Suffix 1
            if professional_suffix_1_column_letter:
                suffix_cell = provider_sheet[f"{professional_suffix_1_column_letter}{row_idx}"]
                template_suffix = suffix_cell.value
                extract_suffix = npi_extract_data.get('Suffix Derived', None)
                
                if compare_values(template_suffix, extract_suffix):
                    suffix_cell.fill = yellow_fill
                    differences_found += 1
        
        # Check Professional Statement column for issues
        if professional_statement_column_letter:
            for row_idx in range(2, max_row + 1):  # Start from row 2 (skip header)
                statement_cell = provider_sheet[f"{professional_statement_column_letter}{row_idx}"]
                statement_value = statement_cell.value
                
                if statement_value is not None and not pd.isna(statement_value):
                    statement_str = str(statement_value)
                    
                    # Check if length exceeds 2000 characters
                    if len(statement_str) > 2000:
                        statement_cell.fill = yellow_fill
                        differences_found += 1
                    # Check if contains invalid symbols
                    elif has_invalid_symbols(statement_value):
                        statement_cell.fill = yellow_fill
                        differences_found += 1
        
        # Check Headshot Link column for empty cells
        empty_headshots = 0
        if headshot_link_column_letter:
            for row_idx in range(2, max_row + 1):  # Start from row 2 (skip header)
                headshot_cell = provider_sheet[f"{headshot_link_column_letter}{row_idx}"]
                headshot_value = headshot_cell.value
                
                # Check if cell is empty (None, NaN, or empty string)
                is_empty = False
                if headshot_value is None or pd.isna(headshot_value):
                    is_empty = True
                elif isinstance(headshot_value, str) and not headshot_value.strip():
                    is_empty = True
                
                if is_empty:
                    headshot_cell.fill = red_fill
                    empty_headshots += 1
        
        # Check Location ID 1 column for empty cells
        empty_location_ids = 0
        if location_id_1_column_letter:
            for row_idx in range(2, max_row + 1):  # Start from row 2 (skip header)
                location_id_cell = provider_sheet[f"{location_id_1_column_letter}{row_idx}"]
                location_id_value = location_id_cell.value
                
                # Check if cell is empty (None, NaN, or empty string)
                is_empty = False
                if location_id_value is None or pd.isna(location_id_value):
                    is_empty = True
                elif isinstance(location_id_value, str) and not location_id_value.strip():
                    is_empty = True
                
                if is_empty:
                    location_id_cell.fill = red_fill
                    empty_location_ids += 1
        
        # Save the workbook
        wb.save(template_file)
        return True
        
    except Exception as e:
        print(f"Error in check_differences: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    check_differences()

